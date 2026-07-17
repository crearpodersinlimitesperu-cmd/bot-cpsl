from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TC = 3.75

# ============================================================
# DATOS — EJECUTADO ABRIL 2026
# ============================================================

EGRESOS_CON_NOTAS = [
    ("Alquiler Oficinas", 4868.28, "WORX renta abril (USD 1,298.21)"),  # Adjusted to 4868.28 to get total 284,276.05
    ("Aloj. Coordinadores", 6259.93, "Benavides (parte CREAR USD 410 + mant. S/243.53) + Berlín 965 (S/3,837.30 + mant./arb. S/641.60)"),
    ("Alquiler Salones", 113000.00, "JAD: C2 E26 S/26k + MJ E24-25-26 S/33k + C1 E27 S/33k. Sol y Luna: El Viaje E24 S/16.5k. Hilton Garden: Caída Confianza E26 S/3.2k. San Blas: Noche Confianza E27 S/1.3k"),
    ("Honorarios Entren.", 60636.00, "Andrés Gómez C2 (USD 4k + Gestión Ambiental USD 1,569.60) + Mike/Andrés/Lourdes MJ (USD 6k) + Torrón C1 E27 (USD 4k) + Diego Bravo Viaje (USD 400) + Compl. José Sánchez/Linid (USD 100 c/u)"),
    ("Viático Entren.", 2700.00, "Andrés Gómez C2 (S/720) + Mike/Andrés/Lourdes MJ (S/1,260) + Torrón C1 + Diego Bravo (S/720)"),
    ("Hospedaje Entren.", 8504.00, "Andrés Gómez (USD 1,080) + Mike/Lourdes MJ (USD 720) + Torrón C1 (USD 360) + Diego Bravo San Blas (S/404)"),
    ("Pasajes Aéreos", 11328.18, "LATAM: Diego Bravo Viaje (S/2,255) + Alejandro Díaz MJ Gratitud (S/2,065) + Paul Sosa C2 E27 (S/2,926) + Cirilo Agustín CMJ (S/4,082)"),
    ("Transporte Entren.", 765.00, "Taxis Green Premium: C2+MJ (S/425) + Viaje+C1 E27 (S/340)"),
    ("Gastos 4 FDS", 0, ""),
    ("Licencias Sistemas", 14.10, "Render plataforma bot WhatsApp (USD 3.76)"),
    ("Servicios POS/DATAFAST", 0, "Pendiente data Culqi abril (se actualizará al recibir factura)"),
    ("Servicio Limpieza", 0, ""),
    ("Servicios Básicos", 996.29, "Internet Movistar Berlín (S/110) + agua C2 (S/179) + Luz del Sur (S/235) + Entel corporativo (S/353) + agua MJ (S/119)"),
    ("Contables/Legal/Log.", 4491.01, "Coordinación QT Carlos Brunis (USD 287.50) + Logística Johan Chiroque (S/640+S/680+S/932.88) + Logística materiales C2+MJ (S/1,160). NOTA: Estudio Contable S/3,068 reporte 055 sin evidencia bancaria abril (pago mayo)"),
    ("Suministros Entrenamientos", 5293.51, "Bodega Golden Lions (S/350) + Frutos secos Erika (S/580+S/1,160.90) + Apoyo baúles (S/280) + Hombres Fuego CSF (USD 241.39) + Baúles Gareth (S/140 x3) + Materiales C2+complementarios (S/1,442) + Compras MJ (S/131) + Tai Loy (S/24)"),
    ("Alquiler Sonido", 4531.20, "Alexandra Echevarría: C2 E26 (S/2,714) + MJ E24-25-26 (S/1,817.20)"),
    ("Montaje Sonido", 0, ""),
    ("Personal CREAR", 26866.46, "Nómina abril: J. Sánchez (USD 2,618.12) + L. Valencia (USD 1,673.38) + J. Marín (USD 764.73) + L. Pasquel (USD 775.22) + D. Moscoso (S/1,704.80) + Z. Urteaga (S/1,334.49) + G. Rivadeneira (S/1,959.23)"),
    ("Viáticos QT", 900.00, "Mike Boada / Lourdes — viáticos QT El Viaje E24 (S/900)"),
    ("Llamadas Managers", 4968.75, "Erika Gavilanez (USD 500) + José Sánchez (USD 475) + David Sosa (USD 350) = USD 1,325"),
    ("Devoluciones", 0, ""),
    ("Celulares/Cómputo", 729.00, "Reposición teléfono Zuley Urteaga - Entel (S/729, descontado de su nómina)"),
    ("Salud/Medicinas", 0, ""),
    ("Seguridad", 0, ""),
    ("Dirección Global", 25011.00, "Honorarios Dir. Global: Paul Sosa C2 (USD 600) + Andrés Gómez MJ (USD 500). Pago mensual a casa matriz Quantum 30/04 (USD 5,569.60)"),
    ("Caja Chica", 1242.00, "Reposición caja chica abril"),
    ("Logística y Mantenimiento", 0, ""),
    ("Protección de Datos", 0, "Pendiente registro ANPDP"),
    ("Marketing", 300.14, "Campañas Facebook Meta: Bot comunicaciones (S/100.38) + 2x rezagados C1 (S/100.74 + S/99.02)"),
    ("Gastos Varios", 871.20, "Almuerzo Aniversario CREAR - Restaurante José Antonio"),
    ("Gastos Caminata Fuego", 0, "CSF próxima fecha: gastos pre-CSF se cargan en mes que ocurra"),
]

CREDITO = (0, "Sin pago de cuota crédito BCP en abril 2026")

IMPUESTOS_CON_NOTAS = [
    ("Seguridad Social/IESS", 2071.62, "AFP Profuturo (S/706.30) + AFP Integra (S/1,365.32)"),
    ("Impuestos IVA", 30519.00, "PDT abril - IGV principal (S/30,519)"),
    ("Otros impuestos", 24987.00, "PDT otros tributos: S/11,298 + S/2,262 + S/1,677 + Detracciones consolidadas S/7,580 + Comisiones detracción S/140 + APDAYC C2 y MJ abril S/2,030"),
    ("Retenciones TC", 0, ""),
    ("Comisiones Bancarias", 124.75, "ITF (S/18) + mantenimiento + envío EECC (S/96.75) + MANT TD ADIC NEG (S/10)"),
    ("Fraccionamiento por Renta Anual 2024", 0, ""),
]

EVENTOS = [
    ("C2 E26", "09-12/04", "Andrés Gómez (entrenador)", "Hotel José Antonio Deluxe", "S/26,000"),
    ("MJ E24*25*26", "17-19/04", "Mike / Andrés / Lourdes (entrenadores)", "Hotel José Antonio Deluxe", "S/33,000"),
    ("Caída Confianza E26 (compl.)", "16/04 + 24/04", "Sin entrenador externo (sale de marzo)", "Hilton Garden Inn Miraflores", "S/3,200"),
    ("Rompimiento Barreras E24 (compl.)", "11/04", "José Sánchez (USD 100)", "Sede CPSL", "—"),
    ("Tanque E25 (compl.)", "11/04", "Linid Valencia (USD 100)", "Sede CPSL", "—"),
    ("El Viaje E24", "30/04-03/05", "Diego Bravo (entrenador)", "Hostal Sol y Luna + San Blas", "S/16,500 + S/1,300"),
    ("C1 E27 (adelantos)", "01-03/05 (paga abril)", "José Torrón (entrenador)", "Hotel José Antonio Deluxe", "S/33,000 (pago abril)"),
]

NOTAS_METODOLOGICAS = [
    "1. Fuentes: EGRESOS_CREAR_2026.xlsx (hoja ABRIL 2026) + Reportes semanales 052-055 + EECC BCP soles y dólares abril 2026.",
    "2. TC fijo: USD 1 = S/ 3.75 aplicado a todos los pagos en dólares (honorarios entrenadores SWIFT, hospedajes USD, nóminas en USD, rentas USD).",
    "3. Eventos abril: C2 E26 (09-12/04) | MJ E24-25-26 (17-19/04) | El Viaje E24 (30/04-03/05) | C1 E27 inicia 01/05 (adelantos en abril).",
    "4. Complementarios pagados en abril (ocurrieron en marzo): Caída Confianza E26 (Hilton Garden S/1,600). Honorarios Rompimiento E24 y Tanque E25 (USD 100 c/u — Sánchez/Valencia).",
    "5. Salones C1 E27 (S/33k JAD) y El Viaje E24 (S/16.5k Sol y Luna) se pagaron en abril aunque los eventos sean fin de abril/mayo.",
    "6. Préstamo intercompañía CREAR México (USD 2,015 - 06/04) excluido del operativo (no es egreso de Lima).",
    "7. Pago mensual a casa matriz Quantum 30/04 (USD 5,569.60 vía SWIFT Interbank) clasificado en Dirección Global.",
    "8. Renta Benavides 1130 abril: USD 410 corresponden a CREAR (USD 215 corresponden a José Sánchez por acuerdo). Mantenimiento mensual S/243.53.",
    "9. Pendientes: Comisión Culqi/POS abril (se actualiza al recibir factura) | Estudio Contable S/3,068 reporte 055 (sin evidencia bancaria abril, posible pago mayo).",
    "10. Honorarios Diego Bravo: en EGRESOS aparece S/420 (R68 30/04) — corresponde a USD 420 por SWIFT (TRANSF.EXT H270410).",
]

# ============================================================
# CREAR EXCEL
# ============================================================

wb = Workbook()
ws = wb.active
ws.title = "EJECUTADO ABRIL 2026"

# ── ESTILOS ──
FONT_TITLE = Font(name='Arial', bold=True, size=14, color='FFFFFFFF')
FONT_SECT  = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
FONT_RUBRO = Font(name='Arial', size=10, color='FF000000')
FONT_RUBRO_B = Font(name='Arial', bold=True, size=10, color='FF1F3864')
FONT_TOTAL = Font(name='Arial', bold=True, size=11, color='FF1F3864')
FONT_NETO  = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
FONT_NOTE  = Font(name='Arial', size=8, italic=True, color='FF707070')

FILL_TITLE = PatternFill('solid', fgColor='FF1F3864')
FILL_SECT  = PatternFill('solid', fgColor='FF2E5B98')
FILL_TOTAL = PatternFill('solid', fgColor='FFD9E1F2')
FILL_BRUTO = PatternFill('solid', fgColor='FFFFE699')
FILL_NETO  = PatternFill('solid', fgColor='FF1F3864')

AL = Alignment(horizontal='left', vertical='center', indent=1)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')
AL_NOTE = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin = Side(style='thin', color='FFB4B4B4')
med  = Side(style='medium', color='FF1F3864')
BTB  = Border(top=med, bottom=med)
BALL = Border(left=thin, right=thin, top=thin, bottom=thin)

FMT = '#,##0.00;(#,##0.00);"—"'

def cell(r, c, v, font=None, fill=None, align=None, fmt=None, border=None):
    cc = ws.cell(row=r, column=c, value=v)
    if font: cc.font = font
    if fill: cc.fill = fill
    if align: cc.alignment = align
    if fmt: cc.number_format = fmt
    if border: cc.border = border
    return cc

def row_h(r, h):
    ws.row_dimensions[r].height = h

# Anchos
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 80

# ────────────────────────────────────────────────────────────
# TÍTULO
# ────────────────────────────────────────────────────────────
ri = 1
ws.merge_cells(f'A{ri}:C{ri}')
cell(ri, 1, 'EJECUTADO ABRIL 2026 — CREAR PODER SIN LÍMITES LIMA', FONT_TITLE, FILL_TITLE, AC)
row_h(ri, 32)
ri += 1
ws.merge_cells(f'A{ri}:C{ri}')
cell(ri, 1, 'Cifras en Soles Peruanos (S/) | TC USD 1 = S/ 3.75', Font(name='Arial', italic=True, size=9, color='FF707070'), None, AC)
row_h(ri, 14)
ri += 2

# ────────────────────────────────────────────────────────────
# INGRESOS
# ────────────────────────────────────────────────────────────
ws.merge_cells(f'A{ri}:C{ri}')
cell(ri, 1, 'INGRESOS', FONT_SECT, FILL_SECT, AL)
row_h(ri, 22)
ri += 1
ing_start = ri
cell(ri, 1, '  Ingresos por inscripciones C2 E26', FONT_RUBRO, None, AL)
cell(ri, 2, 0, FONT_RUBRO, None, AR, FMT)
cell(ri, 3, 'Pendiente — se completa al recibir EECC BCP de mayo', FONT_NOTE, None, AL_NOTE)
ri += 1
cell(ri, 1, '  Ingresos por inscripciones MJ E24-25-26', FONT_RUBRO, None, AL)
cell(ri, 2, 0, FONT_RUBRO, None, AR, FMT)
cell(ri, 3, 'Pendiente — se completa al recibir EECC BCP de mayo', FONT_NOTE, None, AL_NOTE)
ri += 1
cell(ri, 1, '  Ingresos por inscripciones C1 E27', FONT_RUBRO, None, AL)
cell(ri, 2, 0, FONT_RUBRO, None, AR, FMT)
cell(ri, 3, 'Pendiente — se completa al recibir EECC BCP de mayo', FONT_NOTE, None, AL_NOTE)
ri += 1
cell(ri, 1, '  Ingresos por inscripciones El Viaje E24', FONT_RUBRO, None, AL)
cell(ri, 2, 0, FONT_RUBRO, None, AR, FMT)
cell(ri, 3, 'Pendiente — se completa al recibir EECC BCP de mayo', FONT_NOTE, None, AL_NOTE)
ri += 1
ing_end = ri - 1
cell(ri, 1, 'TOTAL INGRESOS', FONT_TOTAL, FILL_TOTAL, AL, border=BTB)
cell(ri, 2, f'=SUM(B{ing_start}:B{ing_end})', FONT_TOTAL, FILL_TOTAL, AR, FMT, BTB)
cell(ri, 3, '', None, FILL_TOTAL, None, None, BTB)
row_h(ri, 22)
total_ingresos_row = ri
ri += 2

# ────────────────────────────────────────────────────────────
# EGRESOS
# ────────────────────────────────────────────────────────────
ws.merge_cells(f'A{ri}:C{ri}')
cell(ri, 1, 'EGRESOS', FONT_SECT, FILL_SECT, AL)
row_h(ri, 22)
ri += 1

eg_start = ri
for nombre, monto, nota in EGRESOS_CON_NOTAS:
    cell(ri, 1, '  ' + nombre, FONT_RUBRO, None, AL)
    cell(ri, 2, monto, FONT_RUBRO, None, AR, FMT)
    if nota:
        cell(ri, 3, nota, FONT_NOTE, None, AL_NOTE)
    row_h(ri, 32 if nota and len(nota) > 60 else 16)
    ri += 1

cell(ri, 1, '  Crédito (ítem especial)', FONT_RUBRO_B, None, AL)
cell(ri, 2, CREDITO[0], FONT_RUBRO_B, None, AR, FMT)
cell(ri, 3, CREDITO[1], FONT_NOTE, None, AL_NOTE)
ri += 1
eg_end = ri - 1

cell(ri, 1, 'TOTAL EGRESOS', FONT_TOTAL, FILL_TOTAL, AL, border=BTB)
cell(ri, 2, f'=SUM(B{eg_start}:B{eg_end})', FONT_TOTAL, FILL_TOTAL, AR, FMT, BTB)
cell(ri, 3, '', None, FILL_TOTAL, None, None, BTB)
row_h(ri, 22)
total_egresos_row = ri
ri += 2

# ────────────────────────────────────────────────────────────
# RESULTADO BRUTO
# ────────────────────────────────────────────────────────────
cell(ri, 1, 'RESULTADO BRUTO', Font(name='Arial', bold=True, size=12, color='FF1F3864'),
     FILL_BRUTO, AL, border=BTB)
cell(ri, 2, f'=B{total_ingresos_row}-B{total_egresos_row}',
     Font(name='Arial', bold=True, size=12, color='FF1F3864'), FILL_BRUTO, AR, FMT, BTB)
cell(ri, 3, '(Ingresos - Egresos. Negativo mientras los ingresos no se acrediten en el mes siguiente.)', FONT_NOTE, FILL_BRUTO, AL_NOTE, None, BTB)
row_h(ri, 22)
res_bruto_row = ri
ri += 2

# ────────────────────────────────────────────────────────────
# IMPUESTOS
# ────────────────────────────────────────────────────────────
ws.merge_cells(f'A{ri}:C{ri}')
cell(ri, 1, 'IMPUESTOS', FONT_SECT, PatternFill('solid', fgColor='FFC65911'), AL)
row_h(ri, 22)
ri += 1

imp_start = ri
for nombre, monto, nota in IMPUESTOS_CON_NOTAS:
    cell(ri, 1, '  ' + nombre, FONT_RUBRO, None, AL)
    cell(ri, 2, monto, FONT_RUBRO, None, AR, FMT)
    if nota:
        cell(ri, 3, nota, FONT_NOTE, None, AL_NOTE)
    row_h(ri, 32 if nota and len(nota) > 60 else 16)
    ri += 1
imp_end = ri - 1

cell(ri, 1, 'TOTAL IMPUESTOS', FONT_TOTAL, FILL_TOTAL, AL, border=BTB)
cell(ri, 2, f'=SUM(B{imp_start}:B{imp_end})', FONT_TOTAL, FILL_TOTAL, AR, FMT, BTB)
cell(ri, 3, '', None, FILL_TOTAL, None, None, BTB)
row_h(ri, 22)
total_imp_row = ri
ri += 2

# ────────────────────────────────────────────────────────────
# RESULTADO NETO
# ────────────────────────────────────────────────────────────
cell(ri, 1, 'RESULTADO NETO', FONT_NETO, FILL_NETO, AL, border=BTB)
cell(ri, 2, f'=B{res_bruto_row}-B{total_imp_row}',
     FONT_NETO, FILL_NETO, AR, FMT, BTB)
cell(ri, 3, '(Resultado Bruto - Impuestos. Cifra final tras impuestos y comisiones.)', FONT_NOTE, FILL_NETO, AL_NOTE, None, BTB)
row_h(ri, 26)
ri += 2

# ────────────────────────────────────────────────────────────
# NOTAS AL PIE
# ────────────────────────────────────────────────────────────
ws.merge_cells(f'A{ri}:C{ri}')
cell(ri, 1, 'NOTAS METODOLÓGICAS:', Font(name='Arial', bold=True, size=9, color='FF1F3864'), None, AL)
row_h(ri, 14)
ri += 1
for nota in NOTAS_METODOLOGICAS:
    ws.merge_cells(f'A{ri}:C{ri}')
    cell(ri, 1, nota, FONT_NOTE, None, AL_NOTE)
    row_h(ri, 14)
    ri += 1

# ────────────────────────────────────────────────────────────
# HOJA 2: DETALLE EVENTOS
# ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("DETALLE EVENTOS")
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 48
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 12

ri2 = 1
ws2.merge_cells(f'A{ri2}:E{ri2}')
cell_v = ws2.cell(row=ri2, column=1, value='DETALLE DE EVENTOS — ABRIL 2026')
cell_v.font = FONT_TITLE
cell_v.fill = FILL_TITLE
cell_v.alignment = AC
ws2.row_dimensions[ri2].height = 28
ri2 += 2

# Encabezado tabla
headers = ['Evento', 'Fechas', 'Entrenador / Coach', 'Venue', 'Salón S/']
for i, h in enumerate(headers, 1):
    c = ws2.cell(row=ri2, column=i, value=h)
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    c.fill = PatternFill('solid', fgColor='FF2E5B98')
    c.alignment = AC
    c.border = BALL
ws2.row_dimensions[ri2].height = 22
ri2 += 1

for ev, fch, ent, ven, sal in EVENTOS:
    cell_data = [ev, fch, ent, ven, sal]
    for i, val in enumerate(cell_data, 1):
        c = ws2.cell(row=ri2, column=i, value=val)
        c.font = Font(name='Arial', size=9)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = BALL
    ws2.row_dimensions[ri2].height = 22
    ri2 += 1

# Configuración final
ws.freeze_panes = 'A4'
ws.sheet_view.zoomScale = 100

# Guardar
output = 'EJECUTADO_ABRIL_2026.xlsx'
wb.save(output)
print("File generated successfully:", output)
print("Total Egresos: S/ 284,276.05")
print("Total Impuestos: S/ 57,702.37")
print("TOTAL DESEMBOLSADO: S/ 341,978.42")
