import os
import json
import logging
import re
import openpyxl
from datetime import datetime, timedelta
import requests

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
log = logging.getLogger("Sync2027")

# Configurar API URL 
API_URL = "https://script.google.com/macros/s/AKfycbySw3nJ0gmOqPtLURGrJeH7ja51MbkLjEDO2exqZTUAzW3-p35s4cU7uKSUUz4fEhGD/exec"

# Rutas a los Excels.
EXCEL_MAPEO = r"G:\Unidades compartidas\LIMA\PROGRAMACION 2026 CREAR LIMA.xlsx"



def procesar_sede(wb, sheet_name, prefix_sede, updates_list):
    log.info(f"Procesando sede {prefix_sede} en pestaña {sheet_name}...")
    if sheet_name not in wb.sheetnames:
        log.warning(f"Pestaña {sheet_name} no encontrada.")
        return
    
    eventos_agregados = 0
    ws = wb[sheet_name]
    for r in range(2, ws.max_row + 1):
        inicio = ws.cell(row=r, column=1).value
        fin = ws.cell(row=r, column=2).value
        equipo = ws.cell(row=r, column=4).value
            
        # Saltamos si es un año (entero) o no son fechas
        if isinstance(inicio, int) or not isinstance(inicio, datetime) or not isinstance(fin, datetime):
            continue
            
        nombre = ws.cell(row=r, column=3).value
        trainer = ws.cell(row=r, column=5).value
        
        if not nombre or "REUNIÓN" in str(nombre).upper():
            continue
            
        # Forzar a que recoja 2026 y 2027
        if inicio.year not in [2026, 2027]:
            continue
            
        # Extraer equipo resolviendo fórmulas simples si `data_only=True` falla
        equipo_val = ws.cell(row=r, column=4).value
        
        def resolve_equipo(val, current_ws, depth=0):
            if depth > 5 or val is None: return ""
            if isinstance(val, (int, float)): return str(int(val))
            val_str = str(val).strip()
            # Si es una fórmula como =+D285+1 o =D36+1
            import re
            match = re.match(r'^=?[+]?D(\d+)(?:\+(\d+))?$', val_str.replace(' ', ''))
            if match:
                target_row = int(match.group(1))
                add_val = int(match.group(2)) if match.group(2) else 0
                target_val = current_ws.cell(row=target_row, column=4).value
                base_str = resolve_equipo(target_val, current_ws, depth+1)
                if base_str.isdigit():
                    return str(int(base_str) + add_val)
            return val_str

        equipo_str = resolve_equipo(equipo_val, ws)
        equipo_str = equipo_str.replace('*', ' ').replace('+', ' ').strip()
        
        # ID Único
        evento_id = f"EVENTO_{prefix_sede}_{equipo_str}_{inicio.strftime('%Y%m%d')}".replace(" ", "_")
        
        data = {
            "id": evento_id,
            "data": {
                "sede": prefix_sede,
                "fecha_inicio": inicio.strftime('%Y-%m-%dT00:00:00Z'),
                "fecha_fin": fin.strftime('%Y-%m-%dT00:00:00Z'),
                "nombre": str(nombre),
                "equipo": equipo_str,
                "trainer": str(trainer) if trainer else "Por confirmar",
                "lugar": "CREAR PODER SIN LÍMITES",
                "direccion": "Por definir",
                "force_update": int(datetime.now().timestamp() * 1000)
            },
            "ts": int(datetime.now().timestamp() * 1000)
        }
        updates_list.append(data)
        eventos_agregados += 1
        
    log.info(f"-> {eventos_agregados} eventos extraídos para {prefix_sede}.")

def push_to_cloud(updates_list):
    if not updates_list:
        log.info("No hay actualizaciones para enviar.")
        return
        
    chunk_size = 50
    for i in range(0, len(updates_list), chunk_size):
        chunk = updates_list[i:i + chunk_size]
        payload = {
            "action": "batchUpdate",
            "updates": chunk
        }
        log.info(f"Enviando lote {i//chunk_size + 1} ({len(chunk)} registros) a la nube...")
        try:
            response = requests.post(API_URL, json=payload)
            response.raise_for_status()
            log.info(f"Respuesta del lote {i//chunk_size + 1}: {response.text[:100]}...")
        except Exception as e:
            log.error(f"Error subiendo lote {i//chunk_size + 1}: {e}")

def main():
    if not os.path.exists(EXCEL_MAPEO):
        log.error(f"No se encontró el excel maestro: {EXCEL_MAPEO}")
        return
        
    log.info("Abriendo Excel (puede tomar unos segundos)...")
    wb = openpyxl.load_workbook(EXCEL_MAPEO, data_only=True)
    
    updates = []
    
    sedes = {
        "LIM": "LIM",
        "UIO C1": "UIO C1",
        "UIO C2": "UIO C2",
        "GYE": "GYE",
        "CUE": "CUE",
        "MED": "MED",
        "CDMX": "MEX"
    }
    
    for tab_name, sede_name in sedes.items():
        procesar_sede(wb, tab_name, sede_name, updates)
        
    wb.close()
    
    push_to_cloud(updates)
    log.info("Sincronización finalizada.")

if __name__ == "__main__":
    main()
