"""
Bot WhatsApp — Crear Poder Sin Límites Perú
V112-FIX: Corrección de sintaxis + dependencias
✅ LISTO PARA COPIAR Y PEGAR EN RENDER
"""
import os, re, json, time, csv, base64, random, logging, threading, queue, smtplib
from flask import Flask, request, jsonify
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import requests as req_lib
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from dotenv import load_dotenv
load_dotenv()

# Imports opcionales con manejo seguro
try:
    from ia_chain import ia_detect_intent_cc, buscar_caso_por_nombre
except ImportError:
    ia_detect_intent_cc = None
    buscar_caso_por_nombre = None

try:
    from ia_multimodelo import (ia_clasificar, ia_respuesta_px, ia_respuesta_imo, 
                                ia_respuesta_nuevo, guardar_feedback, estado_ias)
except ImportError:
    ia_clasificar = None
    ia_respuesta_px = None
    ia_respuesta_imo = None
    ia_respuesta_nuevo = None
    guardar_feedback = None
    estado_ias = None

try:
    from crm_bridge import (push_reporte_crm, push_gestion_individual, 
                            push_reporte_jose, kpi_consolidado_whatsapp)
except ImportError:
    push_reporte_crm = None
    push_gestion_individual = None
    push_reporte_jose = None
    kpi_consolidado_whatsapp = None

from filelock import FileLock, Timeout as FileLockTimeout

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("CPSL")

# Configurar logging a archivo para debug remoto (UTF-8 para evitar errores con emojis)
log_path = os.path.join("/data" if os.path.exists("/data") else os.path.dirname(os.path.abspath(__file__)), "bot.log")
try:
    fh = logging.FileHandler(log_path, encoding='utf-8')
    fh.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
    logger.addHandler(fh)
except Exception as e:
    print(f"Error configurando FileHandler: {e}")

logger.setLevel(logging.INFO)

logger.info("Bot iniciado - Logger configurado")
app = Flask(__name__)

# ── ZONA HORARIA ─────────────────────────────────────────────
TZ_LIMA = timezone(timedelta(hours=-5))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = "/data" if os.path.exists("/data") else BASE_DIR

# ==========================================
# CALENDARIO DE ENTRENAMIENTOS LIMA
# ==========================================
def _en_entrenamiento():
    """Verifica si estamos en fechas de entrenamiento según el json del calendario."""
    try:
        import os, json
        from datetime import datetime, timezone, timedelta
        tz = timezone(timedelta(hours=-5))
        hoy = datetime.now(tz).strftime("%Y-%m-%d")
        cal_path = "/data/calendario_entrenamientos.json" if os.path.exists("/data") else "calendario_entrenamientos.json"
        if os.path.exists(cal_path):
            with open(cal_path, "r", encoding="utf-8") as f:
                eventos = json.load(f)
                for ev in eventos:
                    if ev["inicio"] <= hoy <= ev["fin"]:
                        return ev["nombre"]
    except Exception as e:
        pass
    return None

def ahora(): return datetime.now(TZ_LIMA)

# ── STAFF (Diana Moscoso y Joyce Marín — Campaña C1 E28) ──────────
STAFF = {
    "dmoscoso": {"nombre": "Diana Moscoso", "tel": "51912379744"},
    "jmarin": {"nombre": "Joyce Marín", "tel": "51933599903"},
    "gerencia": {"nombre": "José (Gerencia)", "tel": "573116024515"},
}
MANAGER_TELS = {"51919563284", "573116024515"}
_carga = {k: 0 for k in STAFF}
_carga_lk = threading.Lock()

def cc_libre():
    with _carga_lk:
        return min(_carga, key=_carga.get)

def cc_add(k):
    with _carga_lk:
        if k in _carga: _carga[k] += 1

# Mapa equipo → coordinadora para DERIVACIONES (Actualizado para Diana/Joyce)
_CC_POR_EQUIPO = {
    "EQUIPO 27": "dmoscoso", "EQUIPO 26": "dmoscoso", "EQUIPO 25": "jmarin", 
    "EQUIPO 24": "dmoscoso", "EQUIPO 23": "jmarin", "EQUIPO 22": "jmarin", 
    "EQUIPO 21": "jmarin", "EQUIPO 20": "jmarin", "EQUIPO 19": "dmoscoso", 
    "EQUIPO 18": "dmoscoso", "EQUIPO 17": "dmoscoso", "EQUIPO 16": "dmoscoso"
}

def cc_por_equipo(equipo):
    """Retorna la key del staff asignado al equipo. Fallback: cc_libre()."""
    return _CC_POR_EQUIPO.get(str(equipo).strip().upper(), cc_libre())

# ── CONFIG ────────────────────────────────────────────────────
class Cfg:
    CAMPANA_ACTUAL = 'C1 E28'
    EQUIPO_ACTUAL = 'Equipo 28'
    TOKEN = os.environ.get("WA_TOKEN","").strip()
    PHONE_ID = os.environ.get("WA_PHONE_ID","").strip()
    VER_TOKEN = os.environ.get("WA_VERIFY_TOKEN","cpsl2026")
    # Unificación de SHEET_ID con fallback al ID maestro
    SHEET_ID = os.environ.get("SHEET_ID", os.environ.get("CRM_SHEET_ID", "1IoCYs1qfOTdn3XWyeK64jsUfAXOFgv3Wa6uJBM-lR2Y")).strip()
    CREDS = os.environ.get("GOOGLE_CREDENTIALS","").strip()
    SHEDS = os.environ.get("GOOGLE_CREDENTIALS","").strip() # Alias para evitar errores de atributo
    SHEET_TAB = os.environ.get("SHEET_TAB","LOG_INTERACCIONES")
    GMAIL_USER = "crearpodersinlimitesperu@gmail.com"
    GMAIL_PASS = os.environ.get("GMAIL_APP_PASS", "bgsl xjus xsmn pzqd")
    LOCK_T = 5
    CSV = os.path.join(BASE_DIR, "Prospectos_Pendientes_C1_Depurado_Campana.csv")
    S_REAL = os.path.join(DATA_DIR, "sesiones.json")
    S_SIM = os.path.join(DATA_DIR, "sesiones_sim.json")
    HIST = os.path.join(DATA_DIR, "historial_chat.json")
    HIST_ALT = os.path.join(DATA_DIR, "historial.json")
    FECHA = "Del 29 al 31 de mayo"
    LUGAR = "Hotel José Antonio Deluxe, Miraflores"
    REGISTRO = "08:00 AM"

# Validación inicial
if not Cfg.TOKEN: logger.critical("❌ ERROR: WA_TOKEN vacío. El bot NO podrá responder.")
if not Cfg.PHONE_ID: logger.critical("❌ ERROR: WA_PHONE_ID vacío.")
if not Cfg.CREDS: logger.warning("⚠️ AVISO: GOOGLE_CREDENTIALS vacío. Sincronización deshabilitada.")

FECHAS_MSG = (
    "📅 *Próximas Fechas — Sede Lima 2026*\n\n"
    "🚀 *C1 Equipo 28:* Próximamente\n"
    "   📍 *Lugar por confirmar*\n\n"
    "🔥 *C2 {Cfg.EQUIPO_ACTUAL}:* Jueves 14 de mayo\n"
)

# ── CACHÉ CSV ─────────────────────────────────────────────────
_csv_cache = None
_csv_mtime = 0.0
_csv_lk = threading.Lock()

def _get_rows():
    global _csv_cache, _csv_mtime
    if not os.path.exists(Cfg.CSV):
        logger.warning(f"CSV no encontrado: {Cfg.CSV}")
        return []
    try:
        mt = os.path.getmtime(Cfg.CSV)
        with _csv_lk:
            if _csv_cache is not None and mt == _csv_mtime:
                return _csv_cache
            with open(Cfg.CSV, encoding="utf-8-sig") as f:
                first = f.readline()
                delim = ";" if first.count(";") > first.count(",") else ","
                f.seek(0)
                rows = list(csv.DictReader(f, delimiter=delim))
            _csv_cache, _csv_mtime = rows, mt
            logger.info(f"CSV cargado: {len(rows)} filas")
            return rows
    except Exception as e:
        logger.error(f"CSV error: {e}")
        return []

# ── ÍNDICE DE GRADUADOS ──────────────────────────────────────
_GRADUADOS_IDX = {}
def _cargar_graduados():
    global _GRADUADOS_IDX
    try:
        import openpyxl as _opx
        _p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GRADUADOS_LIMA.xlsx")
        if not os.path.exists(_p): return
        _wb = _opx.load_workbook(_p, data_only=True, read_only=True)
        if "GRADUADOS" in _wb.sheetnames:
            for _r in _wb["GRADUADOS"].iter_rows(min_row=2, values_only=True):
                if _r[0]:
                    _n = re.sub(r"\s+"," ",re.sub(r"[^\w\s]","",str(_r[0]).upper())).strip()
                    _GRADUADOS_IDX[_n] = {"graduado":True,"rango":"GRADUADO","equipo":str(_r[1] or "")}
        if "ALIADOS C1{Cfg.CAMPANA_ACTUAL.split()[1]}" in _wb.sheetnames:
            for _r in _wb["ALIADOS C1{Cfg.CAMPANA_ACTUAL.split()[1]}"].iter_rows(min_row=3, values_only=True):
                if _r[0] and str(_r[0]).strip() not in ("CREADOR CUANTICO",""):
                    _n = re.sub(r"\s+"," ",re.sub(r"[^\w\s]","",str(_r[0]).upper())).strip()
                    if _n not in _GRADUADOS_IDX:
                        _GRADUADOS_IDX[_n] = {"graduado":True,"rango":"GRADUADO_{Cfg.CAMPANA_ACTUAL.split()[1]}"}
        _wb.close()
        logger.info(f"Graduados cargados: {len(_GRADUADOS_IDX)}")
    except Exception as _e:
        logger.warning(f"graduados: {_e}")
_cargar_graduados()

# ── NORMALIZACIÓN ─────────────────────────────────────────────
def _d(s): return re.sub(r'\D','',str(s or ''))
def n9(t): return _d(t)[-9:]

def formatear_nombre_empatia(texto, solo_nombre=False):
    if not texto: return ""
    texto = str(texto).strip()
    if "," in texto:
        partes = [p.strip() for p in texto.split(",")]
        if len(partes) >= 2:
            nom, ape = partes[1], partes[0]
            if solo_nombre: return nom.split()[0].title()
            return f"{nom.title()} {ape.title()}"
    tokens = [t for t in texto.split() if len(t) > 1]
    if not tokens: return texto.title()
    if len(tokens) >= 3:
        nombres = " ".join(tokens[2:])
        apellidos = " ".join(tokens[:2])
        if solo_nombre: return tokens[2].title()
        return f"{nombres.title()} {apellidos.title()}"
    if len(tokens) == 2:
        if solo_nombre: return tokens[1].title()
        return f"{tokens[1].title()} {tokens[0].title()}"
    return tokens[0].title()

def np(s): return formatear_nombre_empatia(s, solo_nombre=True)

# ── PERFIL CRM ────────────────────────────────────────────────
def perfil_crm(tel):
    t9 = n9(tel)
    rows = _get_rows()
    p = {"tipo": "NUEVO", "nombre": None, "apellido": "", "nombre_full": "", "equipo": "",
         "imo_nombre": "", "imo_tel": "", "staff_key": None, "staff_tel": None, "staff_nom": None, "pendientes": []}
    px_row, imo_rows = None, []
    for r in rows:
        if n9(r.get("Teléfono","")) == t9: px_row = r
        if n9(r.get("Tel. IMO","")) == t9 and n9(r.get("Tel. IMO","")): imo_rows.append(r)
    if imo_rows:
        p["tipo"] = "IMO"
        p["nombre"] = np(imo_rows[0].get("IMO",""))
        p["imo_nombre"] = formatear_nombre_empatia(imo_rows[0].get("IMO",""))
        p["pendientes"] = [f"• {formatear_nombre_empatia(f'{r.get('Apellido','')} {r.get('Nombre','')}')} ({r.get('Equipo','')})" for r in imo_rows]
    elif px_row:
        p["tipo"] = "PX"
        p["nombre"] = np(px_row.get("Nombre",""))
        p["apellido"] = px_row.get("Apellido","").strip().title()
        p["nombre_full"] = f"{p['nombre']} {p['apellido']}".strip()
        p["equipo"] = px_row.get("Equipo","")
        p["imo_nombre"] = px_row.get("IMO","").strip()
        p["imo_tel"] = _d(px_row.get("Tel. IMO",""))
    if p["tipo"] == "PX":
        k = cc_por_equipo(p.get("equipo",""))
    elif p["tipo"] == "IMO":
        import re as _re2
        equipos = [r.get("Equipo","") for r in imo_rows] if imo_rows else []
        nums = [int(m.group()) for eq in equipos for m in [_re2.search(r"\d+",eq)] if m]
        eq_top = f"EQUIPO {max(nums)}" if nums else ""
        k = cc_por_equipo(eq_top)
    else:
        k = cc_libre()
    p["staff_key"] = k
    p["staff_tel"] = STAFF[k]["tel"]
    p["staff_nom"] = STAFF[k]["nombre"]
    cc_add(k)
    nom_norm = re.sub(r"\s+"," ",re.sub(r"[^\w\s]","", f"{p.get('nombre','')} {p.get('apellido','f')}".upper())).strip()
    grad_info = _GRADUADOS_IDX.get(nom_norm, {})
    if not grad_info:
        partes = nom_norm.split()
        if len(partes) >= 2:
            clave = " ".join(partes[:2])
            grad_info = next((v for n,v in _GRADUADOS_IDX.items() if n.startswith(clave)), {})
    p["graduado"] = grad_info.get("graduado", False)
    p["rango"] = grad_info.get("rango", "")
    
    # ── ASIGNACIÓN DE TIPO DEFINITIVA ──
    if p["tipo"] == "NUEVO" and p["graduado"]:
        p["tipo"] = "GRADUADO"
    
    return p

# ── SESIONES ──────────────────────────────────────────────────
def _sp(tel): return Cfg.S_SIM if str(tel).startswith("SIM_") else Cfg.S_REAL

def get_s(tel):
    path = _sp(tel)
    try:
        with FileLock(path+".lock", timeout=Cfg.LOCK_T):
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f:
                    return json.load(f).get(str(tel), {})
    except FileLockTimeout: logger.warning(f"lock get {tel}")
    except Exception as e: logger.error(f"get_s {e}")
    return {}

def set_s(tel, data):
    path = _sp(tel)
    try:
        with FileLock(path+".lock", timeout=Cfg.LOCK_T):
            d = {}
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f: d = json.load(f)
            d[str(tel)] = data
            with open(path,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False,indent=2)
    except FileLockTimeout: logger.warning(f"lock set {tel}")
    except Exception as e: logger.error(f"set_s {e}")

def del_s(tel):
    path = _sp(tel)
    try:
        with FileLock(path+".lock", timeout=Cfg.LOCK_T):
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f: d = json.load(f)
                d.pop(str(tel), None)
                with open(path,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False,indent=2)
    except Exception: pass

def add_hist(tel, nom, txt, tipo):
    path = Cfg.HIST
    try:
        with FileLock(path+".lock", timeout=Cfg.LOCK_T):
            h = []
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f: h = json.load(f)
            h.append({"telefono":str(tel),"nombre":nom or "?","texto":txt,"tipo":tipo,"hora":ahora().strftime("%d/%m %H:%M")})
            if len(h)>5000: h=h[-5000:]
            with open(path,"w",encoding="utf-8") as f: json.dump(h,f,ensure_ascii=False,indent=2)
    except Exception as e: logger.error(f"hist {e}")

# ── GOOGLE SHEETS JWT ─────────────────────────────────────────
_stok, _stok_exp = None, 0
_stok_lk = threading.Lock()

def _sheets_tok():
    global _stok, _stok_exp
    with _stok_lk:
        if _stok and time.time() < _stok_exp - 60: return _stok
        if not Cfg.CREDS: return None
        try:
            from cryptography.hazmat.primitives import hashes, serialization
            from cryptography.hazmat.primitives.asymmetric import padding as cp
            now = int(time.time())
            creds = json.loads(Cfg.CREDS)
            pem = creds["private_key"].replace("\\n","\n")
            hdr = base64.urlsafe_b64encode(json.dumps({"alg":"RS256","typ":"JWT"}).encode()).rstrip(b"=")
            pld = base64.urlsafe_b64encode(json.dumps({"iss":creds["client_email"],"scope":"https://www.googleapis.com/auth/spreadsheets","aud":"https://oauth2.googleapis.com/token","iat":now,"exp":now+3600}).encode()).rstrip(b"=")
            msg_b = hdr+b"."+pld
            pk = serialization.load_pem_private_key(pem.encode(),password=None)
            sig = pk.sign(msg_b,cp.PKCS1v15(),hashes.SHA256())
            jwt = (msg_b+b"."+base64.urlsafe_b64encode(sig).rstrip(b"=")).decode()
            r = req_lib.post("https://oauth2.googleapis.com/token",data={"grant_type":"urn:ietf:params:oauth:grant-type:jwt-bearer","assertion":jwt},timeout=10)
            if r.status_code==200:
                d=r.json(); _stok=d["access_token"]; _stok_exp=now+d.get("expires_in",3600)
                return _stok
            logger.error(f"sheets tok {r.status_code}")
        except Exception as e: logger.error(f"sheets tok err {e}")
    return None

_q = queue.Queue()
def _wsheets():
    while True:
        try:
            t = _q.get()
            if Cfg.SHEET_ID:
                tok = _sheets_tok()
                if tok:
                    tab = Cfg.SHEET_TAB.replace(" ","%20")
                    url = f"https://sheets.googleapis.com/v4/spreadsheets/{Cfg.SHEET_ID}/values/{tab}!A:K:append"
                    payload = {"values":[[ahora().strftime("%d/%m/%Y %H:%M:%S"),t.get("dir",""),str(t.get("tel","")),t.get("nom",""),t.get("tipo",""),t.get("staff",""),t.get("msg","")[:500],t.get("evento",""),t.get("estado",""),]]}
                    r = req_lib.post(url,
                        params={"valueInputOption":"RAW","insertDataOption":"INSERT_ROWS"},
                        json=payload,
                        headers={"Authorization":f"Bearer {tok}","Content-Type":"application/json"},timeout=10)
                    if r.status_code != 200:
                        logger.error(f"wsheets ERROR status={r.status_code} resp={r.text[:100]}")
                else:
                    logger.warning("wsheets: No se pudo obtener token de Google (Cfg.CREDS vacío?)")
            else:
                logger.warning("wsheets: Cfg.SHEET_ID no configurado")
            time.sleep(0.5)
        except Exception as e: logger.error(f"wsheets EXCEPCIÓN: {e}")
        finally: _q.task_done()

threading.Thread(target=_wsheets,daemon=True,name="wsheets").start()

def reg(tel, nom, tipo, msg, evento, estado="", dir_="IN", staff=""):
    if str(tel).startswith("SIM_"): return
    if not Cfg.SHEET_ID: return
    _q.put({"tel":tel,"nom":nom,"tipo":tipo,"msg":msg,"evento":evento,"estado":estado,"dir":dir_,"staff":staff})

# ── ENVÍO WA ──────────────────────────────────────────────────
def wa(tel, txt, log="BOT"):
    if str(tel).startswith("SIM_"): add_hist(tel, log, txt, "out"); return True
    if not Cfg.TOKEN: logger.critical("wa(): WA_TOKEN vacío"); return False
    try:
        r = req_lib.post(f"https://graph.facebook.com/v19.0/{Cfg.PHONE_ID}/messages",
            json={"messaging_product":"whatsapp","to":str(tel),"type":"text","text":{"body":txt,"preview_url":False}},
            headers={"Authorization":f"Bearer {Cfg.TOKEN}","Content-Type":"application/json"},timeout=10)
        if r.status_code == 200:
            logger.info(f"wa() EXITOSO tel={tel}")
            add_hist(tel, log, txt, "out"); reg(tel, log, "", txt, "BOT_OUT", dir_="OUT"); return True
        else:
            err = r.json().get("error", {}); code_err = err.get("code", 0); msg_err = err.get("message","?")[:120]
            logger.error(f"wa() FALLO tel={tel} status={r.status_code} code={code_err}: {msg_err}")
            if code_err == 190: logger.critical("⚠️ WA_TOKEN EXPIRADO")
            elif code_err == 100: logger.error(f"⚠️ PHONE_ID incorrecto: {Cfg.PHONE_ID}")
            return False
    except Exception as e: logger.error(f"wa() excepción tel={tel}: {e}"); return False


EMAILS_CC = {
    "dmoscoso": "diana.moscoso@crearpsl.com",
    "jmarin": "joyce.marin@crearpsl.com"
}

def sync_cc_all(p, motivo, extra=""):
    try:
        cc_key = p.get("staff_key") or cc_libre()
        nom_cc = p.get("staff_nom") or STAFF.get(cc_key, {}).get("nombre", "Coordinación")
        nom_px = p.get("nombre_full") or p.get("nombre", "Sin nombre")
        tel_px = p.get("_tel", "")
        
        # 1. ACTUALIZAR GOOGLE SHEETS
        if Cfg.SHEET_ID and Cfg.CREDS:
            try:
                scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
                creds = ServiceAccountCredentials.from_json_keyfile_dict(json.loads(Cfg.CREDS), scope)
                client = gspread.authorize(creds)
                sh = client.open_by_key(Cfg.SHEET_ID)
                
                # A. LOG_DERIVACIONES
                ws_log = sh.worksheet("LOG_DERIVACIONES")
                ws_log.append_row([
                    ahora().strftime("%d/%m/%Y"), 
                    ahora().strftime("%H:%M"), 
                    nom_cc, nom_px, tel_px, motivo, "DERIVADO"
                ])
                
                # B. RESUMEN_DERIVACIONES (Mantener el formato del Excel)
                ws_res = sh.worksheet("RESUMEN_DERIVACIONES")
                cell = ws_res.find(nom_cc)
                if cell:
                    row = cell.row
                    # Incrementar CASOS TOTALES (Col B)
                    val = ws_res.cell(row, 2).value or 0
                    ws_res.update_cell(row, 2, int(val) + 1)
                    
                    # Incrementar URGENTES si aplica (Col C)
                    urgente = "DEVOLUCION" in motivo.upper() or "PLATA" in motivo.upper() or "URGENTE" in motivo.upper()
                    if urgente:
                        val_u = ws_res.cell(row, 3).value or 0
                        ws_res.update_cell(row, 3, int(val_u) + 1)
                    
                    # Incrementar CONFIRMACIONES (LOG) (Col D)
                    if "CONFIRMA" in motivo.upper():
                        val_c = ws_res.cell(row, 4).value or 0
                        ws_res.update_cell(row, 4, int(val_c) + 1)

                    # Incrementar OPCION 4 / INFO (Col E)
                    if "OPCION 4" in motivo.upper() or "INFO" in motivo.upper():
                        val_i = ws_res.cell(row, 5).value or 0
                        ws_res.update_cell(row, 5, int(val_i) + 1)
            except Exception as e:
                logger.error(f"sync_cc_all Sheets Error: {e}")

        # 2. ENVIAR CORREO ELECTRÓNICO
        email_dest = EMAILS_CC.get(cc_key)
        if email_dest:
            try:
                msg = MIMEMultipart()
                msg['From'] = Cfg.GMAIL_USER
                msg['To'] = email_dest
                msg['Subject'] = f"🚨 DERIVACIÓN: {nom_px} ({motivo})"
                
                body = f"""
                <html>
                <body style='font-family: Arial, sans-serif; color: #333;'>
                    <h2 style='color: #1d4ed8;'>Nuevo Caso Derivado - CPSL Lima</h2>
                    <p><b>Participante:</b> {nom_px}</p>
                    <p><b>Teléfono:</b> <a href='https://wa.me/{tel_px}'>{tel_px}</a></p>
                    <p><b>Motivo:</b> {motivo}</p>
                    <p><b>Detalle:</b> {extra}</p>
                    <hr>
                    <p style='font-size: 12px; color: #666;'>Sincronizado automáticamente por el Cerebro Cuántico CPSL.</p>
                </body>
                </html>
                """
                msg.attach(MIMEText(body, 'html'))
                
                with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                    server.login(Cfg.GMAIL_USER, Cfg.GMAIL_PASS)
                    server.send_message(msg)
                logger.info(f"📧 Correo enviado a {email_dest} para {nom_px}")
            except Exception as e:
                logger.error(f"sync_cc_all Email Error: {e}")
                
    except Exception as e:
        logger.error(f"sync_cc_all General Error: {e}")

def notif_cc(p, motivo, extra=""):
    sync_cc_all(p, motivo, extra)
    tel_cc = p.get("staff_tel") or STAFF[cc_libre()]["tel"]
    nom_cc = p.get("staff_nom") or "Coordinación"
    nom_full = p.get("nombre_full") or ""
    nom_pila = p.get("nombre") or ""
    nom_px = nom_full if nom_full and len(nom_full.split())>1 else nom_pila or "Sin nombre"
    tel_px = p.get("_tel","")
    tipo = p.get("tipo","")
    equipo = p.get("equipo","")
    pend_n = len(p.get("pendientes",[]))
    if tipo == "IMO": ctx = f"*IMO | {pend_n} enrolados pendientes C1*"; ctx += f" | {equipo}" if equipo else ""
    elif tipo == "PX": ctx = f"*Creador {Cfg.CAMPANA_ACTUAL}{(' | '+equipo) if equipo else ''}*"; imo_n = p.get("imo_nombre",""); ctx += f"\n*Su IMO:* {imo_n}" if imo_n else ""
    else: ctx = "*Nuevo contacto*"
    logger.info(f"notif_cc INICIO → {nom_cc} tel={tel_cc} | px={nom_px} | {motivo[:40]}")
    if not tel_cc or not Cfg.TOKEN: logger.critical("notif_cc: tel_cc o WA_TOKEN vacío"); return nom_cc
    _cc_key = p.get("staff_key") or cc_libre()
    if "solicita" in motivo.lower() or "CONFIRMA" in motivo or "NO ASISTE" in motivo or "DEVOLUCION" in motivo or "directo" in motivo.lower():
        urgente = "DEVOLUCION" in motivo or "PLATA" in motivo or "URGENTE" in motivo
        try: abrir_caso(str(tel_px), nom_px, _cc_key, motivo, urgente=urgente)
        except: pass
    imo_n = p.get("imo_nombre","") or p.get("imo","")
    imo_tel = p.get("imo_tel","")
    imo_str = f"\n*IMO:* {imo_n}" if imo_n else ""
    if imo_tel and imo_tel != tel_px: imo_str += f" (wa.me/{imo_tel})"
    exito = wa(tel_cc, f"📋 *CASO DERIVADO — CPSL Lima*\n_{ahora().strftime('%d/%m/%Y %H:%M')}_\n\n*👤 Nombre:* {nom_px}\n*📱 WhatsApp:* wa.me/{tel_px}\n*🏷 {ctx}*{imo_str}\n\n*📝 Asunto:* {motivo}"+(f"\n*Detalle:* {extra}" if extra else "")+f"\n\n¿Atendiste este caso?\n1️⃣ Sí, resuelto\n2️⃣ En gestión\n3️⃣ Necesito apoyo", f"SIS→{nom_cc}")
    if not exito: logger.error(f"notif_cc: wa() falló enviando a {tel_cc}")
    else:
        try:
            s_cc = get_s(tel_cc) or {}
            s_cc["caso_followup"] = str(tel_px)
            s_cc["modo"] = "CC"
            s_cc["st_cc"] = "MAIN"
            set_s(tel_cc, s_cc)
        except: pass
    return nom_cc

# ══════════════════════════════════════════════════════════════
# FLUJO COORDINADORAS
# ══════════════════════════════════════════════════════════════
_CC_TELS = {
    "51912379744": {"key":"dmoscoso","nombre":"Diana","nombre_full":"Diana Moscoso"},
    "51933599903": {"key":"jmarin","nombre":"Joyce","nombre_full":"Joyce Marín"},
}

def _menu_cc(tel_cc, nom):
    cc_key = _CC_TELS.get(tel_cc, {}).get("key", "")
    mis_casos = casos_abiertos(cc_key) if cc_key and 'casos_abiertos' in globals() else []
    urgentes = sum(1 for c in mis_casos if c.get("estado") == "URGENTE")
    en_gestion = sum(1 for c in mis_casos if c.get("estado") == "EN_GESTION")
    abiertos = sum(1 for c in mis_casos if c.get("estado") == "ABIERTO")
    if mis_casos:
        if urgentes: alerta = f"\n\n🚨 *{urgentes} URGENTE{'S' if urgentes>1 else ''} + {abiertos+en_gestion} en seguimiento*"
        else: alerta = f"\n\n⏳ *{len(mis_casos)} caso(s) derivado(s) pendiente(s)*"; alerta += f" ({en_gestion} en gestión)" if en_gestion else ""
    else: alerta = "\n\n✅ *Sin casos derivados pendientes.*"
    wa(tel_cc, f"🏆 *TORRE DE CONTROL — CPSL Lima*\nHola {nom}!{alerta}\n\n1️⃣ Reporte de llamadas del día\n2️⃣ Registrar confirmación de PX\n3️⃣ Reportar devolución\n4️⃣ 📊 Ver mis casos derivados\n8️⃣ 📦 Ver mis casos archivados\n0️⃣ Salir\n\n💡 *Tip:* Puedes simplemente escribirme qué hiciste.\n_Ej: 'Resolví el caso de Bertha' o 'Le escribí a Juan'_", f"SIS→{nom}")

# Funciones placeholder para casos (se cargarán desde módulo si existe)
def casos_abiertos(cc_key=None): return []
def casos_cerrados(cc_key=None, limite=5): return []
def abrir_caso(tel_px, nombre, cc_key, asunto, urgente=False): return {}
def cerrar_caso(tel_px, nota): return True
def actualizar_caso(tel_px, estado, nota): return True
def casos_para_followup(horas=12): return []
def marcar_notificado(tel_px): pass
def resumen_casos(): return {"total":0,"urgentes":0,"abiertos":0,"en_gestion":0,"cerrados":0,"por_cc":{}}
def registrar_reporte(tel, txt): return "Reporte registrado.", {}
def consolidar_reportes(): return None
def reportes_pendientes(): return []
def marcar_stop(tel): pass
def marcar_confirmado(tel): pass
def resumen_recordatorios(): return {}
def run_bienvenida_e27(**k): return {"error":"módulo no disponible"}
def estado_bienvenida(): return {}

try:
    from casos_derivados import (
        casos_abiertos, casos_cerrados, abrir_caso, cerrar_caso, 
        actualizar_caso, casos_para_followup, marcar_notificado, resumen_casos
    )
except ImportError:
    pass
def detener_bienvenida(): return {"ok": False}
def ejecutar_campana(path, modo_prueba=False, limite=50): pass

def _flujo_cc(tel, up, texto, cc_info):
    nom = cc_info["nombre"]; nom_full = cc_info["nombre_full"]; cc_key = cc_info["key"]
    s = get_s(tel) or {}; st = s.get("st_cc","MAIN"); JOSE_TEL = "51919563284"
    if not s or up in {"HOLA","MENU","0","INICIO"}:
        caso_fw = s.get("caso_followup") if s else None
        s = {"modo":"CC","cc_key":cc_key,"st_cc":"MAIN"}
        if caso_fw and up == "HOLA": s["st_cc"] = "VER_CASOS"; set_s(tel, s); _flujo_cc(tel, "VER", texto, cc_info); return
        set_s(tel, s); _menu_cc(tel, nom); return
    if st == "MAIN":
        if up in {"1","2","3"} and s.get("caso_followup"):
            tel_caso = s["caso_followup"]
            if up == "1": cerrar_caso(tel_caso, f"Resuelto por {nom_full}"); wa(tel, f"✅ Caso cerrado. Registrado en el sistema.", f"SIS→{nom}"); wa(JOSE_TEL, f"✅ *Caso cerrado* por {nom_full}\nPX: wa.me/{tel_caso}", f"SIS→JOSE")
            elif up == "2": actualizar_caso(tel_caso, "EN_GESTION", f"En gestión por {nom_full}"); wa(tel, f"📋 Entendido — marcado como En gestión.", f"SIS→{nom}")
            elif up == "3": actualizar_caso(tel_caso, "ABIERTO", f"Sin respuesta — {nom_full} pide apoyo"); wa(tel, f"🆘 Avisado a coordinación para apoyo.", f"SIS→{nom}"); wa(JOSE_TEL, f"🆘 *{nom_full} necesita apoyo*\nCaso: wa.me/{tel_caso}", f"SIS→JOSE")
            s.pop("caso_followup", None); set_s(tel, s); _menu_cc(tel, nom); return
        if up == "1": wa(tel, f"📋 *Reporte del día — {nom_full}*\n\nEscribe tu reporte:\n✅ Confirmados: N\n🔀 Gestionando: N\n🛑 Devoluciones: N\n💬 Notas: texto libre\n\n_O escribe libremente._", f"SIS→{nom}"); s["st_cc"] = "ESPERANDO_REPORTE"; set_s(tel, s)
        elif up == "2": wa(tel, f"Escribe el nombre del PX que confirmó asistencia al {Cfg.CAMPANA_ACTUAL}\n\n_Ejemplo: Juan Pérez — equipo 26_\n\n9️⃣ Volver", f"SIS→{nom}"); s["st_cc"] = "ESPERANDO_CONFIRMACION"; set_s(tel, s)
        elif up == "3": wa(tel, f"Escribe los datos del PX que solicita devolución:\n\n_Ejemplo: María García — +51999888777 — monto S/250_\n\n9️⃣ Volver", f"SIS→{nom}"); s["st_cc"] = "ESPERANDO_DEVOLUCION"; set_s(tel, s)
        elif up == "4":
            h = []; hist_path = Cfg.HIST
            if os.path.exists(hist_path):
                with open(hist_path, encoding="utf-8") as f: h = json.load(f)
            notifs = [m for m in h if m.get("telefono") == tel and "TORRE DE CONTROL" in m.get("texto","") and "DERIVACIONES" not in m.get("texto","")]
            if notifs:
                lista = "\n".join([f"• [{m['hora']}] {m['texto'][m['texto'].find('Nombre:')+8:m['texto'].find('Tel:')-1].strip()}" for m in notifs[-5:] if 'Nombre:' in m.get('texto','')])
                wa(tel, f"📋 *Tus últimas derivaciones recibidas:*\n\n{lista}\n\n_Para ver el historial completo revisa la Torre de Control._\n\n9️⃣ Volver", f"SIS→{nom}")
            else: wa(tel, f"No tienes derivaciones pendientes registradas.\n\n9️⃣ Volver", f"SIS→{nom}")
        elif up in {"9","VOLVER"}: s["st_cc"] = "MAIN"; set_s(tel, s); _menu_cc(tel, nom)
        else: _menu_cc(tel, nom)
    elif st == "ESPERANDO_REPORTE":
        if up in {"9","VOLVER"}: s["st_cc"] = "MAIN"; set_s(tel, s); _menu_cc(tel, nom); return
        if len(up) == 1 and up.isdigit(): wa(tel, "Para enviar tu reporte escribe el texto\nEj: Confirmados: 5 — Gestionando: 10\n9️⃣ Cancelar", f"SIS→{nom}"); return
        hora_s = ahora().strftime("%d/%m/%Y %H:%M:%S"); reg(tel, nom_full, "", texto, "REPORTE_CC", dir_="IN", staff=nom_full); add_hist(tel, f"CC/{nom}", texto, "in")
        resumen_parsed, parsed = registrar_reporte(tel, texto)
        try:
            if push_reporte_crm: push_reporte_crm(nom_full, parsed, texto)
        except Exception as crm_e: logger.warning(f"CRM_Bridge: {crm_e}")
        wa(tel, f"✅ *Reporte registrado* — {hora_s}\n\n{resumen_parsed}\n\n_Si hay algún error, reenvía el reporte corregido._\n\n0️⃣ Salir | 9️⃣ Menú", f"SIS→{nom}")
        consolidado = consolidar_reportes()
        msg_jose = f"📊 *NUEVO REPORTE — {nom_full}*\n_{hora_s}_\n\n{resumen_parsed}"
        if consolidado: msg_jose += f"\n\n{'─'*30}\n{consolidado}"
        wa(JOSE_TEL, msg_jose, f"SIS→JOSE")
        pendientes = reportes_pendientes()
        if pendientes:
            noms_pend = ", ".join(p["nombre"].split()[0] for p in pendientes)
            wa(JOSE_TEL, f"⏳ *Reportes pendientes:* {noms_pend}", f"SIS→JOSE")
        s["st_cc"] = "MAIN"; set_s(tel, s)
    elif st == "ESPERANDO_CONFIRMACION":
        if up in {"9","VOLVER"}: s["st_cc"] = "MAIN"; set_s(tel, s); _menu_cc(tel, nom); return
        hora_s = ahora().strftime("%d/%m/%Y %H:%M:%S"); reg(tel, nom_full, "", texto, "CONFIRMA_CC", dir_="IN", staff=nom_full); add_hist(tel, f"CC/{nom}", texto, "in")
        wa(tel, f"✅ Confirmación registrada:\n_{texto}_\n\n0️⃣ Salir | 9️⃣ Menú", f"SIS→{nom}")
        wa(JOSE_TEL, f"✅ *CONFIRMACIÓN registrada por {nom_full}:*\n{texto}\n_{hora_s}_", f"SIS→JOSE")
        s["st_cc"] = "MAIN"; set_s(tel, s)
    elif st == "ESPERANDO_DEVOLUCION":
        if up in {"9","VOLVER"}: s["st_cc"] = "MAIN"; set_s(tel, s); _menu_cc(tel, nom); return
        hora_s = ahora().strftime("%d/%m/%Y %H:%M:%S"); reg(tel, nom_full, "", texto, "DEVOLUCION_CC", dir_="IN", staff=nom_full); add_hist(tel, f"CC/{nom}", texto, "in")
        wa(tel, f"⚠️ Devolución registrada:\n_{texto}_\n\n0️⃣ Salir | 9️⃣ Menú", f"SIS→{nom}")
        wa(JOSE_TEL, f"⚠️ *DEVOLUCIÓN reportada por {nom_full}:*\n{texto}\n_{hora_s}_", f"SIS→JOSE")
        s["st_cc"] = "MAIN"; set_s(tel, s)

# ── FLUJO PRINCIPAL ───────────────────────────────────────────
STOP_W = {"STOP","BAJA","DETENER","NO MAS"}
RESET_W = {"HOLA","MENU","MENÚ","0","INICIO","START","HI"}
_NEG_PATTERNS = ["NO QUIERO", "NO DESEO", "NO PUEDO", "NO ME INTERESA", "DEVUELVAN", "NO VOY", "NO ASIST", "IMPOSIBLE", "NO ESTOY INTERESADO", "OCUPADO", "DE VIAJE", "POR SALUD", "DELICAD", "NO ESTÁ EN MIS PLANES", "OTRA ACTIVIDAD", "TENGO OTRO", "FERIADO", "TRABAJO ESE"]
_AUTORESPONDER = ["GRACIAS POR COMUNICARTE CON", "FUERA DEL HORARIO", "TE RESPONDEREMOS TAN PRONTO", "HORARIO DISPONIBLE", "MENSAJE AUTOMÁTICO", "NUESTRO HORARIO"]

def _es_autoresponder(txt): return any(p in txt for p in _AUTORESPONDER)
def _detectar_negativa(txt): return any(p in txt for p in _NEG_PATTERNS)
def _limpiar_input(txt):
    limpio = re.sub(r'[^\w\s]', '', txt).strip()
    if limpio in ('0','1','2','3','4','5','6','7','8','9'): return limpio
    return txt

class ResponseTracker:
    def __init__(self): self.respuestas = 0
    def track(self): self.respuestas += 1

def flujo(tel, texto):
    tracker = ResponseTracker()
    # Parchear wa localmente para este hilo
    original_wa = globals().get('wa')
    def tracked_wa(*args, **kwargs):
        tracker.track()
        return original_wa(*args, **kwargs)
    
    # Backup y reemplazo temporal (solo para este hilo, requiere cuidado)
    # Mejor: pasar el tracker a las funciones de menú.
    # Pero para rapidez, usaré una variable en el hilo.
    
    try:
        up = texto.strip().upper()
        up_clean = _limpiar_input(up)
        if up_clean != up and up_clean in ('0','1','2','3','4','5','6','7','8','9'): up = up_clean
        if _es_autoresponder(up): logger.info(f"[AUTO-RESP] Ignorado de {tel}: {texto[:60]}"); return
        if up in STOP_W: del_s(tel); marcar_stop(tel); wa(tel,"Has sido dado de baja. Escribe HOLA para reiniciar.\n\n*Crear Poder Sin Límites Perú*","SIS"); reg(tel,"","","STOP","STOP",dir_="SYS"); return

        # --- LÓGICA DE CAMPAÑA C1 E28 (BOTONES DE RESPUESTA) ---
        if up in {"SÍ CONFIRMO", "SI CONFIRMO"}:
            s = get_s(tel) or {}
            p = s.get("p") or perfil_crm(tel)
            nom_full = p.get("nombre_full") or p.get("nombre", "")
            nom_cc = notif_cc(p, "✅ PX CONFIRMA E28 (Campaña Reactivación)", "Clic en SÍ CONFIRMO")
            reg(tel, nom_full, p.get("tipo", "PX"), "Clic en SÍ CONFIRMO", "CONFIRMA_E28", dir_="SYS", staff=nom_cc)
            wa(tel, f"¡Excelente decisión, {p.get('nombre', '')}! ✅\n\nTu confirmación para el **Capítulo 1 - Equipo 28** ha sido registrada.\n\n📍 *Lugar por confirmar (Sede Lima)*\n🗓 Del 29 al 31 de mayo\n\nTu coordinadora *{nom_cc}* ha sido notificada y se pondrá en contacto contigo muy pronto. 💪", "SIS")
            del_s(tel)
            return

        if up == "NO CONFIRMO":
            s = get_s(tel) or {}
            p = s.get("p") or perfil_crm(tel)
            nom_full = p.get("nombre_full") or p.get("nombre", "")
            nom_cc = p.get("staff_nom", "Coordinación")
            notif_cc(p, "⚠️ PX NO CONFIRMA E28", "Clic en NO CONFIRMO")
            reg(tel, nom_full, p.get("tipo", "PX"), "Clic en NO CONFIRMO", "NO_ASISTE_E28", dir_="SYS", staff=nom_cc)
            wa(tel, f"Entendemos perfectamente, {p.get('nombre', '')}. 🌟\n\nTu entrenamiento sigue disponible para cuando decidas retomarlo. Sabemos que los tiempos perfectos existen.\n\nSi deseas consultar fechas futuras o conversar con tu coordinadora *{nom_cc}*, solo escríbenos *HOLA* en cualquier momento.\n\n_Tu espacio quedará pausado._", "SIS")
            del_s(tel)
            return
        # -------------------------------------------------------
        s = get_s(tel)
        if tel in MANAGER_TELS: _flujo_gerente(tel, up, texto); return
        p = s.get("p") or perfil_crm(tel); s["p"] = p
        if p.get("tipo") in ("PX", "IMO", "NUEVO") and not s.get("notificado_entrenamiento"):
            evento_actual = _en_entrenamiento()
            if evento_actual:
                wa(tel, f"⚠️ *Aviso automático:*\nActualmente todo el equipo se encuentra en el entrenamiento presencial *{evento_actual}*.\n\n_Nuestro tiempo de respuesta será mayor al habitual. Agradecemos tu paciencia. 🙏_", "SIS")
                s["notificado_entrenamiento"] = True; set_s(tel, s)
        if not s or up in RESET_W:
            if tel in _CC_TELS: s = {"modo":"CC","cc_key":_CC_TELS[tel]["key"],"st_cc":"MAIN"}; set_s(tel, s); _menu_cc(tel, _CC_TELS[tel]["nombre"]); return
            p = perfil_crm(tel); p["_tel"] = tel; s = {"p": p, "st": "MAIN"}; set_s(tel, s); _menu_main(tel, p); return
        if tel in _CC_TELS: _flujo_cc(tel, up, texto, _CC_TELS[tel]); return
        p = s.get("p", {}); p["_tel"] = tel
        if p.get("tipo") != "NUEVO" and not p.get("staff_tel"):
            equipo = p.get("equipo",""); k = cc_por_equipo(equipo) if equipo else cc_libre()
            p["staff_key"] = k; p["staff_tel"] = STAFF[k]["tel"]; p["staff_nom"] = STAFF[k]["nombre"]; s["p"] = p; set_s(tel, s)
        st = s.get("st","MAIN"); sb = s.get("sb")
        if up in {"9","VOLVER","REGRESAR","ATRAS","ATRÁS"}: s["st"]="MAIN"; s["sb"]=None; set_s(tel,s); _menu_main(tel,p); return
        if _detectar_negativa(up) and p.get("tipo") in ("PX", "IMO"):
            nom_full = p.get("nombre_full") or p.get("nombre","")
            nom_cc = notif_cc(p, f"⚠️ PX expresa NEGATIVA/NO ASISTE", f"Mensaje: '{texto[:150]}'")
            reg(tel, nom_full, p.get("tipo",""), texto[:100], "NEGATIVA", dir_="SYS", staff=nom_cc)
            wa(tel, f"Entendido. Tu mensaje ha sido enviado a tu coordinadora *{nom_cc}*.\n\nSi cambias de parecer, escribe *HOLA* en cualquier momento.\n\n_STOP para darte de baja._", p.get("nombre","")); del_s(tel); return
        if st == "DER":
            tel_cc = p.get("staff_tel") or STAFF[cc_libre()]["tel"]; nom_cc = p.get("staff_nom","Coord"); nom_px = p.get("nombre_full") or p.get("nombre",""); nom_full_der = p.get("nombre_full") or nom_px
            wa(tel_cc, f"💬 *Mensaje de {nom_full_der}*\nTel: wa.me/{tel}\n\n{texto}", f"RELAY→{nom_cc}"); wa(tel,"✅ Mensaje entregado a tu coordinadora.\n_Escribe *0* para volver al menú._",p.get("nombre","")); return
        tipo = p.get("tipo","NUEVO")
        if os.environ.get("MODO_ENTRENAMIENTO","").lower() == "true":
            tipo_chk = p.get("tipo","NUEVO"); st_chk = s.get("st","")
            if tipo_chk in ("PX","NUEVO","GRADUADO") and st_chk in ("MAIN","NEW","") and up not in RESET_W:
                nom_e = p.get("nombre","") or "amigo/a"; s["msg_entrena"] = texto; set_s(tel, s)
                wa(tel, f"🙏 Hola {nom_e}, gracias por escribirnos.\n\nEn este momento nuestro equipo está en *entrenamiento activo* de liderazgo. Nuestras respuestas pueden tardar más de lo habitual.\n\nPara atenderte rápido cuando salgamos, déjanos el detalle:\n\n• ¿Cuál es tu consulta o situación?\n• ¿Es urgente?\n\n_Tu mensaje queda registrado. Con gusto te contactamos. 🙏_", "ENTRENA")
                wa("51919563284", f"📨 *Msg durante entrenamiento*\nDe: {p.get('nombre','?')} (wa.me/{tel})\nMsg: {texto[:120]}", "SIS→JOSE"); return
        
        if tipo == "IMO": _imo(tel, up, texto, s, p)
        elif tipo == "PX": _px(tel, up, texto, s, p)
        elif tipo == "GRADUADO": _graduado(tel, up, texto, s, p)
        else: _nuevo(tel, up, texto, s, p)
        
        # --- LÓGICA ANTI-SILENCIO ---
        # Si después de todo el flujo no hubo una llamada a wa() (detectado por log o lógica)
        # O si el usuario mandó algo que no disparó ninguna opción:
        # Nota: La implementación con tracker local es compleja por los hilos. 
        # Usaremos un chequeo de estado: si sigue en MAIN y el mensaje no fue un RESET_W.
    except Exception as e: 
        logger.error(f"flujo {tel}: {e}", exc_info=True)
        wa(tel, "🙏 *Aviso:* Tuvimos un inconveniente técnico procesando tu mensaje, pero ya estamos aquí. ¿En qué podemos ayudarte?", "FALLBACK")
        _menu_main(tel, perfil_crm(tel))

def _graduado(tel, up, texto, s, p):
    nom = p.get("nombre","Líder"); st = s.get("st","MAIN")
    if st == "MAIN":
        if up == "1": nom_cc = notif_cc(p,"Graduado solicita ser ALIADO C1 E28"); wa(tel, f"✅ ¡Excelente {nom}! Tu interés para ser *Aliado C1 E28* ha sido notificado a *{nom_cc}*.\n\nTe contactaremos para los siguientes pasos. 💪\n\n9️⃣ Volver",nom)
        elif up == "2": nom_cc = notif_cc(p,"Graduado solicita RE-ENROLAMIENTO C1 E28"); wa(tel, f"✨ ¡Qué alegría {nom}! Nada como volver a vivir la experiencia.\n\nTu coordinadora *{nom_cc}* te enviará los detalles de inversión y registro.\n\n9️⃣ Volver",nom)
        elif up == "3": nom_cc = notif_cc(p,"Graduado solicita ser STAFF/EQUIPO APOYO"); wa(tel, f"🙌 Gracias por tu servicio, {nom}.\n\nTu solicitud para ser *Staff* ha sido enviada a *{nom_cc}*. Pronto te daremos más info.\n\n9️⃣ Volver",nom)
        elif up == "4": nom_cc = notif_cc(p,"Graduado solicita atención directa"); s["st"]="DER"; set_s(tel,s); wa(tel,f"✅ Entendido. Puedes escribir tu consulta aquí y *{nom_cc}* te responderá directo.",nom)
        elif up == "0": del_s(tel); wa(tel,"Hasta pronto. Escribe HOLA para volver. 🌟",nom)
        else: _menu_main(tel, p)
    else: s["st"]="MAIN"; set_s(tel,s); _menu_main(tel,p)

def _menu_main(tel, p):
    tipo = p.get("tipo","NUEVO"); nom = p.get("nombre") or "Líder"
    if tipo == "IMO":
        n = len(p.get("pendientes",[])); al = f"\n⚠️ Tienes *{n}* enrolado{'s' if n!=1 else ''} pendiente{'s' if n!=1 else 'f'} de C1." if n else "\n✅ Todos tus enrolados al día."
        wa(tel, f"👑 *Hola {nom}* — Portal IMO{al}\n\n1️⃣ Ver mis pendientes de C1\n2️⃣ Ver TODOS mis enrolados\n3️⃣ Solicitar ser Aliado {Cfg.CAMPANA_ACTUAL}\n4️⃣ Fechas activas\n5️⃣ Hablar con Coordinación\n0️⃣ Salir\n\n_STOP para darte de baja._", nom)
    elif tipo == "PX":
        nom_cc = p.get("staff_nom","Coordinación")
        wa(tel, f"🌟 *Hola {nom}!*\nTu coordinadora: *{nom_cc}*\n\n1️⃣ Confirmar asistencia al C1 {Cfg.EQUIPO_ACTUAL}\n2️⃣ Fechas y logística\n3️⃣ Inversión y pagos\n4️⃣ Hablar con mi coordinadora\n0️⃣ Salir\n\n_STOP para darte de baja._", nom)
    elif tipo == "GRADUADO":
        wa(tel, f"🎓 *Hola {nom}!* — Portal Graduado\n\n1️⃣ Quiero ser Aliado C1 E28\n2️⃣ Re-enrolarme al C1\n3️⃣ Solicitar Staff / Equipo de Apoyo\n4️⃣ Hablar con Coordinación\n0️⃣ Salir\n\n_STOP para darte de baja._", nom)
    else: wa(tel, f"🌟 *Bienvenido a Crear Poder Sin Límites Perú*\nCanal Corporativo Oficial — Sede Lima.\n\n1️⃣ Ya participé antes (cambié de número)\n2️⃣ Soy nuevo — quiero información\n0️⃣ Salir\n\n_STOP para darte de baja._", "Sistema")

def _imo(tel, up, texto, s, p):
    nom = p.get("nombre","Líder"); pend = p.get("pendientes",[]); st = s.get("st","MAIN")
    if st == "MAIN":
        if up == "1":
            if pend:
                lista = "\n".join(pend[:20]); lista += f"\n_...y {len(pend)-20} más_" if len(pend)>20 else ""
                wa(tel, f"⏳ *Pendientes de C1 — {Cfg.EQUIPO_ACTUAL}*\n📅 {Cfg.FECHA}\n📍 {Cfg.LUGAR}\n\n{lista}\n\n¿Cómo avanzan tus gestiones?\n1️⃣ Reportar una confirmación\n2️⃣ Sigo gestionando\n3️⃣ Necesito apoyo de Coordinación\n9️⃣ Volver", nom); s["st"]="IMO_PEND"; set_s(tel,s)
            else: wa(tel,"🎉 ¡Todos tus enrolados ya se sentaron! Felicitaciones.\n\n9️⃣ Volver",nom)
        elif up == "2":
            rows = _get_rows(); t9 = n9(tel); todos = []
            for r in rows:
                if n9(r.get("Tel. IMO","")) == t9:
                    nom_px = formatear_nombre_empatia(f"{r.get('Apellido','')} {r.get('Nombre','f')}"); eq = r.get("Equipo",""); c1 = str(r.get("C1","")).strip().upper(); st_px = "✅ Sentado" if c1=="SI" else "⏳ Pendiente"
                    todos.append(f"• {nom_px} ({eq}) — {st_px}")
            if todos: lista = "\n".join(todos[:25]); lista+=f"\n_...y {len(todos)-25} más_" if len(todos)>25 else ""; wa(tel,f"📋 *Todos tus enrolados:*\n\n{lista}\n\n9️⃣ Volver",nom)
            else: wa(tel,"Sin enrolados vinculados en el sistema.\n\n9️⃣ Volver",nom)
        elif up == "3": nom_cc = notif_cc(p,f"Solicita ser Aliado {Cfg.CAMPANA_ACTUAL}",f"IMO: {p.get('imo_nombre',nom)}"); wa(tel, f"✅ Solicitud registrada.\n\nTu coordinadora *{nom_cc}* te escribirá para confirmar tu rol como Aliado.\n\n9️⃣ Volver",nom); reg(tel,nom,"IMO","Solicita ser Aliado","ALIADO",dir_="SYS",staff=nom_cc)
        elif up == "4": wa(tel,FECHAS_MSG+"\n\n9️⃣ Volver",nom)
        elif up == "5": nom_cc = notif_cc(p,"IMO solicita atención directa"); s["st"]="DER"; set_s(tel,s); wa(tel,f"✅ Derivado a *{nom_cc}*. Puedes escribirle directamente aquí.",nom)
        elif up == "0": del_s(tel); wa(tel,"Hasta pronto. Escribe HOLA para volver. 🌟",nom)
        else:
            if ia_clasificar:
                cat = ia_clasificar(texto)
                if cat == "CONFIRMA" or any(w in up for w in ["CONFIRMA","VA ASISTIR","VA A SENTARSE","ASISTIRA","ASISTIRÁ","SI VA"]):
                    nom_cc = notif_cc(p, "IMO reporta confirmación en texto libre", f"Mensaje: '{texto[:150]}f'")
                    wa(tel, f"✅ Recibido. Tu mensaje fue enviado a *{nom_cc}* para procesarlo.\n\nSi deseas confirmar formalmente, usa la opción *1* → *1*.\n\n9️⃣ Volver", nom); reg(tel, nom, "IMO", f"Texto libre: {texto[:100]}", "CONF_TEXTO", dir_="SYS", staff=nom_cc)
                elif len(texto.strip()) > 15 and ia_respuesta_imo:
                    resp_ia = ia_respuesta_imo(nom, texto, len(pend))
                    if resp_ia: wa(tel, f"{resp_ia}\n\n_Escribe *0* para menú o *5* para coordinación._", nom); guardar_feedback and guardar_feedback("IMO", texto, resp_ia)
                    else: _menu_main(tel, p)
                else: _menu_main(tel, p)
            else: _menu_main(tel, p)
    elif st == "IMO_PEND":
        if up == "1": s["st"]="IMO_CONF"; set_s(tel,s); wa(tel,"Escribe el nombre de quien confirma.\n_Escribe 9 para volver._",nom)
        elif up == "2": s["st"]="MAIN"; set_s(tel,s); wa(tel,"Perfecto. Cuando tengas confirmaciones escríbenos. 💪\n\n9️⃣ Volver / 0️⃣ Menú",nom)
        elif up == "3": nom_cc = notif_cc(p,f"IMO necesita apoyo para gestionar pendientes {Cfg.CAMPANA_ACTUAL}",f"{len(pend)} pendientes"); s["st"]="DER"; set_s(tel,s); wa(tel,f"✅ Derivado. *{nom_cc}* te apoyará directamente.",nom)
        else: s["st"]="MAIN"; set_s(tel,s); _menu_main(tel,p)
    elif st == "IMO_CONF":
        nom_cc = notif_cc(p,"IMO reporta confirmación de enrolado",f"Nombre: '{texto}'"); reg(tel,nom,"IMO",f"Confirma: {texto}","CONF_ENROLADO",dir_="SYS",staff=nom_cc)
        wa(tel, f"✅ *{texto}* registrado como confirmado.\nCoordinación ({nom_cc}) lo procesará.\n\n¿Otra confirmación? Escribe el nombre o *9* para volver.",nom)

def _px(tel, up, texto, s, p):
    nom = p.get("nombre","Líder"); nom_cc = p.get("staff_nom","Coordinación"); st = s.get("st","MAIN")
    if st == "MAIN":
        if up == "1": nom_cc2 = notif_cc(p,f"PX CONFIRMA asistencia {Cfg.CAMPANA_ACTUAL}"); reg(tel,p.get("nombre_full",nom),"PX",f"Confirma {Cfg.CAMPANA_ACTUAL}","CONFIRMA",dir_="SYS",staff=nom_cc2); wa(tel, f"¡Confirmado {nom}! ✅\n\n📍 *{Cfg.LUGAR}*\n🗓 {Cfg.FECHA}\n⏰ {Cfg.REGISTRO}\n\nRopa cómoda y botella de agua. Bloquea los 3 días.\n\nTu coordinadora *{nom_cc2}* recibirá tu confirmación. 💪",nom); del_s(tel)
        elif up == "2": wa(tel,FECHAS_MSG+"\n\n9️⃣ Volver",nom)
        elif up == "3": wa(tel, "💳 *Inversión y Pagos*\n\nBCP — Creación Cuántica E.I.R.L.\nCuenta Soles: *1934218307060*\n\n1️⃣ Enviar voucher a Coordinación\n9️⃣ Volver",nom); s["st"]="PX_PAGO"; set_s(tel,s)
        elif up == "4": nom_cc2 = notif_cc(p,"PX solicita atención directa"); s["st"]="DER"; set_s(tel,s); wa(tel,f"✅ Te derivo con *{nom_cc2}*. Escribe tu consulta aquí.",nom)
        elif up == "0": del_s(tel); wa(tel,"Hasta pronto. Escribe HOLA para volver. 🌟",nom)
        else:
            if ia_clasificar:
                cat = ia_clasificar(texto)
                if cat == "CONFIRMA":
                    nom_cc2 = notif_cc(p, "PX CONFIRMA asistencia (detectado por IA)", f"Mensaje: '{texto[:100]}f'")
                    reg(tel, p.get("nombre_full",nom), "PX", f"Confirma (IA): {texto[:80]}", "CONFIRMA", dir_="SYS", staff=nom_cc2)
                    wa(tel, f"¡Confirmado {nom}! ✅\n\n📍 *{Cfg.LUGAR}*\n🗓 {Cfg.FECHA}\n⏰ {Cfg.REGISTRO}\n\nTu coordinadora *{nom_cc2}* recibirá tu confirmación. 💪", nom)
                    guardar_feedback and guardar_feedback("PX", texto, "CONFIRMA_AUTO"); del_s(tel)
                elif cat == "PREGUNTA_FECHA": wa(tel, FECHAS_MSG + "\n\n9️⃣ Volver", nom)
                elif cat == "PREGUNTA_PAGO": wa(tel, "💳 *Inversión y Pagos*\n\nBCP — Creación Cuántica E.I.R.L.\nCuenta Soles: *1934218307060*\n\n9️⃣ Volver", nom)
                elif len(texto.strip()) > 15 and ia_respuesta_px:
                    resp_ia = ia_respuesta_px(nom, texto, nom_cc)
                    if resp_ia: wa(tel, f"{resp_ia}\n\n_Escribe *0* para menú o *4* para tu coordinadora._", nom); guardar_feedback and guardar_feedback("PX", texto, resp_ia)
                    if cat in ("QUEJA", "CONSULTA_GENERAL"): notif_cc(p, f"PX escribió (IA respondió): {cat}", f"Msg: '{texto[:100]}'")
                else: _menu_main(tel, p)
            else: _menu_main(tel, p)
    elif st == "PX_PAGO":
        if up == "1": nom_cc2 = notif_cc(p,"PX envía voucher de pago"); s["st"]="DER"; set_s(tel,s); wa(tel,f"✅ Derivado a *{nom_cc2}*. Adjunta el voucher en el siguiente mensaje.",nom)
        else: s["st"]="MAIN"; set_s(tel,s); _menu_main(tel,p)

def _nuevo(tel, up, texto, s, p):
    st = s.get("st","MAIN")
    if st == "MAIN":
        if up == "1": s["st"]="NVO_NUM"; set_s(tel,s); wa(tel, "Para encontrar tu registro escríbeme:\n\n*Nombre completo y DNI* en un solo mensaje.\n_Ej: Juan Pérez 12345678_\n\n_Escribe 9 para volver._","Sistema")
        elif up == "2": s["st"]="NVO_INFO"; set_s(tel,s); wa(tel, "🌟 *Crear Poder Sin Límites Perú*\n\nEntrenamientos de liderazgo y transformación de alto rendimiento. Salir del modo automático y crear resultados extraordinarios.\n\n1️⃣ Información del Capítulo 1\n2️⃣ Fechas 2026\n3️⃣ Inversión\n4️⃣ Hablar con Coordinación\n9️⃣ Volver","Sistema")
        elif up == "0": del_s(tel); wa(tel,"Hasta pronto. Escribe HOLA cuando quieras. 🌟","Sistema")
        else:
            if len(texto.strip()) > 10 and ia_respuesta_nuevo:
                resp_ia = ia_respuesta_nuevo(texto)
                if resp_ia: wa(tel, f"{resp_ia}\n\n_Escribe *2* para info o *0* para salir._", "Sistema"); guardar_feedback and guardar_feedback("NUEVO", texto, resp_ia)
                else: _menu_main(tel, p)
            else: _menu_main(tel,p)
    elif st == "NVO_NUM":
        k = cc_libre(); cc_add(k); tel_cc=STAFF[k]["tel"]; nom_cc=STAFF[k]["nombre"]
        wa(tel_cc, f"🔍 *VERIFICACIÓN DE IDENTIDAD*\nTel: wa.me/{tel}\nDato: '{texto}f'\nBuscar en sistema y actualizar.","SIS")
        p["staff_key"]=k; p["staff_tel"]=tel_cc; p["staff_nom"]=nom_cc; s["p"]=p; s["st"]="DER"; set_s(tel,s)
        wa(tel,f"✅ Datos enviados a Coordinación ({nom_cc}). Te responderán pronto.","Sistema"); reg(tel,texto,"NUEVO",texto,"CAMBIO_NUM",dir_="SYS",staff=nom_cc)
    elif st == "NVO_INFO":
        if up == "1": wa(tel, f"🚀 *Capítulo 1 — El Descubrimiento*\n\n3 días vivenciales para observar los mecanismos que frenan tus resultados. No es una conferencia — es transformación.\n\n*Próxima fecha:* {Cfg.FECHA}\n*Lugar:* {Cfg.LUGAR}\n\n1️⃣ Inscribirme\n9️⃣ Volver","Sistema"); s["st"]="NVO_C1"; set_s(tel,s)
        elif up == "2": wa(tel,FECHAS_MSG+"\n\n9️⃣ Volver","Sistema")
        elif up == "3": wa(tel, "💳 *Inversión*\n\nEl costo es personalizado. Tu coordinadora te dará todos los detalles.\n\n1️⃣ Contactar Coordinación\n9️⃣ Volver","Sistema"); s["st"]="NVO_INV"; set_s(tel,s)
        elif up == "4": k=cc_libre(); cc_add(k); tel_cc=STAFF[k]["tel"]; nom_cc=STAFF[k]["nombre"]; wa(tel_cc,f"🆕 *NUEVO PROSPECTO*\nTel: wa.me/{tel}","SIS"); p["staff_key"]=k; p["staff_tel"]=tel_cc; p["staff_nom"]=nom_cc; s["p"]=p; s["st"]="DER"; set_s(tel,s); wa(tel,f"✅ Derivado a Coordinación ({nom_cc}). Te escribirán pronto.","Sistema")
        elif up == "9": s["st"]="MAIN"; set_s(tel,s); _menu_main(tel,p)
    elif st in ("NVO_C1","NVO_INV"):
        if up == "1": k=cc_libre(); cc_add(k); tel_cc=STAFF[k]["tel"]; nom_cc=STAFF[k]["nombre"]; wa(tel_cc, f"🆕 *NUEVO PROSPECTO INTERESADO*\nTel: wa.me/{tel}\nInterés: {'C1' if st=='NVO_C1' else 'Inversión'}","SIS"); p["staff_key"]=k; p["staff_tel"]=tel_cc; p["staff_nom"]=nom_cc; s["p"]=p; s["st"]="DER"; set_s(tel,s); wa(tel,f"✅ Coordinación ({nom_cc}) te escribirá en breve. 🌟","Sistema")
        elif up == "9": s["st"]="NVO_INFO"; set_s(tel,s); wa(tel,"1️⃣ Info C1\n2️⃣ Fechas\n3️⃣ Inversión\n4️⃣ Coordinación\n9️⃣ Volver","Sistema")

# ── ENDPOINTS ─────────────────────────────────────────────────
@app.route("/")
def index(): return "<h1>🤖 Bot CPSL IA — Torre de Control Activa</h1><p>API Endpoint: /api/interactions</p>"

@app.route("/dashboard")
def dashboard():
    import io
    html_path = os.path.join(BASE_DIR, "dashboard.html")
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f: return f.read()
    return "Dashboard no encontrado", 404

@app.route("/api/admin/transicion_e28", methods=["GET"])
def transicion_e28():
    try:
        from casos_derivados import _cargar, _guardar, _lk, ahora
        with _lk:
            casos = _cargar()
            modificados = 0
            for k, c in casos.items():
                if c.get("estado") in ["ABIERTO", "EN_GESTION", "URGENTE"]:
                    c["estado"] = "ARCHIVADO_E27"
                    c["ts_cierre"] = ahora().isoformat()
                    if "historial" not in c: c["historial"] = []
                    c["historial"].append({"ts": ahora().isoformat(), "nota": "Cierre automático por transición a campaña C1E28"})
                    modificados += 1
            if modificados > 0:
                _guardar(casos)
        return jsonify({"ok": True, "msg": f"{modificados} casos del E27 archivados con éxito para limpiar el dashboard."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/interactions")
def api_interactions():
    date_filter = request.args.get('date')
    hist_path = Cfg.HIST if os.path.exists(Cfg.HIST) else Cfg.HIST_ALT
    try:
        with open(hist_path, 'r', encoding='utf-8') as f: data = json.load(f)
    except Exception: return jsonify({"ok": False, "msg": "Historial no disponible", "interacciones": []}), 500
    if date_filter: data = [d for d in data if d.get('ts', '').startswith(date_filter)]
    data = data[-3000:]
    return jsonify({"ok": True, "interacciones": data}), 200

@app.route("/webhook", methods=["GET"])
def wh_get():
    if request.args.get("hub.verify_token")==Cfg.VER_TOKEN: return request.args.get("hub.challenge"),200
    return "error",403

@app.route("/webhook", methods=["POST"])
def wh_post():
    d = request.get_json(silent=True)
    logger.info(f"WEBHOOK RAW: {d}")
    if not d: return jsonify({"status":"ok"}),200
    try:
        chg = d["entry"][0]["changes"][0]["value"]
        if "messages" not in chg: return jsonify({"status":"ok"}),200
        msg = chg["messages"][0]; tel = msg.get("from",""); tipo = msg.get("type","")
        if tipo=="text":
            txt = str(msg["text"]["body"]); s_wh = get_s(tel); p_wh = s_wh.get("p") or perfil_crm(tel)
            nom_d = (p_wh.get("nombre_full") or p_wh.get("imo_nombre") or p_wh.get("nombre") or tel)
            nom_h = f"({p_wh.get('tipo','?')}) {nom_d}"; add_hist(tel, nom_h, txt, "in")
            reg(tel, p_wh.get("nombre",""), p_wh.get("tipo",""), txt, "MSG_IN", staff=p_wh.get("staff_nom",""))
            threading.Thread(target=flujo,args=(tel,txt),daemon=False,name=f"f{tel[-4:]}").start()
        else:
            s_wh = get_s(tel); p_wh = s_wh.get("p") or {"tipo":"?","nombre":None}
            nom_h = f"({p_wh.get('tipo','?')}) {p_wh.get('nombre') or tel}"; add_hist(tel, nom_h, f"[{tipo}]", "in"); wa(tel,"Por favor responde con texto o el número de tu opción.","SIS")
    except Exception as e: logger.error(f"wh {e}",exc_info=True)
    return jsonify({"status":"ok"}),200

@app.route("/api/historial")
def hist_all():
    try:
        merged = {}
        for path in [Cfg.HIST, Cfg.HIST_ALT]:
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f:
                    for m in json.load(f):
                        k = f"{m.get('telefono','')}|{m.get('hora','')}|{m.get('texto','')[:30]}"; merged[k] = m
        resultado = sorted(merged.values(), key=lambda x: x.get("hora",""))
        return jsonify(resultado), 200
    except Exception as e: logger.error(f"hist_all {e}"); return jsonify([]), 200

@app.route("/api/historial/<tel>")
def hist_tel(tel):
    try:
        merged = {}
        for path in [Cfg.HIST, Cfg.HIST_ALT]:
            if os.path.exists(path):
                with open(path, encoding="utf-8") as f:
                    for m in json.load(f):
                        if str(m.get("telefono",""))==str(tel): k = f"{m.get('hora','')}|{m.get('texto','')[:30]}"; merged[k] = m
        resultado = sorted(merged.values(), key=lambda x: x.get("hora",""))
        return jsonify(resultado), 200
    except Exception as e: logger.error(f"hist_tel {e}"); return jsonify([]), 200

@app.route("/api/carga_coordinadoras")
def carga_cc(): return jsonify({k:{"nombre":v["nombre"],"casos":_carga.get(k,0),"tel":v["tel"]} for k,v in STAFF.items()}),200

@app.route("/api/enviar",methods=["POST"])
def api_enviar():
    d = request.json or {}; tel = d.get("telefono","").strip(); msg = d.get("mensaje", d.get("texto","")).strip()
    if not tel or not msg: return jsonify({"error":"faltan datos"}), 400
    if not Cfg.TOKEN: logger.error("api_enviar: TOKEN vacío"); return jsonify({"status":"error","msg":"Token WA no configurado"}), 500
    ok = wa(tel, msg, "PANEL")
    if ok: reg(tel, "PANEL", "MANUAL", msg, "MANUAL_OUT", dir_="OUT"); return jsonify({"status":"ok", "enviado":True}), 200
    else: logger.error(f"api_enviar: wa() falló para {tel}"); return jsonify({"status":"error","enviado":False,"msg":"Meta rechazó el envío — ver logs"}), 500

@app.route("/api/wa", methods=["POST"])
def api_wa(): return api_enviar()

@app.route("/api/mensaje_simulador",methods=["POST"])
def api_sim():
    d = request.json or {}; tel = d.get("telefono",""); txt = d.get("texto","")
    if not tel or not txt: return jsonify({"error":"faltan datos"}), 400
    s_sim = get_s(tel)
    if not s_sim.get("p"):
        tel_up = tel.upper()
        if "SIM_IMO" in tel_up or "SIM_GRAD" in tel_up:
            p_iny = {"tipo": "IMO", "nombre": "Gareth", "apellido": "Ramos Pérez", "nombre_full":"Gareth Said Ramos Pérez", "equipo": "EQUIPO 26", "imo_nombre": "Gareth Said Ramos Pérez", "imo_tel": tel, "staff_key": "dmoscoso", "staff_tel": STAFF["dmoscoso"]["tel"], "staff_nom": STAFF["dmoscoso"]["nombre"], "pendientes": ["• Juan Carlos Soto García (EQUIPO 26)", "• María Fernanda López Ruiz (EQUIPO 25)", "• Carlos Alberto Mendoza (EQUIPO 26)"]}
        elif "SIM_PX" in tel_up:
            p_iny = {"tipo": "PX", "nombre": "Kely", "apellido": "Arcce Rojas", "nombre_full":"Kely Arcce Rojas", "equipo": "EQUIPO 26", "imo_nombre": "Gareth Said Ramos Pérez", "imo_tel": "", "staff_key": "jmarin", "staff_tel": STAFF["jmarin"]["tel"], "staff_nom": STAFF["jmarin"]["nombre"], "pendientes": []}
        else: p_iny = {"tipo": "NUEVO", "nombre": None, "nombre_full": "", "staff_key": None, "staff_tel": None, "staff_nom": None, "pendientes": []}
        if not any(x in tel_up for x in ["SIM_IMO","SIM_PX","SIM_GRAD","SIM_NEW"]):
            p_real = perfil_crm(tel)
            if p_real.get("tipo") != "NUEVO": p_iny = p_real
        s_sim["p"] = p_iny; s_sim["st"] = "MAIN"; set_s(tel, s_sim)
    p_log = s_sim.get("p") or {}; nom_h = f"({p_log.get('tipo','SIM')}) {p_log.get('nombre_full') or p_log.get('nombre') or tel}"
    add_hist(tel, nom_h, txt, "in"); threading.Thread(target=flujo, args=(tel, txt), daemon=True, name=f"sim{tel[-4:]}").start()
    return jsonify({"status":"ok"}), 200

@app.route("/api/test_notif", methods=["POST"])
def test_notif():
    d = request.json or {}; key = d.get("cc","dmoscoso"); msg = d.get("msg","Test de notificación desde Torre de Control")
    tel = STAFF.get(key,{}).get("tel",""); nom = STAFF.get(key,{}).get("nombre","?")
    if not tel: return jsonify({"error":f"CC '{key}f' no encontrada"}), 400
    logger.info(f"TEST NOTIF → {nom} ({tel})"); exito = wa(tel, f"🧪 *TEST Torre de Control*\n{msg}\n\nSi ves esto, las notificaciones funcionan ✅", "TEST")
    return jsonify({"enviado":exito,"cc":nom,"tel":tel}), 200

# ── ENDPOINTS BIENVENIDA {Cfg.CAMPANA_ACTUAL.split()[1]} ──────────────────────────────────
@app.route("/api/bienvenida/e27/iniciar", methods=["POST"])
def api_bienvenida_iniciar():
    d = request.json or {}; limite = min(int(d.get("limite", 50) or 50), 100)
    def _run(): run_bienvenida_e27(limite=limite)
    threading.Thread(target=_run, daemon=True, name="bienvenida_e27").start()
    return jsonify({"ok": True, "limite": limite, "msg": f"Bienvenida {Cfg.CAMPANA_ACTUAL.split()[1]} iniciada — máx {limite} envíos"}), 200

@app.route("/api/bienvenida/e27/estado")
def api_bienvenida_estado(): return jsonify(estado_bienvenida()), 200

@app.route("/api/bienvenida/e27/progreso")
def api_bienvenida_progreso():
    try:
        from bienvenida_e27 import cargar_estado_envio, cargar_participantes
        estado_env = cargar_estado_envio(); pxs = cargar_participantes(); rows = []
        for px in pxs:
            tel = px.get("Telefono",""); ev = estado_env.get(tel, {})
            rows.append({"tel": tel, "nombre": f"{px.get('Apellidos','')} {px.get('Nombres','').split()[0] if px.get('Nombres') else ''}".strip(), "cc": px.get("CC_Nombre",""), "estado": ev.get("estado","PENDIENTE"), "ts": ev.get("ts","")})
        resumen_cc = {}
        for r in rows:
            cc = r["cc"]
            if cc not in resumen_cc: resumen_cc[cc] = {"enviados":0,"pendientes":0,"errores":0,"total":0}
            resumen_cc[cc]["total"] += 1; resumen_cc[cc][r["estado"].lower() if r["estado"] in ("ENVIADO","ERROR") else "pendientes"] += 1
        return jsonify({"filas": rows, "resumen_cc": resumen_cc, "total": len(rows), "enviados": sum(1 for r in rows if r["estado"]=="ENVIADO"), "pendientes": sum(1 for r in rows if r["estado"]=="PENDIENTE"), "errores": sum(1 for r in rows if r["estado"]=="ERROR")}), 200
    except: return jsonify({"error": "Módulo no disponible"}), 500

@app.route("/api/bienvenida/e27/detener", methods=["POST"])
def api_bienvenida_detener(): return jsonify(detener_bienvenida()), 200

@app.route("/api/recordatorios/resumen")
def api_recordatorios_resumen(): return jsonify(resumen_recordatorios()), 200

@app.route("/api/recordatorios/<tel>/confirmado", methods=["POST"])
def api_marcar_confirmado(tel): marcar_confirmado(tel); return jsonify({"ok": True, "tel": tel}), 200

@app.route("/api/casos/abrir", methods=["POST"])
def api_abrir_caso():
    d = request.json or {}; tel = str(d.get("tel","")).strip()
    if not tel: return jsonify({"error":"tel requerido"}), 400
    caso = abrir_caso(tel_px=tel, nombre=d.get("nombre", tel), cc_key=d.get("cc", "dmoscoso"), asunto=d.get("asunto","Derivado desde panel"), urgente=bool(d.get("urgente", False)))
    return jsonify({"ok": True, "caso": caso}), 200

@app.route("/api/casos")
def api_casos():
    cc_key = request.args.get("cc"); casos = casos_abiertos(cc_key); resumen = resumen_casos()
    return jsonify({"casos": casos, "resumen": resumen}), 200

@app.route("/api/casos/<tel_px>/cerrar", methods=["POST"])
def api_cerrar_caso(tel_px):
    d = request.json or {}; nota = d.get("nota","Cerrado desde panel"); ok = cerrar_caso(tel_px, nota)
    if ok: logger.info(f"Caso cerrado: {tel_px} — {nota}")
    return jsonify({"ok": ok}), 200

@app.route("/api/casos/<tel_px>/estado", methods=["POST"])
def api_estado_caso(tel_px):
    d = request.json or {}; estado = d.get("estado","EN_GESTION"); nota = d.get("nota",""); ok = actualizar_caso(tel_px, estado, nota)
    return jsonify({"ok": ok}), 200

@app.route("/api/followup", methods=["POST"])
def api_followup():
    horas = int((request.json or {}).get("horas", 12)); pendientes = casos_para_followup(horas); enviados = 0
    for caso in pendientes:
        cc_key = caso.get("cc_key","dmoscoso"); cc = STAFF.get(cc_key, STAFF["dmoscoso"]); nom_px = caso.get("nombre","?"); tel_px = caso.get("tel_px","?"); asunto = caso.get("asunto","?"); estado = caso.get("estado","?")
        ok = wa(cc["tel"], f"⏰ *Seguimiento de caso — CPSL Lima*\n\n*{nom_px}*\nwa.me/{tel_px}\nAsunto: {asunto}\nEstado: {estado}\n\n¿Pudiste contactar a esta persona? Responde:\n1️⃣ Sí, resuelto\n2️⃣ En gestión\n3️⃣ Sin respuesta — necesito apoyo", f"SIS→{cc['nombre']}")
        if ok: marcar_notificado(tel_px); enviados += 1; time.sleep(1)
    logger.info(f"Followup enviado: {enviados}/{len(pendientes)} casos"); return jsonify({"enviados": enviados, "total": len(pendientes)}), 200

@app.route("/api/reporte_consolidado")
def reporte_consolidado():
    consolidado = consolidar_reportes(); pendientes = reportes_pendientes()
    return jsonify({"consolidado": consolidado, "pendientes": [p["nombre"] for p in pendientes], "reportes": len(globals().get('_reportes_hoy',[])), "hora": ahora().strftime("%d/%m/%Y %H:%M")}), 200

@app.route("/api/solicitar_reporte", methods=["POST"])
def solicitar_reporte():
    d = request.json or {}; dest = d.get("cc","todas"); hora_s = ahora().strftime("%H:%M"); targets = []
    if dest == "todas": targets = list(STAFF.items())
    elif dest in STAFF: targets = [(dest, STAFF[dest])]
    else: return jsonify({"error": f"CC '{dest}' no encontrada"}), 400
    enviados = []
    for key, cc in targets:
        if key not in ("dmoscoso","jmarin","zurteaga"): continue
        ok = wa(cc["tel"], f"📊 *Torre de Control — CPSL Lima*\n\nHola {cc['nombre'].split()[0]}, por favor envía tu reporte del día:\n\n✅ Confirmados hoy: ?\n🔀 En gestión: ?\n🛑 Devoluciones: ?\n💬 Notas:\n\n_Responde este mensaje con los datos o escribe *HOLA* para ver el menú._", f"SIS→JOSE")
        enviados.append({"cc": cc["nombre"], "tel": cc["tel"], "enviado": ok}); logger.info(f"Solicitud de reporte enviada a {cc['nombre']}")
    return jsonify({"ok": True, "enviados": enviados}), 200

@app.route("/api/ia/estado")
def api_ia_estado():
    ias = estado_ias() if estado_ias else []; activas = sum(1 for ia in ias if ia.get("activa"))
    return jsonify({"total": len(ias), "activas": activas, "detalle": ias})

@app.route("/api/ia/test", methods=["POST"])
def api_ia_test():
    data = request.json or {}; texto = data.get("texto", "Hola, quiero información del C1"); ctx = data.get("contexto","px_respuesta")
    if ia_respuesta_nuevo:
        from ia_multimodelo import ia_responder as ia20; resp = ia20(texto, contexto=ctx, timeout=10)
        return jsonify({"respuesta": resp, "contexto": ctx})
    return jsonify({"respuesta": "IA no disponible", "contexto": ctx})

@app.route("/api/clear_sessions", methods=["POST"])
def clear_sessions():
    import glob; borradas = 0
    for path in [Cfg.S_REAL, Cfg.S_SIM]:
        if os.path.exists(path):
            with open(path,"w") as f: json.dump({},f); borradas += 1
    logger.info(f"Sesiones borradas ({borradas} archivos)"); return jsonify({"ok":True,"archivos_borrados":borradas}), 200

@app.route("/api/token_status")
def token_status():
    token_ok = bool(Cfg.TOKEN) and len(Cfg.TOKEN) > 20; phone_ok = bool(Cfg.PHONE_ID)
    result = {"token_configurado":token_ok,"phone_id_configurado":phone_ok,"token_len":len(Cfg.TOKEN) if Cfg.TOKEN else 0}
    if token_ok and phone_ok:
        try:
            r = req_lib.get(f"https://graph.facebook.com/v19.0/{Cfg.PHONE_ID}", headers={"Authorization":f"Bearer {Cfg.TOKEN}"},timeout=8)
            result["meta_ok"] = r.status_code == 200; result["meta_resp"] = r.status_code
        except Exception as e: result["meta_ok"] = False; result["meta_err"] = str(e)
    return jsonify(result), 200

@app.route("/status")
def status():
    rows = _get_rows()
    return jsonify({"version":"v112-FIX","status":"activo","csv_filas":len(rows),"csv_path":Cfg.CSV,"csv_ok":len(rows)>0,"hora":ahora().strftime("%d/%m/%Y %H:%M:%S"),"c1_e27":Cfg.FECHA,"carga_cc":{k:v for k,v in _carga.items()}}),200

@app.route("/chat")
def panel():
    try:
        with open(os.path.join(BASE_DIR,"panel_chat.html"),encoding="utf-8") as f: return f.read()
    except: return "<h2>Panel no disponible</h2>",200

# ── INTEGRACIÓN WORKER DE SEGUIMIENTO (CORREGIDO) ─────────────
_SEG_OK = False
_estado_worker = {"corriendo":False,"ok":0,"err":0,"total":0,"ultimo":"No disponible","log":[]}

def run_seguimiento(**kw): return {"error":"Worker no encontrado"}

try:
    from seguimiento_github import run_seguimiento as _run_seg, _estado as _estado_worker, AUTO as SEG_AUTO, HORA_AUTO as SEG_HORA
    _SEG_OK = True
    logger.info(f"✅ Worker seguimiento GitHub cargado (AUTO={SEG_AUTO}, HORA={SEG_HORA})")
except ImportError:
    try:
        from seguimiento_autonomo import run_seguimiento as _run_seg, _estado_worker
        _SEG_OK = True
        logger.info("✅ Worker seguimiento_autonomo cargado")
    except ImportError:
        _SEG_OK = False
        logger.warning("⚠️ Worker seguimiento no encontrado")

def run_seguimiento(**kw):
    """Función segura que acepta cualquier parámetro"""
    if not _SEG_OK: return {"error": "Worker no disponible"}
    try:
        return _run_seg(**kw)
    except TypeError:
        # Si hay error de parámetros, llamar sin kwargs extra
        return _run_seg(modo=kw.get("modo","ambos"))

@app.route("/api/seguimiento/estado")
def seg_estado(): return jsonify(_estado_worker), 200

@app.route("/api/seguimiento/iniciar", methods=["POST"])
def seg_iniciar():
    """✅ FIX: Compatible con cualquier versión de run_seguimiento"""
    try:
        d = request.json or {}; modo = d.get("modo", "ambos")
        if not _SEG_OK: return jsonify({"error": "Worker de seguimiento no disponible"}), 503
        res = run_seguimiento(modo=modo)
        return jsonify(res or {"ok": True}), 200
    except Exception as e:
        logger.error(f"Error en seg_iniciar: {e}", exc_info=True)
        return jsonify({"error": str(e), "detalle": "Revisa los logs del servidor"}), 500

@app.route("/api/seguimiento/log")
def seg_log(): return jsonify(_estado_worker.get("log",[])), 200

@app.route("/api/seguimiento/reenvio", methods=["POST"])
def seg_reenvio():
    if not _SEG_OK: return jsonify({"error":"Worker no disponible"}), 503
    try:
        from seguimiento_autonomo import run_reenvio
        d = request.json or {}; res = run_reenvio(horas_espera = d.get("horas_espera", 48), limite = d.get("limite"))
        return jsonify(res), 200
    except ImportError: return jsonify({"error":"Función run_reenvio no disponible"}), 503

@app.route("/api/seguimiento/detectar", methods=["GET"])
def seg_detectar():
    if not _SEG_OK: return jsonify({"error":"Worker no disponible"}), 503
    try:
        from seguimiento_autonomo import detectar_sin_respuesta
        horas = int(request.args.get("horas", 48)); sin_resp = detectar_sin_respuesta(horas_espera=horas)
        return jsonify({"total": len(sin_resp), "contactos": sin_resp[:50]}), 200
    except ImportError: return jsonify({"error":"Función detectar_sin_respuesta no disponible"}), 500

@app.route("/api/seguimiento/reenviar", methods=["POST"])
def seg_reenviar():
    d = request.json or {}
    if not _SEG_OK: return jsonify({"error":"Worker no disponible"}), 503
    try:
        from seguimiento_autonomo import run_reenvio
        res = run_reenvio(horas_espera = d.get("horas_espera", 48), limite = d.get("limite"))
        return jsonify(res), 200
    except ImportError: return jsonify({"error":"Función run_reenvio no disponible"}), 503

@app.route("/api/seguimiento/sin_respuesta")
def seg_sin_resp():
    if not _SEG_OK: return jsonify([]), 200
    try:
        from seguimiento_autonomo import detectar_sin_respuesta
        horas = int(request.args.get("horas", 48)); resultado = detectar_sin_respuesta(horas_espera=horas)
        return jsonify(resultado), 200
    except ImportError: return jsonify([]), 200

# ── SCHEDULER FOLLOWUP CASOS DERIVADOS (08:00 y 20:00) ───────
def _scheduler_followup():
    ya_hecho = set()
    while True:
        try:
            hora = ahora().strftime("%H:%M"); clave = ahora().strftime("%d/%m") + hora
            if hora in ("08:00","20:00") and clave not in ya_hecho:
                pendientes = casos_para_followup(horas=12)
                if pendientes:
                    logger.info(f"Followup automatico: {len(pendientes)} casos")
                    for caso in pendientes:
                        cc_k = caso.get("cc_key","dmoscoso"); cc = STAFF.get(cc_k, STAFF["dmoscoso"])
                        msg = (f"Seguimiento CPSL Lima\n\n{caso.get('nombre','?')}\nwa.me/{caso.get('tel_px','?')}\nAsunto: {caso.get('asunto_original') or caso.get('asunto','?')}\n\nResponde: 1=Resuelto 2=En gestion 3=Necesito apoyo")
                        wa(cc["tel"], msg, f"SIS->{cc['nombre']}"); marcar_notificado(caso.get("tel_px","")); time.sleep(1.5)
                ya_hecho.add(clave)
                if len(ya_hecho) > 100: ya_hecho.clear()
        except Exception as e: logger.error(f"followup_sched: {e}")
        time.sleep(60)

threading.Thread(target=_scheduler_followup, daemon=True, name="followup").start()
logger.info("Scheduler followup activo — 08:00 y 20:00")

# ── FLUJO GERENTE (José Sánchez) ────────────────────────────────
def _flujo_gerente(tel, up, texto):
    s = get_s(tel) or {}; st = s.get("st_jose", ""); OPCIONES = {"1","2","3","4","5","6","0"}
    def menu():
        wa(tel, f"⚡ *Torre de Control CPSL Lima*\n_José Sánchez · Gerente · {ahora().strftime('%d/%m %H:%M')}_\n\n1️⃣ Estado del sistema\n2️⃣ Derivados activos\n3️⃣ Reporte consolidado\n4️⃣ Aviso masivo a CCs\n5️⃣ Activar modo ENTRENAMIENTO\n6️⃣ Desactivar modo ENTRENAMIENTO\n7️⃣ Bienvenida {Cfg.CAMPANA_ACTUAL.split()[1]} — estado/iniciar\n8️⃣ 📊 Pegar reporte de CC al CRM\n0️⃣ Salir", "GERENTE"); s["st_jose"] = "MENU"; set_s(tel, s)
    if up in RESET_W or (not st and up not in OPCIONES): menu(); return
    if st == "AVISO_MASIVO":
        if up == "0": wa(tel, "❌ Aviso cancelado.", "GERENTE")
        else:
            enviados = 0
            for cc_tel in ["51912379744","51933599903"]:
                if wa(cc_tel, f"📢 *Mensaje de Gerencia:*\n\n{texto}", "GERENTE"): enviados += 1; time.sleep(1)
            wa(tel, f"✅ Mensaje enviado a {enviados} coordinadoras.\n\nEscribe un número para otra opción.", "GERENTE")
        s["st_jose"] = "MENU"; set_s(tel, s); return
    if st == "PEGAR_REPORTE":
        if up == "0": wa(tel, "❌ Cancelado.", "GERENTE"); s["st_jose"] = "MENU"; set_s(tel, s); return
        cc_detectada, exito = push_reporte_jose(texto) if push_reporte_jose else ("DESCONOCIDA", False)
        if cc_detectada == "DESCONOCIDA":
            s["_reporte_pendiente"] = texto; s["st_jose"] = "CONFIRMAR_CC"; set_s(tel, s)
            wa(tel, f"🤖 *No pude detectar de qué coordinadora es.*\n\n¿De quién es este reporte?\n\n1️⃣ Diana Moscoso\n2️⃣ Joyce Marín\n0️⃣ Cancelar", "GERENTE"); return
        if exito: wa(tel, f"✅ *Reporte registrado en el CRM*\nCC detectada: *{cc_detectada}*\n\n_El CRM ya puede ver estos datos en el Buscador 360°._\n\nEscribe un número para otra opción.", "GERENTE")
        else: wa(tel, f"⚠️ No pude enviar el reporte al CRM.\nCC detectada: *{cc_detectada}*\n\nVerifica que GOOGLE_CREDENTIALS esté configurado en Render.", "GERENTE")
        s["st_jose"] = "MENU"; set_s(tel, s); return
    if st == "CONFIRMAR_CC":
        cc_map = {"1": "DIANA", "2": "JOYCE", "3": "ZULEY"}
        if up in cc_map:
            cc_nombre = cc_map[up]; texto_pend = s.pop("_reporte_pendiente", "")
            if texto_pend and push_reporte_crm:
                from reportes_cc import parsear_reporte; parsed = parsear_reporte(texto_pend, cc_nombre); exito = push_reporte_crm(cc_nombre, parsed, texto_pend)
                if exito: wa(tel, f"✅ *Reporte de {cc_nombre} registrado en el CRM*\n\n_El CRM ya puede ver estos datos._\n\nEscribe un número para otra opción.", "GERENTE")
                else: wa(tel, f"⚠️ Error al enviar el reporte de {cc_nombre}.", "GERENTE")
            else: wa(tel, "⚠️ Reporte perdido. Intenta de nuevo con opción 8.", "GERENTE")
        elif up == "0": s.pop("_reporte_pendiente", None); wa(tel, "❌ Cancelado.", "GERENTE")
        else: wa(tel, "Responde 1️⃣ Diana, 2️⃣ Joyce, o 0️⃣ Cancelar", "GERENTE"); return
        s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "1":
        res = resumen_casos(); por_cc = res.get("por_cc", {}); cc_txt = "\n".join(f"  · {STAFF.get(k,{}).get('nombre',k)}: {n} casos" for k,n in por_cc.items()) or "  Sin casos asignados"
        wa(tel, f"📊 *Estado CPSL Lima*\n_{ahora().strftime('%d/%m/%Y %H:%M')}_\n\n🔴 Urgentes: {res.get('urgentes',0)}\n⏳ Abiertos: {res.get('abiertos',0)}\n🔵 En gestión: {res.get('en_gestion',0)}\n✅ Cerrados: {res.get('cerrados',0)}\n\nPor coordinadora:\n{cc_txt}\n\n_Escribe un número para otra opción._", "GERENTE"); s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "2":
        activos = casos_abiertos()
        if not activos: wa(tel, "✅ Sin casos derivados activos.\n\n_Escribe un número para otra opción._", "GERENTE")
        else:
            lineas = []; emojis = {"URGENTE":"🔴","EN_GESTION":"⏳","ABIERTO":"🔵"}
            for c in activos[:12]: emoji = emojis.get(c["estado"],"▪"); cc_n = STAFF.get(c.get("cc_key",""),{}).get("nombre","?"); lineas.append(f"{emoji} {c.get('nombre','?f')[:22]} → {cc_n}")
            wa(tel, f"🗂 *Derivados activos ({len(activos)}):*\n\n" + "\n".join(lineas) + "\n\n_Escribe un número para otra opción._", "GERENTE")
        s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "3":
        try:
            if kpi_consolidado_whatsapp: kpi_msg = kpi_consolidado_whatsapp(); wa(tel, kpi_msg, "GERENTE")
            else:
                consolidado = consolidar_reportes()
                if consolidado: wa(tel, consolidado, "GERENTE")
                else: wa(tel, "Sin datos de consolidado.", "GERENTE")
        except Exception as e: logger.error(f"KPI consolidado: {e}"); wa(tel, "Error al obtener KPIs.", "GERENTE")
        s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "4": s["st_jose"] = "AVISO_MASIVO"; set_s(tel, s); wa(tel, "📢 Escribe el mensaje que quieres enviar a *todas las coordinadoras* (Diana y Joyce):\n\n_O escribe 0 para cancelar._", "GERENTE"); return
    if up == "5": os.environ["MODO_ENTRENAMIENTO"] = "true"; wa(tel, "⚡ *Modo ENTRENAMIENTO activado.*\nEl bot pedirá más detalle antes de derivar y notificará la demora.\n\n_Escribe un número para otra opción._", "GERENTE"); s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "6": os.environ["MODO_ENTRENAMIENTO"] = ""; wa(tel, "✅ *Modo ENTRENAMIENTO desactivado.*\nEl bot opera en modo normal.\n\n_Escribe un número para otra opción._", "GERENTE"); s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "7":
        est = estado_bienvenida()
        if est.get("corriendo"): wa(tel, f"📤 *Bienvenida {Cfg.CAMPANA_ACTUAL.split()[1]} en curso*\nEnviados: {est.get('enviados',0)}\nPendientes: {est.get('pendientes_total',0)}\n\nEscribe *7A* para detener.", "GERENTE")
        elif up == "7A": detener_bienvenida(); wa(tel, "⏹ Bienvenida detenida.", "GERENTE")
        else: wa(tel, f"📤 *Bienvenida {Cfg.CAMPANA_ACTUAL.split()[1]} — Estado*\nEnviados histórico: {est.get('enviados_total_historico',0)}/275\nPendientes: {est.get('pendientes_total',275)}\n\nEscribe *7S* para iniciar envío (50 por ciclo).", "GERENTE")
        s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "7S":
        def _b(): run_bienvenida_e27(limite=50)
        import threading as _th; _th.Thread(target=_b, daemon=True).start()
        wa(tel, f"📤 Bienvenida {Cfg.CAMPANA_ACTUAL.split()[1]} iniciada — 50 mensajes en cola (45s/msg).", "GERENTE"); s["st_jose"] = "MENU"; set_s(tel, s); return
    if up == "8": s["st_jose"] = "PEGAR_REPORTE"; set_s(tel, s); wa(tel, "📊 *Pegar Reporte al CRM*\n\nPega aquí el reporte de cualquier coordinadora.\nLa IA detectará automáticamente de quién es (Diana o Joyce) basado en el contenido, y lo enviará directo al CRM.\n\n_O escribe 0 para cancelar._", "GERENTE"); return
    if up == "0": wa(tel, "👋 Hasta pronto, José.", "GERENTE"); s["st_jose"] = ""; set_s(tel, s); return
    menu()

# ── KEEPALIVE DE VENTANA 24H ────────────────────────────────────
def _keepalive_loop():
    import time as _t; INTERVALO = 23 * 3600; _t.sleep(3600); _ka_intentos = {}; _ka_intervalo = 3 * 3600
    while True:
        try:
            ahora_s = ahora().strftime("%d/%m/%Y %H:%M")
            CCS_KEEPALIVE = [("51912379744", "Diana"), ("51933599903", "Joyce")]
            for tel_cc, nom in CCS_KEEPALIVE:
                ok = wa(tel_cc, f"👋 Hola {nom} — CPSL Lima al día.\nEstamos disponibles. Escribe *HOLA* si necesitas algo.", "KEEPALIVE")
                if ok: logger.info(f"Keepalive enviado a {nom}"); _t.sleep(3)
            try:
                activos = casos_abiertos()
                for caso in activos[:20]:
                    tel_px = caso.get("tel_px",""); _nom_raw = caso.get("nombre_full","") or caso.get("nombre","")
                    if _nom_raw:
                        _nom_raw = _nom_raw.strip(); _p = _nom_raw.split()
                        if _nom_raw.isupper() and len(_p) >= 3: nom_px = " ".join(w.title() for w in _p[2:] + _p[:2])
                        else: nom_px = _nom_raw.title()
                    else: nom_px = ""
                    if tel_px: wa(tel_px, f"Hola {nom_px} — te escribimos desde Crear Poder Sin Límites Perú.\nTu coordinadora está en contacto. Escribe *HOLA* si tienes alguna consulta.", "KEEPALIVE"); _t.sleep(2)
            except Exception as e: logger.error(f"keepalive_casos: {e}")
            logger.info(f"Keepalive completado — {ahora_s}")
        except Exception as e: logger.error(f"keepalive_loop: {e}")
        _t.sleep(INTERVALO)

threading.Thread(target=_keepalive_loop, daemon=True, name="keepalive").start()
logger.info("✅ Keepalive loop activo — ciclo 23h")

@app.route("/api/bienvenida/preview", methods=["POST"])
def api_bienvenida_preview():
    d = request.json or {}; nom = d.get("nombre","Participante"); cc_key = d.get("cc","dmoscoso"); cc = STAFF.get(cc_key, STAFF["dmoscoso"]); CC_EMOJI = {"dmoscoso":"🌟","jmarin":"⚡","zurteaga":"🔥"}
    pila = nom.split()[0].title() if nom else "Hola"
    msg = f"Hola {pila} 👋\n\nBienvenido/a a *Crear Poder Sin Límites Perú* 🇵🇪\n\nTu inscripción para *C1 {Cfg.EQUIPO_ACTUAL}* ha sido confirmada.\n📅 Viernes 01, Sábado 02 y Domingo 03 de Mayo 2026\n🏨 Hotel José Antonio Deluxe, Miraflores\n\nTu coordinadora asignada es *{cc['nombre']}* {CC_EMOJI.get(cc_key,'🌟')}\nGuárdala en tus contactos:\n📱 wa.me/{cc['tel']}\n\nSi tienes alguna consulta, escríbele directamente.\n\n_¡Nos vemos en el salón!_ ⚡"
    return jsonify({"mensaje": msg, "cc": cc["nombre"], "tel_cc": cc["tel"]}), 200

@app.route("/api/bienvenida_plantilla/iniciar", methods=["GET", "POST"])
def api_bienvenida_plantilla_iniciar():
    import threading; d = request.get_json(silent=True) or {}; limite = min(int(d.get("limite", 50) or 50), 200)
    def _run():
        try:
            from enviar_bienvenida_plantilla import ejecutar_masivo; ejecutar_masivo(limite=limite)
        except Exception as e: logger.error(f"bienvenida_plantilla_iniciar: {e}", exc_info=True)
    threading.Thread(target=_run, daemon=True, name="masivo_plantilla").start()
    return jsonify({"ok": True, "msg": f"Envío masivo con plantilla iniciado ({limite} pxs)."}), 200

# ══════════════════════════════════════════════════════════════
# ENDPOINT: ENVÍO MASIVO DE PLANTILLA APROBADA POR LOTE
# Recibe JSON con lista de contactos [{tel, nombre}] y plantilla
# ══════════════════════════════════════════════════════════════
_envio_plantilla_estado = {"corriendo": False, "total": 0, "enviados": 0, "errores": 0, "log": []}

def _wa_template_send(tel, nombre, template_name="emergencia_enrolamiento", lang="es"):
    """Envía una plantilla aprobada de WhatsApp con parámetro {{1}}=nombre."""
    if not Cfg.TOKEN:
        logger.error("_wa_template_send: WA_TOKEN vacío")
        return False
    payload = {
        "messaging_product": "whatsapp",
        "to": str(tel),
        "type": "template",
        "template": {
            "name": template_name,
            "language": {"code": lang},
            "components": [
                {"type": "body", "parameters": [{"type": "text", "text": nombre}]}
            ]
        }
    }
    try:
        r = req_lib.post(
            f"https://graph.facebook.com/v19.0/{Cfg.PHONE_ID}/messages",
            json=payload,
            headers={"Authorization": f"Bearer {Cfg.TOKEN}", "Content-Type": "application/json"},
            timeout=15
        )
        if r.status_code == 200:
            logger.info(f"TEMPLATE OK -> {tel} ({nombre})")
            return True
        err = r.json().get("error", {})
        logger.error(f"TEMPLATE FAIL {tel}: {r.status_code} - {err.get('message','?')[:150]}")
        return False
    except Exception as e:
        logger.error(f"TEMPLATE EXC {tel}: {e}")
        return False

@app.route("/api/plantilla/enviar_lote", methods=["POST"])
def api_plantilla_enviar_lote():
    """Recibe {contactos: [{tel, nombre}], plantilla: str, pausa: int} y envía en hilo."""
    global _envio_plantilla_estado
    if _envio_plantilla_estado["corriendo"]:
        return jsonify({"ok": False, "msg": "Ya hay un envío en curso.", "estado": _envio_plantilla_estado}), 409

    d = request.get_json(silent=True) or {}
    contactos = d.get("contactos", [])
    plantilla = d.get("plantilla", "emergencia_enrolamiento")
    pausa = max(int(d.get("pausa", 20)), 5)  # mínimo 5 segundos entre mensajes

    if not contactos:
        return jsonify({"ok": False, "msg": "Lista de contactos vacía."}), 400

    _envio_plantilla_estado = {"corriendo": True, "total": len(contactos), "enviados": 0, "errores": 0, "log": []}

    def _run_lote():
        global _envio_plantilla_estado
        logger.info(f"HILO PLANTILLA ARRANCADO: {len(contactos)} contactos, plantilla={plantilla}, pausa={pausa}s")
        try:
            for i, c in enumerate(contactos):
                logger.info(f"PLANTILLA [{i+1}/{len(contactos)}] procesando...")
                tel = str(c.get("tel", "")).strip()
                nombre = str(c.get("nombre", "Amigo/a")).strip().title()
                if not tel or len(tel) < 10:
                    _envio_plantilla_estado["errores"] += 1
                    _envio_plantilla_estado["log"].append(f"SKIP: tel inválido '{tel}'")
                    continue
                ok = _wa_template_send(tel, nombre, template_name=plantilla)
                if ok:
                    _envio_plantilla_estado["enviados"] += 1
                    _envio_plantilla_estado["log"].append(f"OK: {nombre} ({tel})")
                else:
                    _envio_plantilla_estado["errores"] += 1
                    _envio_plantilla_estado["log"].append(f"FAIL: {nombre} ({tel})")
                if i < len(contactos) - 1:
                    time.sleep(pausa)
        except Exception as e:
            logger.error(f"enviar_lote error: {e}", exc_info=True)
            _envio_plantilla_estado["log"].append(f"ERROR FATAL: {e}")
        finally:
            _envio_plantilla_estado["corriendo"] = False
            logger.info(f"Lote plantilla finalizado: {_envio_plantilla_estado['enviados']}/{_envio_plantilla_estado['total']}")

    import threading
    threading.Thread(target=_run_lote, daemon=True, name="envio_plantilla_lote").start()
    return jsonify({"ok": True, "msg": f"Envío iniciado: {len(contactos)} contactos con plantilla '{plantilla}' (pausa {pausa}s)", "total": len(contactos)}), 200

@app.route("/api/plantilla/estado")
def api_plantilla_estado():
    """Devuelve el estado actual del envío masivo de plantillas."""
    return jsonify(_envio_plantilla_estado), 200

@app.route("/api/plantilla/reset", methods=["POST"])
def api_plantilla_reset():
    """Resetea el estado del envío para poder re-lanzar."""
    global _envio_plantilla_estado
    _envio_plantilla_estado = {"corriendo": False, "total": 0, "enviados": 0, "errores": 0, "log": ["RESET manual"]}
    return jsonify({"ok": True, "msg": "Estado reseteado."}), 200
@app.route("/api/bienvenida/v1/iniciar", methods=["POST"])
def api_bienvenida_v1_iniciar():
    import threading; d = request.json or {}; limite = min(int(d.get("limite", 50) or 50), 100)
    def _run():
        try:
            from bienvenida_e27 import ejecutar_campana
            csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "asignacion_c1_e27.csv")
            xlsx_path = csv_path.replace(".csv",".xlsx"); path = xlsx_path if os.path.exists(xlsx_path) else csv_path
            if not os.path.exists(path): logger.error(f"bienvenida: archivo no encontrado: {path}"); return
            ejecutar_campana(path, modo_prueba=False, limite=limite)
        except Exception as e: logger.error(f"bienvenida_iniciar: {e}", exc_info=True)
    threading.Thread(target=_run, daemon=True, name="bienvenida_e27").start()
    return jsonify({"ok": True, "limite": limite, "msg": f"Campaña iniciada — {limite} mensajes"}), 200

@app.route("/api/bienvenida/v1/estado")
def api_bienvenida_v1_estado():
    try:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bienvenida_estado.json")
        if os.path.exists(path):
            with open(path) as f: return jsonify(json.load(f)), 200
    except: pass
    return jsonify({"corriendo": False, "enviados": 0, "total": 0}), 200

def _disparar_recordatorios_imos():
    logger.info("⏸️ Envío de recordatorios a IMOs PAUSADO por reprogramación de fecha C1.")
    return
    try:
        from seguimiento_imos import enviar_recordatorios_imos
        from sync_cloud import conectar_sheets
        c = conectar_sheets()
        if c:
            enviar_recordatorios_imos(c, os.environ.get("CRM_SHEET_ID", "1IoCYs1qfOTdn3XWyeK64jsUfAXOFgv3Wa6uJBM-lR2Y"))
            logger.info("✅ Recordatorios enviados a IMOs correctamente.")
        else: logger.error("❌ Sin conexión a Sheets para enviar recordatorios IMO.")
    except Exception as e: logger.error(f"Error en _disparar_recordatorios_imos: {e}")

# ── SCHEDULER: SEGUIMIENTO IMOs Y RECORDATORIOS ──
def _scheduler_imos():
    import time as _time; import subprocess; ya_enviado_hoy = set()
    while True:
        try:
            hora = ahora().strftime("%H:%M"); fecha = ahora().strftime("%d/%m"); minuto = int(ahora().strftime("%M"))
            if minuto in (0, 15, 30, 45):
                try:
                    from vigilante_ia import ejecutar_vigilancia; ejecutar_vigilancia()
                except Exception as e: logger.error(f"[VIGILANTE] Error: {e}")
                try:
                    script_auditoria = r"C:\Users\josem\Downloads\CONTROL_SISTEMA_CREARLIMA\robot_doble_chequeo.py"
                    if os.path.exists(script_auditoria): subprocess.Popen(["python", script_auditoria], cwd=os.path.dirname(script_auditoria))
                except Exception as e: logger.error(f"[IMO-SCHED] Error ejecutando auditoria: {e}")
            if minuto in (0, 30):
                logger.info("[IMO-SCHED] Refrescando datos de gestion...")
                try:
                    script_path = r"C:\Users\josem\Downloads\CONTROL_SISTEMA_CREARLIMA\robot_gestion_llamadas.py"
                    if os.path.exists(script_path): subprocess.Popen(["python", script_path], cwd=os.path.dirname(script_path)); logger.info("[IMO-SCHED] Ejecutando robot_gestion_llamadas.py en segundo plano...")
                except Exception as e: logger.error(f"[IMO-SCHED] Error ejecutando robot: {e}")
            es_dia_previo = (fecha == "30/04"); hora_imos = "18:00" if es_dia_previo else "07:30"; hora_recordatorios = "18:00" if es_dia_previo else "10:00"
            clave_imo = f"{fecha}-imo"
            if hora == hora_imos and clave_imo not in ya_enviado_hoy: ya_enviado_hoy.add(clave_imo); logger.info(f"[IMO-SCHED] Enviando mensajes principales a IMOs ({hora_imos})..."); _enviar_mensajes_imos()
            clave_rec = f"{fecha}-rec"
            if hora == hora_recordatorios and clave_rec not in ya_enviado_hoy: ya_enviado_hoy.add(clave_rec); logger.info(f"[RECORDATORIOS] Enviando recordatorios a IMOs ({hora_recordatorios})..."); _disparar_recordatorios_imos()
        except Exception as e: logger.error(f"[IMO-SCHED] Error general: {e}")
        _time.sleep(60)

def _enviar_mensajes_imos():
    logger.info("⏸️ Seguimiento principal a IMOs PAUSADO por reprogramación de fecha C1.")
    return
    try:
        from seguimiento_imos import enviar_seguimiento_diario, en_horario
        if not en_horario(): logger.info("[IMO] Fuera de horario (9-17h)"); return
        from sync_cloud import conectar_sheets
        SHEET_ID = os.environ.get("CRM_SHEET_ID", "1IoCYs1qfOTdn3XWyeK64jsUfAXOFgv3Wa6uJBM-lR2Y"); c = conectar_sheets()
        if not c: logger.error("[IMO] Sin conexion Sheets"); return
        n = enviar_seguimiento_diario(c, SHEET_ID); logger.info(f"[IMO] Seguimiento completado: {n} mensajes enviados")
    except Exception as e: logger.error(f"[IMO] Error: {e}", exc_info=True)

@app.route("/api/imo/respuesta", methods=["POST"])
def api_imo_respuesta():
    d = request.json or {}; imo_nombre = d.get("imo_nombre", ""); imo_tel = d.get("imo_tel", ""); px_nombre = d.get("px_nombre", ""); respuesta = d.get("respuesta", ""); cc_alias = d.get("cc_alias", "")
    if not all([imo_nombre, px_nombre, respuesta]): return jsonify({"error": "Faltan campos"}), 400
    try:
        from seguimiento_imos import guardar_respuesta_imo
        from sync_cloud import conectar_sheets
        SHEET_ID = os.environ.get("CRM_SHEET_ID", "1IoCYs1qfOTdn3XWyeK64jsUfAXOFgv3Wa6uJBM-lR2Y"); c = conectar_sheets()
        ok = guardar_respuesta_imo(c, SHEET_ID, imo_nombre, imo_tel, px_nombre, respuesta, cc_alias)
        return jsonify({"ok": ok}), 200
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/imo/pendientes/<cc>")
def api_imo_pendientes(cc):
    try:
        from seguimiento_imos import obtener_respuestas_pendientes_cc
        from sync_cloud import conectar_sheets
        SHEET_ID = os.environ.get("CRM_SHEET_ID", "1IoCYs1qfOTdn3XWyeK64jsUfAXOFgv3Wa6uJBM-lR2Y"); c = conectar_sheets()
        pend = obtener_respuestas_pendientes_cc(c, SHEET_ID, cc.upper())
        return jsonify(pend), 200
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/imo/trigger", methods=["GET", "POST"])
def api_imo_trigger():
    def _run(): _enviar_mensajes_imos()
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "msg": "Seguimiento IMO disparado"}), 200

@app.route("/api/imo/force-send", methods=["GET", "POST"])
def api_imo_force_send():
    try:
        from seguimiento_imos import enviar_seguimiento_diario
        from sync_cloud import conectar_sheets
        SHEET_ID = os.environ.get("CRM_SHEET_ID", "1IoCYs1qfOTdn3XWyeK64jsUfAXOFgv3Wa6uJBM-lR2Y"); c = conectar_sheets()
        if not c: return jsonify({"error": "Sin conexion Sheets"}), 500
        n = enviar_seguimiento_diario(c, SHEET_ID); return jsonify({"ok": True, "enviados": n}), 200
    except Exception as e: return jsonify({"error": str(e)}), 500

@app.route("/api/imo/force-recordatorio", methods=["GET", "POST"])
def api_imo_force_recordatorio():
    try:
        from seguimiento_imos import enviar_recordatorios_imos
        from sync_cloud import conectar_sheets
        import threading
        SHEET_ID = os.environ.get("CRM_SHEET_ID", "1IoCYs1qfOTdn3XWyeK64jsUfAXOFgv3Wa6uJBM-lR2Y")
        
        def run_async():
            c = conectar_sheets()
            if c:
                enviar_recordatorios_imos(c, SHEET_ID)
                
        threading.Thread(target=run_async, daemon=True).start()
        return jsonify({"ok": True, "msg": "Enviando recordatorios en segundo plano..."}), 200
    except Exception as e: return jsonify({"error": str(e)}), 500

# ── CRM DATA CACHE ──────────────────────────────────────────
_CRM_CACHE = {"last_refresh": 0, "data": None, "otros": {}, "stats": ""}

def refresh_crm_cache(force=False):
    import time; now = time.time()
    if not force and _CRM_CACHE["data"] is not None and (now - _CRM_CACHE["last_refresh"] < 300): return _CRM_CACHE["data"], _CRM_CACHE.get("otros", {})
    try:
        from sync_cloud import load_master_cloud, load_productividad_cloud, load_asignaciones_cloud, load_gestion_llamadas_cloud
        df = load_master_cloud(); df_prod = load_productividad_cloud(); df_asig = load_asignaciones_cloud(); df_llamas = load_gestion_llamadas_cloud()
        otros = {"productividad": df_prod, "asignaciones": df_asig, "gestion_llamadas": df_llamas}
        if not df.empty:
            _CRM_CACHE["data"] = df; _CRM_CACHE["otros"] = otros; _CRM_CACHE["last_refresh"] = now
            total = len(df); sentados = df['Estatus C1'].apply(lambda x: "SI" in str(x).upper() or "SENTADO" in str(x).upper()).sum() if 'Estatus C1' in df.columns else 0
            graduados = df['Participación'].str.contains('GRADUADO', case=False, na=False).sum() if 'Participación' in df.columns else 0
            prod_gestiones = len(df_prod) if not df_prod.empty else 0; asig_total = len(df_asig) if not df_asig.empty else 0
            _CRM_CACHE["stats"] = f"Total Base (Hoja 1): {total}. Sentados C1: {sentados}. Graduados: {graduados}. Meta C1: 325.\nGestiones de Productividad: {prod_gestiones}. Asignaciones registradas: {asig_total}."
            logger.info("Caché de CRM refrescada (Múltiples Pestañas) ✅")
        return df, otros
    except Exception as e: logger.error(f"Error refrescando caché CRM: {e}"); return _CRM_CACHE["data"], _CRM_CACHE.get("otros", {})

@app.route("/api/chat", methods=["POST"])
def api_chat():
    d = request.json or {}; msg = d.get("message", ""); source = d.get("source", "web")
    if not msg: return jsonify({"reply": "No recibí ningún mensaje."}), 400
    try:
        from ia_chain import ia_responder; df, otros = refresh_crm_cache(); stats_txt = _CRM_CACHE["stats"]; search_results = ""; msg_norm = msg.upper().strip()
        if len(msg_norm) > 3 and df is not None:
            cols = {c.strip().upper(): c for c in df.columns}; c_nom = cols.get('NOMBRES', 'Nombres'); c_ape = cols.get('APELLIDOS', 'Apellidos'); c_dni = cols.get('DNI', 'DNI'); c_tel = cols.get('TELÉFONO', cols.get('TELEFONO', 'Teléfono')); c_est = cols.get('ESTATUS C1', 'Estatus C1'); c_coo = cols.get('COORDINADOR', 'Coordinador')
            matches = []
            try:
                mask = (df[c_nom].astype(str).str.upper().str.contains(msg_norm, na=False) | df[c_ape].astype(str).str.upper().str.contains(msg_norm, na=False) | df[c_dni].astype(str).str.contains(msg_norm, na=False) | df[c_tel].astype(str).str.contains(msg_norm, na=False))
                results = df[mask].head(3)
                for _, r in results.iterrows():
                    info = f"- {r.get(c_nom)} {r.get(c_ape)} (DNI: {r.get(c_dni)}) -> C1: {r.get(c_est)}, CC: {r.get(c_coo)}"
                    df_prod = otros.get("productividad")
                    if df_prod is not None and not df_prod.empty:
                        prod_mask = df_prod['NombreCompleto'].astype(str).str.upper().str.contains(str(r.get(c_nom)).upper(), na=False)
                        if prod_mask.any():
                            last_gest = df_prod[prod_mask].iloc[-1]; info += f" | Última Gestión: {last_gest.get('Resultado Gestión')} ({last_gest.get('Fecha Gestión')})"
                    matches.append(info)
            except Exception as se: logger.warning(f"Search error: {se}")
            if matches: search_results = "\nHe encontrado estos registros en la base de datos (Maestro + Productividad):\n" + "\n".join(matches)
        contexto = f"Eres el Cerebro de CPSL (🔱), el asistente de inteligencia artificial definitivo y estratégico para el CRM de Crear Lima. Estás entrenado para responder a consultas de usuarios internos: Gerentes (ej. José), Coordinadoras (Diana, Joyce, Zuley) y el CMJ (Coordinador de Maestría del Juego). Tienes acceso a la información de todas las pestañas del CRM en la nube (Google Sheets): 1. Hoja 1 (Base Maestra), 2. PRODUCTIVIDAD, 3. ASIGNACIONES, 4. GESTION_LLAMADAS, etc.\n\n📊 ESTADO ACTUAL DE LA CAMPAÑA:\n{stats_txt}\n\nTUS DIRECTRICES:\n- Sé profesional, empático y orientado a resultados.\n- Adapta tu respuesta al rol: Si es una coordinadora, dale foco a sus equipos y efectividad. Si es gerencia, enfócate en la Meta C1 (325) y panoramas globales. Si es el Coordinador de Maestría del Juego (CMJ), apóyalo con el estatus detallado de los prospectos y su historial.\n- Si te preguntan por un participante, responde usando ÚNICAMENTE la data encontrada a continuación.\n- Si te piden gráficas o análisis complejos, explícales los números disponibles o guíalos a la pestaña 'Sala de Guerra' del CRM donde se visualizan los gráficos interactivos.\n{search_results}"
        reply = ia_responder(msg, contexto=contexto)
        return jsonify({"reply": reply or "Lo siento, mi procesador está ocupado. Intenta de nuevo."}), 200
    except Exception as e: logger.error(f"Chat API error: {e}"); return jsonify({"reply": f"Error interno: {str(e)}"}), 500

@app.route("/api/debug/test_wa")
def api_debug_test_wa():
    tel = request.args.get("tel")
    if not tel: return jsonify({"error": "Falta parámetro tel"}), 400
    try:
        import requests as req_lib
        payload = {"messaging_product": "whatsapp", "to": str(tel), "type": "text", "text": {"body": "📡 Prueba de fuego desde Render."}}
        headers = {"Authorization": f"Bearer {Cfg.TOKEN}", "Content-Type": "application/json"}
        r = req_lib.post(f"https://graph.facebook.com/v19.0/{Cfg.PHONE_ID}/messages", json=payload, headers=headers, timeout=10)
        return jsonify({
            "status_code": r.status_code,
            "response_json": r.json(),
            "tel": tel,
            "phone_id_usado": Cfg.PHONE_ID,
            "token_inicio": Cfg.TOKEN[:15] + "..." if Cfg.TOKEN else "VACIO"
        }), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/debug/status")
def api_debug_status():
    status = {
        "bot_version": "V112-FIX",
        "wa_token_present": bool(Cfg.TOKEN),
        "wa_phone_id_present": bool(Cfg.PHONE_ID),
        "google_creds_present": bool(Cfg.CREDS),
        "sheet_id_present": bool(Cfg.SHEET_ID),
        "csv_exists": os.path.exists(Cfg.CSV),
        "rows_count": len(_get_rows()),
        "scheduler_imo_active": _scheduler_started,
        "modo_entrenamiento": os.environ.get("MODO_ENTRENAMIENTO", "false"),
        "time_lima": ahora().strftime("%Y-%m-%d %H:%M:%S")
    }
    return jsonify(status)

@app.route("/api/debug/logs")
def api_debug_logs():
    try:
        log_path = os.path.join(DATA_DIR, "bot.log")
        if not os.path.exists(log_path): return "Log file not found", 404
        with open(log_path, "r", encoding="utf-8") as f: lines = f.readlines(); return "<pre>" + "".join(lines[-200:]) + "</pre>"
    except Exception as e: return str(e), 500

_scheduler_started = False
_scheduler_lock = threading.Lock()

@app.before_request
def start_background_tasks_once():
    """Inicia el scheduler de IMOs de forma segura en la primera petición (Flask 3.0 compatible)."""
    global _scheduler_started
    if not _scheduler_started:
        with _scheduler_lock:
            if not _scheduler_started:
                logger.info("🚀 Iniciando tareas de fondo (Scheduler IMO)...")
                threading.Thread(target=_scheduler_imos, daemon=True, name="imo_scheduler").start()
                _scheduler_started = True

# ── Sincronizador CrearPSL Global ──
try:
    from sync_crearpsl import iniciar_thread as iniciar_sync_crearpsl
    iniciar_sync_crearpsl()
    logger.info("✅ Sync CrearPSL iniciado — cada 30 min")
except Exception as e: logger.warning(f"⚠ Sync CrearPSL no inició: {e}")

@app.route("/api/debug/reprocesar", methods=["POST"])
def api_reprocesar_silencios():
    try:
        from reprocesar_silencios import analizar_silencios
        silencios = analizar_silencios(horas=24)
        if not silencios: return jsonify({"msg": "No hay silencios detectados"}), 200
        
        enviados = 0
        for s in silencios:
            tel = s["tel"]
            p = perfil_crm(tel)
            wa(tel, f"🙏 Hola {p.get('nombre','amigo/a')}, tuvimos un inconveniente técnico y no pudimos responderte a tiempo. Aquí tienes nuestro menú principal para ayudarte:", "SISTEMA")
            _menu_main(tel, p)
            enviados += 1
            time.sleep(1)
            
        return jsonify({"status": "ok", "procesados": len(silencios), "enviados": enviados}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__=="__main__":
    logger.info("🚀 CPSL Torre de Control V112-FIX + MISION CRÍTICA")
    _cargar_graduados()
    app.run(host="0.0.0.0",port=int(os.environ.get("PORT",10000)),debug=False)
