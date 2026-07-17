
# =============================================================================
# AGENTE AUTÓNOMO DE CONCILIACIÓN Y ENRIQUECIMIENTO DE PARTICIPANTES
# Quantum Team - Crear Poder Sin Límites
# Versión: 2.0 | Fecha: Junio 2026
# =============================================================================

import os
import sys
import re
import warnings
import unicodedata
from datetime import datetime
from pathlib import Path

# Fix encoding for Windows console
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────
CARPETAS = [
    r"C:\Users\josem\OneDrive - QUANTUM COACHING TECHNOLOGY BVS CIA. LTDA\CREAR LIMA\PORCENTAJE ALIADOS C1",
    r"C:\Users\josem\OneDrive - QUANTUM COACHING TECHNOLOGY BVS CIA. LTDA\CREAR LIMA\PORCENTAJE ALIADOS C2",
]

SALIDA = r"C:\Users\josem\Downloads\CONCILIACION_QUANTUM"
FECHA_PROCESO = datetime.now().strftime("%Y-%m-%d %H:%M")
FECHA_ARCHIVO = datetime.now().strftime("%Y%m%d_%H%M")

# Palabras clave para clasificar estado
KW_CONFIRMADO    = ["confirmado", "c2", "c2+mj", "asegurado", "pagado", "pago", "completo", "ingreso"]
KW_INTERESADO    = ["interesado", "si quiere", "llama", "si", "positivo", "abono", "acuerdo"]
KW_NO_CONTESTA   = ["no contesta", "nc", "no atiende", "buzón", "buzon", "no responde"]
KW_NO_INTERESA   = ["no le interesa", "no interesa", "no quiere", "nni", "no interesado", "rechazó", "rechazo"]
KW_DESERTOR      = ["desertor", "desertó", "deserto", "baja", "canceló", "cancelo", "retiro", "retiró"]
KW_SEGUIMIENTO   = ["seguimiento", "en proceso", "pendiente", "por confirmar", "llamar", "contactar"]
KW_REINGRESO     = ["reingreso", "reactivar", "reactivación", "anterior", "ex alumno", "volver"]

# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def limpiar_texto(txt):
    """Normaliza texto: sin tildes, sin espacios extra, uppercase."""
    if pd.isna(txt) or txt is None:
        return ""
    txt = str(txt).strip().upper()
    txt = unicodedata.normalize("NFD", txt)
    txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
    txt = re.sub(r"\s+", " ", txt)
    return txt

def limpiar_telefono(tel):
    """Extrae solo dígitos del teléfono."""
    if pd.isna(tel) or tel is None:
        return ""
    tel = re.sub(r"\D", "", str(tel))
    return tel[-9:] if len(tel) >= 9 else tel  # últimos 9 dígitos (Perú)

def limpiar_dni(dni):
    """Normaliza DNI a 8 dígitos."""
    if pd.isna(dni) or dni is None:
        return ""
    dni = re.sub(r"\D", "", str(dni))
    return dni.zfill(8) if 6 <= len(dni) <= 9 else dni

def clasificar_estado(texto):
    """Clasifica el estado según palabras clave."""
    t = limpiar_texto(str(texto))
    if any(k in t for k in [x.upper() for x in KW_CONFIRMADO]):
        return "Confirmado"
    if any(k in t for k in [x.upper() for k in KW_INTERESADO for x in [k]]):
        return "Interesado"
    if any(k in t for k in [x.upper() for k in KW_DESERTOR for x in [k]]):
        return "Desertor"
    if any(k in t for k in [x.upper() for k in KW_NO_INTERESA for x in [k]]):
        return "No le interesa"
    if any(k in t for k in [x.upper() for k in KW_NO_CONTESTA for x in [k]]):
        return "No contesta"
    if any(k in t for k in [x.upper() for k in KW_REINGRESO for x in [k]]):
        return "Reingreso potencial"
    if any(k in t for k in [x.upper() for k in KW_SEGUIMIENTO for x in [k]]):
        return "En seguimiento"
    if t in ["C2", "C2+MJ", "ABONO", "ACUERDO", "PAGO", "PAGADO"]:
        return "Confirmado"
    return "En seguimiento"

def nivel_interes(estado):
    mapa = {
        "Confirmado": "Alto interés",
        "Interesado": "Alto interés",
        "En seguimiento": "Interés medio",
        "No contesta": "Bajo interés",
        "No le interesa": "Bajo interés",
        "Desertor": "Riesgo de abandono",
        "Reingreso potencial": "Posible reactivación",
        "Cliente activo": "Alto interés",
    }
    return mapa.get(estado, "Interés medio")

def accion_recomendada(estado):
    mapa = {
        "Confirmado": "Seguimiento semanal",
        "Interesado": "Llamar inmediatamente",
        "En seguimiento": "Llamar inmediatamente",
        "No contesta": "Enviar información",
        "No le interesa": "No contactar temporalmente",
        "Desertor": "Reactivar",
        "Reingreso potencial": "Agendar entrevista",
        "Cliente activo": "Seguimiento semanal",
    }
    return mapa.get(estado, "Seguimiento semanal")

def prioridad(estado):
    mapa = {
        "Confirmado": 2,
        "Interesado": 1,
        "En seguimiento": 1,
        "No contesta": 3,
        "No le interesa": 5,
        "Desertor": 4,
        "Reingreso potencial": 2,
        "Cliente activo": 2,
    }
    return mapa.get(estado, 3)

# ─────────────────────────────────────────────────────────────────────────────
# LECTURA Y NORMALIZACIÓN DE ARCHIVOS
# ─────────────────────────────────────────────────────────────────────────────

def detectar_columnas(df, tipo_archivo):
    """Intenta mapear columnas del DataFrame a un esquema estándar."""
    cols = {c: limpiar_texto(c) for c in df.columns}
    
    mapa = {
        "DNI": None, "NOMBRE": None, "APELLIDO": None,
        "TELEFONO": None, "IMO": None, "STATUS": None,
        "ALIADO": None, "COORDINADOR": None, "OBSERVACION": None,
        "EQUIPO": None,
    }
    
    busqueda = {
        "DNI":         ["DNI", "IDENTIFICACION", "DOCUMENTO", "RUC"],
        "NOMBRE":      ["NOMBRES", "NOMBRE COMPLETO", "NOMBRE QUE PREFIERE", "NOMBRE"],
        "APELLIDO":    ["APELLIDOS", "APELLIDO COMPLETO", "APELLIDO"],
        "TELEFONO":    ["TEL", "TELEFONO", "CELULAR", "MOVIL", "TELEFONO MOVIL"],
        "IMO":         ["IMO", "NOMBRE IMO", "IDENTIFICACION IMO", "IDENTIFICACIONIMO"],
        "STATUS":      ["STATUS", "ESTADO", "RESULTADO", "ASISTENCIA"],
        "ALIADO":      ["ALIADO", "NOMBRE ALIADO"],
        "COORDINADOR": ["COORDINADOR", "COORD", "RESPONSABLE", "USUARIO SEGUIMIENTO"],
        "OBSERVACION": ["OBSERVACION", "OBSERVACIONES", "NOTAS", "COMENTARIOS"],
        "EQUIPO":      ["EQUIPO", "EQUIPO ACTUAL", "NOMBREEQUIPO"],
    }
    
    for campo, palabras in busqueda.items():
        for col_orig, col_limpia in cols.items():
            for p in palabras:
                if p in col_limpia:
                    mapa[campo] = col_orig
                    break
            if mapa[campo]:
                break
    
    return mapa

def leer_hoja(df, archivo, hoja, tipo):
    """Extrae registros de una hoja Excel de forma flexible."""
    if df.empty or len(df) < 2:
        return []
    
    # Detectar si la primera fila es encabezado real o datos
    # Intentar con la fila 0 como encabezado por defecto
    mapa = detectar_columnas(df, tipo)
    
    registros = []
    
    for idx, row in df.iterrows():
        nombre_raw = str(row.get(mapa["NOMBRE"], "")).strip() if mapa["NOMBRE"] else ""
        apellido_raw = str(row.get(mapa["APELLIDO"], "")).strip() if mapa["APELLIDO"] else ""
        
        # Construir nombre completo
        nombre_completo = ""
        if nombre_raw and apellido_raw and nombre_raw not in ["nan", "None", ""]:
            nombre_completo = f"{apellido_raw} {nombre_raw}".strip()
        elif nombre_raw and nombre_raw not in ["nan", "None", ""]:
            nombre_completo = nombre_raw
        
        # Buscar nombre alternativo en columnas no mapeadas
        if not nombre_completo or nombre_completo == "nan":
            for col in df.columns:
                val = str(row.get(col, "")).strip()
                if len(val) > 5 and val not in ["nan", "None"] and re.search(r"[A-Za-záéíóú]{3,}", val):
                    col_l = limpiar_texto(col)
                    if "NOMBRE" in col_l or "APELLIDO" in col_l:
                        nombre_completo = val
                        break
        
        if not nombre_completo or nombre_completo in ["nan", "None", ""]:
            continue  # saltar filas sin nombre
        
        telefono = limpiar_telefono(row.get(mapa["TELEFONO"], "") if mapa["TELEFONO"] else "")
        imo      = str(row.get(mapa["IMO"], "") if mapa["IMO"] else "").strip()
        status   = str(row.get(mapa["STATUS"], "") if mapa["STATUS"] else "").strip()
        aliado   = str(row.get(mapa["ALIADO"], "") if mapa["ALIADO"] else "").strip()
        coord    = str(row.get(mapa["COORDINADOR"], "") if mapa["COORDINADOR"] else "").strip()
        obs      = str(row.get(mapa["OBSERVACION"], "") if mapa["OBSERVACION"] else "").strip()
        equipo   = str(row.get(mapa["EQUIPO"], "") if mapa["EQUIPO"] else "").strip()
        
        # Extraer equipo del nombre del archivo
        eq_match = re.search(r"EQUIPO\s*(\d+)", archivo.upper())
        if eq_match and (not equipo or equipo == "nan"):
            equipo = f"EQUIPO {eq_match.group(1)}"
        
        # Determinar capítulo desde nombre de archivo
        capitulo = "C1" if "C1" in archivo.upper() or "CAPITULO UNO" in archivo.upper() else "C2"
        
        registros.append({
            "NOMBRE_COMPLETO_RAW": nombre_completo,
            "NOMBRE_LIMPIO":       limpiar_texto(nombre_completo),
            "TELEFONO":            telefono,
            "IMO":                 limpiar_texto(imo),
            "STATUS_RAW":          status if status not in ["nan", "None"] else "",
            "ALIADO":              aliado if aliado not in ["nan", "None"] else "",
            "COORDINADOR":         coord  if coord  not in ["nan", "None"] else "",
            "OBSERVACION":         obs    if obs    not in ["nan", "None"] else "",
            "EQUIPO":              equipo if equipo not in ["nan", "None"] else "",
            "CAPITULO":            capitulo,
            "ARCHIVO_FUENTE":      archivo,
            "HOJA_FUENTE":         hoja,
            "FECHA_EXTRACCION":    FECHA_PROCESO,
        })
    
    return registros

def leer_archivo_excel(path):
    """Lee todas las hojas relevantes de un archivo Excel."""
    nombre_archivo = os.path.basename(path)
    todos_registros = []
    
    try:
        xl = pd.ExcelFile(path)
        hojas_relevantes = [
            h for h in xl.sheet_names
            if any(kw in h.upper() for kw in ["PX", "LISTADO", "STATUS", "PARTICIPANTES", "DATOS"])
            or h.upper() in ["PX", "LISTADO", "STATUS-INICIO", "LISTADO LINID"]
        ]
        
        if not hojas_relevantes:
            # Si no hay hojas relevantes por nombre, tomar la primera
            hojas_relevantes = xl.sheet_names[:2]
        
        for hoja in hojas_relevantes:
            try:
                df = pd.read_excel(path, sheet_name=hoja, header=None)
                
                # Detectar fila de encabezado (buscar fila con más texto)
                mejor_fila = 0
                mejor_score = 0
                for i in range(min(5, len(df))):
                    row = df.iloc[i]
                    score = sum(1 for v in row if isinstance(v, str) and len(v) > 2)
                    if score > mejor_score:
                        mejor_score = score
                        mejor_fila = i
                
                df.columns = df.iloc[mejor_fila]
                df = df.iloc[mejor_fila + 1:].reset_index(drop=True)
                df.columns = [str(c) for c in df.columns]
                
                tipo = "C1" if "C1" in nombre_archivo.upper() else "C2"
                registros = leer_hoja(df, nombre_archivo, hoja, tipo)
                todos_registros.extend(registros)
                
            except Exception as e:
                print(f"   ⚠ Hoja '{hoja}' en {nombre_archivo}: {e}")
    
    except Exception as e:
        print(f"   ✗ Error leyendo {nombre_archivo}: {e}")
    
    return todos_registros

# ─────────────────────────────────────────────────────────────────────────────
# CONCILIACIÓN Y DEDUPLICACIÓN
# ─────────────────────────────────────────────────────────────────────────────

def similaridad_nombres(n1, n2):
    """Similaridad simple de nombres por tokens comunes."""
    t1 = set(n1.split())
    t2 = set(n2.split())
    if not t1 or not t2:
        return 0
    comunes = t1 & t2
    return len(comunes) / max(len(t1), len(t2))

def conciliar_registros(registros):
    """Unifica registros duplicados por teléfono y nombre."""
    print(f"\n{'='*60}")
    print(f"  CONCILIACIÓN: {len(registros)} registros brutos")
    print(f"{'='*60}")
    
    # Agrupar por teléfono (clave primaria)
    grupos = {}
    sin_tel = []
    
    for r in registros:
        tel = r["TELEFONO"]
        if tel and len(tel) >= 7:
            if tel not in grupos:
                grupos[tel] = []
            grupos[tel].append(r)
        else:
            sin_tel.append(r)
    
    # Para sin teléfono, agrupar por nombre similar
    grupos_nombre = {}
    for r in sin_tel:
        nombre = r["NOMBRE_LIMPIO"]
        encontrado = False
        for clave in list(grupos_nombre.keys()):
            if similaridad_nombres(nombre, clave) >= 0.7:
                grupos_nombre[clave].append(r)
                encontrado = True
                break
        if not encontrado:
            grupos_nombre[nombre] = [r]
    
    # Combinar todos los grupos
    todos_grupos = list(grupos.values()) + list(grupos_nombre.values())
    
    # Consolidar cada grupo en un registro maestro
    maestros = []
    for grupo in todos_grupos:
        if not grupo:
            continue
        
        # Elegir el registro más completo como base
        base = max(grupo, key=lambda x: sum(1 for v in x.values() if v and v != "nan"))
        
        # Acumular estados únicos de todos los registros
        estados = []
        observaciones = []
        archivos = set()
        capitulos = set()
        aliados = set()
        coordinadores = set()
        equipos = set()
        
        for r in grupo:
            if r["STATUS_RAW"]:
                estados.append(r["STATUS_RAW"])
            if r["OBSERVACION"]:
                observaciones.append(r["OBSERVACION"])
            archivos.add(r["ARCHIVO_FUENTE"])
            capitulos.add(r["CAPITULO"])
            if r["ALIADO"]:
                aliados.add(r["ALIADO"])
            if r["COORDINADOR"]:
                coordinadores.add(r["COORDINADOR"])
            if r["EQUIPO"]:
                equipos.add(r["EQUIPO"])
        
        # Clasificar estado final
        estado_texto = " ".join(estados + observaciones)
        estado_final = clasificar_estado(estado_texto)
        
        # Si tiene C2 en archivos C2, es confirmado
        if "C2" in capitulos and any("C2" in e.upper() for e in estados):
            estado_final = "Confirmado"
        
        # Construir resumen histórico
        eventos = []
        if "C1" in capitulos:
            eventos.append("Participó en Capítulo 1")
        if "C2" in capitulos:
            eventos.append("Avanzó a Capítulo 2")
        for estado in set(estados):
            if estado not in ["nan", "None", ""]:
                eventos.append(f"Estado registrado: {estado}")
        for obs in observaciones[:3]:
            if obs not in ["nan", "None", ""]:
                eventos.append(obs)
        
        if len(grupo) > 1:
            eventos.append(f"Aparece en {len(grupo)} registros de {len(archivos)} archivo(s)")
        
        resumen = ". ".join(eventos) if eventos else "Participante registrado en el sistema."
        
        maestro = {
            "NOMBRE_COMPLETO":       base["NOMBRE_COMPLETO_RAW"],
            "NOMBRE_LIMPIO":         base["NOMBRE_LIMPIO"],
            "TELEFONO":              base["TELEFONO"],
            "IMO":                   base["IMO"],
            "EQUIPO":                " / ".join(sorted(equipos)) if equipos else base["EQUIPO"],
            "ALIADO":                " / ".join(sorted(aliados)) if aliados else base["ALIADO"],
            "COORDINADOR":           " / ".join(sorted(coordinadores)) if coordinadores else base["COORDINADOR"],
            "CAPITULOS":             " + ".join(sorted(capitulos)),
            "ESTADOS_HISTORICOS":    " | ".join(set(estados)) if estados else "",
            "ESTADO_FINAL":          estado_final,
            "NIVEL_INTERES":         nivel_interes(estado_final),
            "PRIORIDAD":             prioridad(estado_final),
            "ACCION_RECOMENDADA":    accion_recomendada(estado_final),
            "RESUMEN_HISTORICO":     resumen,
            "ARCHIVOS_FUENTE":       " | ".join(sorted(archivos)),
            "TOTAL_REGISTROS_ORIG":  len(grupo),
            "FECHA_EXTRACCION":      FECHA_PROCESO,
            "ALERTA":                "⚠ Datos contradictorios" if len(set(estados)) > 2 else "",
        }
        maestros.append(maestro)
    
    print(f"  → {len(maestros)} participantes únicos identificados")
    return maestros

# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DE REPORTES
# ─────────────────────────────────────────────────────────────────────────────

def generar_base_maestra(maestros, salida):
    """Genera BASE_MAESTRA_CONSOLIDADA.xlsx"""
    df = pd.DataFrame(maestros)
    
    columnas_orden = [
        "NOMBRE_COMPLETO", "TELEFONO", "IMO", "EQUIPO",
        "COORDINADOR", "ALIADO", "CAPITULOS",
        "ESTADOS_HISTORICOS", "ESTADO_FINAL",
        "NIVEL_INTERES", "PRIORIDAD", "ACCION_RECOMENDADA",
        "RESUMEN_HISTORICO", "ALERTA",
        "ARCHIVOS_FUENTE", "TOTAL_REGISTROS_ORIG", "FECHA_EXTRACCION",
    ]
    for c in columnas_orden:
        if c not in df.columns:
            df[c] = ""
    
    df = df[columnas_orden].sort_values(["PRIORIDAD", "NOMBRE_COMPLETO"])
    
    path = os.path.join(salida, f"BASE_MAESTRA_CONSOLIDADA_{FECHA_ARCHIVO}.xlsx")
    
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="BASE_MAESTRA")
        
        ws = writer.sheets["BASE_MAESTRA"]
        
        # Colores por estado
        from openpyxl.styles import PatternFill, Font, Alignment
        colores = {
            "Confirmado":         "C6EFCE",
            "Interesado":         "FFEB9C",
            "En seguimiento":     "BDD7EE",
            "No contesta":        "D9D9D9",
            "No le interesa":     "FFC7CE",
            "Desertor":           "FF0000",
            "Reingreso potencial":"E2EFDA",
            "Cliente activo":     "92D050",
        }
        
        # Encabezado
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Datos
        col_estado = list(df.columns).index("ESTADO_FINAL") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            estado_cell = row[col_estado - 1]
            estado_val = str(estado_cell.value or "")
            color = colores.get(estado_val, "FFFFFF")
            fill = PatternFill("solid", fgColor=color)
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Ajustar anchos
        anchos = {"NOMBRE_COMPLETO": 35, "RESUMEN_HISTORICO": 60, "ARCHIVOS_FUENTE": 40}
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = anchos.get(col_name, 18)
        
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    
    print(f"\n  ✅ BASE MAESTRA → {path}")
    return df

def generar_reporte_ejecutivo(maestros, salida):
    """Genera REPORTE_EJECUTIVO.xlsx"""
    df = pd.DataFrame(maestros)
    
    conteos = df["ESTADO_FINAL"].value_counts().to_dict()
    total = len(df)
    
    indicadores = {
        "Total participantes únicos":   total,
        "Confirmados":                  conteos.get("Confirmado", 0),
        "Interesados":                  conteos.get("Interesado", 0),
        "En seguimiento":               conteos.get("En seguimiento", 0),
        "No contestan":                 conteos.get("No contesta", 0),
        "No interesados":               conteos.get("No le interesa", 0),
        "Desertores":                   conteos.get("Desertor", 0),
        "Reingreso potencial":          conteos.get("Reingreso potencial", 0),
        "Alta prioridad (P1)":          len(df[df["PRIORIDAD"] == 1]),
        "Con trayectoria C1+C2":        len(df[df["CAPITULOS"].str.contains("C1") & df["CAPITULOS"].str.contains("C2")]),
        "Solo C1":                      len(df[df["CAPITULOS"] == "C1"]),
        "Solo C2":                      len(df[df["CAPITULOS"] == "C2"]),
        "Con alertas de inconsistencia":len(df[df["ALERTA"] != ""]),
        "Tasa confirmación %":          f"{conteos.get('Confirmado', 0) / total * 100:.1f}%" if total > 0 else "0%",
        "Fecha de proceso":             FECHA_PROCESO,
    }
    
    df_ind = pd.DataFrame(list(indicadores.items()), columns=["INDICADOR", "VALOR"])
    
    # Por capítulo
    df_cap = df.groupby(["CAPITULOS", "ESTADO_FINAL"]).size().reset_index(name="CANTIDAD")
    
    # Por equipo
    df_eq = df.groupby(["EQUIPO", "ESTADO_FINAL"]).size().reset_index(name="CANTIDAD")
    
    path = os.path.join(salida, f"REPORTE_EJECUTIVO_{FECHA_ARCHIVO}.xlsx")
    
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_ind.to_excel(writer, index=False, sheet_name="INDICADORES")
        df_cap.to_excel(writer, index=False, sheet_name="POR_CAPITULO")
        df_eq.to_excel(writer, index=False, sheet_name="POR_EQUIPO")
        
        from openpyxl.styles import PatternFill, Font, Alignment
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    
    print(f"  ✅ REPORTE EJECUTIVO → {path}")
    return indicadores

def generar_ranking_coordinadores(maestros, salida):
    """Genera RANKING_COORDINADORES.xlsx"""
    df = pd.DataFrame(maestros)
    
    if df["COORDINADOR"].str.strip().eq("").all():
        df["COORDINADOR"] = "Sin asignar"
    
    df["COORDINADOR"] = df["COORDINADOR"].replace("", "Sin asignar").fillna("Sin asignar")
    
    # Explotar coordinadores múltiples
    df_exp = df.copy()
    df_exp["COORDINADOR"] = df_exp["COORDINADOR"].str.split(" / ")
    df_exp = df_exp.explode("COORDINADOR")
    df_exp["COORDINADOR"] = df_exp["COORDINADOR"].str.strip()
    
    resumen = df_exp.groupby("COORDINADOR").agg(
        Total_Asignados  = ("NOMBRE_COMPLETO", "count"),
        Confirmados      = ("ESTADO_FINAL", lambda x: (x == "Confirmado").sum()),
        Interesados      = ("ESTADO_FINAL", lambda x: (x == "Interesado").sum()),
        Desertores       = ("ESTADO_FINAL", lambda x: (x == "Desertor").sum()),
        Reactivables     = ("ESTADO_FINAL", lambda x: (x == "Reingreso potencial").sum()),
        No_contestan     = ("ESTADO_FINAL", lambda x: (x == "No contesta").sum()),
    ).reset_index()
    
    resumen["Tasa_Conversion_%"] = (
        resumen["Confirmados"] / resumen["Total_Asignados"] * 100
    ).round(1).astype(str) + "%"
    
    resumen = resumen.sort_values("Confirmados", ascending=False)
    
    path = os.path.join(salida, f"RANKING_COORDINADORES_{FECHA_ARCHIVO}.xlsx")
    
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        resumen.to_excel(writer, index=False, sheet_name="RANKING")
        
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = writer.sheets["RANKING"]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center")
        
        # Colorear top 3
        verde = PatternFill("solid", fgColor="C6EFCE")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(4, ws.max_row)), 1):
            if i <= 3:
                for cell in row:
                    cell.fill = verde
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 25
    
    print(f"  ✅ RANKING COORDINADORES → {path}")

# ─────────────────────────────────────────────────────────────────────────────
# EJECUCIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*60)
    print("  🤖 AGENTE AUTÓNOMO DE CONCILIACIÓN - QUANTUM TEAM")
    print(f"  📅 {FECHA_PROCESO}")
    print("="*60)
    
    # Crear carpeta de salida
    os.makedirs(SALIDA, exist_ok=True)
    
    # PASO 1: Leer todos los archivos
    todos_registros = []
    
    for carpeta in CARPETAS:
        if not os.path.exists(carpeta):
            print(f"\n  ⚠ Carpeta no encontrada: {carpeta}")
            continue
        
        tipo = "C1" if "C1" in carpeta else "C2"
        archivos = [f for f in os.listdir(carpeta) if f.endswith((".xlsx", ".xls", ".csv"))]
        
        print(f"\n📁 {os.path.basename(carpeta)} — {len(archivos)} archivos")
        
        for archivo in sorted(archivos):
            path_completo = os.path.join(carpeta, archivo)
            print(f"   📄 Leyendo: {archivo}", end=" ... ")
            
            if archivo.endswith(".csv"):
                try:
                    df = pd.read_csv(path_completo, encoding="latin-1", on_bad_lines="skip")
                    regs = leer_hoja(df, archivo, "CSV", tipo)
                    todos_registros.extend(regs)
                    print(f"✓ {len(regs)} registros")
                except Exception as e:
                    print(f"✗ {e}")
            else:
                regs = leer_archivo_excel(path_completo)
                todos_registros.extend(regs)
                print(f"✓ {len(regs)} registros")
    
    if not todos_registros:
        print("\n  ✗ No se encontraron registros. Verificar carpetas y archivos.")
        return
    
    print(f"\n  📊 Total registros brutos: {len(todos_registros)}")
    
    # PASO 2: Conciliación
    maestros = conciliar_registros(todos_registros)
    
    # PASO 3: Generar archivos de salida
    print(f"\n{'='*60}")
    print("  📤 GENERANDO ARCHIVOS DE SALIDA")
    print(f"{'='*60}")
    
    df_maestra = generar_base_maestra(maestros, SALIDA)
    indicadores = generar_reporte_ejecutivo(maestros, SALIDA)
    generar_ranking_coordinadores(maestros, SALIDA)
    
    # PASO 4: Resumen final en consola
    print(f"\n{'='*60}")
    print("  📋 RESUMEN EJECUTIVO")
    print(f"{'='*60}")
    for k, v in indicadores.items():
        print(f"  {k:<40} {v}")
    
    print(f"\n  📁 Archivos guardados en: {SALIDA}")
    print("\n  ✅ PROCESO COMPLETADO\n")

if __name__ == "__main__":
    main()
