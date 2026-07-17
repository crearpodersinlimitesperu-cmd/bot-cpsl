import os
import re
import json
import argparse
import logging
import hashlib
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
import pdfplumber
from bs4 import BeautifulSoup
from dateutil import parser as date_parser

# Setup head-less matplotlib for non-GUI environments (like Cloud Shell)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# Timezone configurations
LIMA_TZ = ZoneInfo("America/Lima")

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("agente_financiero.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)

PARSING_ERRORS_LOG = "errores_parseo.log"

def log_parsing_error(message_id, subject, body, error_detail):
    """Logs unparseable emails to errores_parseo.log without interrupting the flow."""
    with open(PARSING_ERRORS_LOG, "a", encoding="utf-8") as f:
        f.write(f"--- PARSING ERROR AT {datetime.now(LIMA_TZ).isoformat()} ---\n")
        f.write(f"Gmail Message ID: {message_id}\n")
        f.write(f"Subject: {subject}\n")
        f.write(f"Error Detail: {error_detail}\n")
        f.write(f"Body snippet:\n{body[:500]}\n")
        f.write("-" * 80 + "\n")

# ============================================================
# 1. APRENDIZAJE INICIAL (LEARNING & DATA EXTRACTION FROM FILES)
# ============================================================

def classify_category(text):
    t = text.lower()
    if "worx" in t or "oficina" in t:
        return "Alquiler Oficinas"
    if "benavides" in t or "berlin" in t or "alojamiento" in t:
        return "Aloj. Coordinadores"
    if "jose antonio" in t or "salon" in t or "deluxe" in t or "hilton" in t or "salón" in t:
        return "Alquiler Salones"
    if "honorarios" in t or "entrenador" in t or "gomez" in t or "torron" in t or "sueldo entrenador" in t:
        return "Honorarios Entren."
    if "viatico" in t or "viáticos" in t or "viatico" in t:
        return "Viático Entren."
    if "hospedaje" in t or "hotel" in t:
        return "Hospedaje Entren."
    if "latam" in t or "vuelo" in t or "pasaje" in t or "aéreo" in t:
        return "Pasajes Aéreos"
    if "transporte" in t or "movilidad" in t or "uber" in t or "taxi" in t:
        return "Transporte Entren."
    if "movistar" in t or "luz del sur" in t or "luz" in t or "entel" in t or "agua" in t or "sedapal" in t or "servicio" in t:
        return "Servicios Básicos"
    if "contable" in t or "legal" in t or "estudio" in t or "asesor" in t:
        return "Contables/Legal/Log."
    if "suministro" in t or "snack" in t or "bodega" in t or "fruto" in t or "bebida" in t or "hidratacion" in t or "agua para entrenamiento" in t or "aguas" in t:
        return "Suministros Entrenamientos"
    if "afp" in t or "pension" in t or "seguridad social" in t or "iess" in t:
        return "Seguridad Social/IESS"
    if "igv" in t or "impuesto" in t or "pdt" in t or "iva" in t:
        return "Impuestos IVA"
    if "itf" in t or "comision" in t or "mantenimiento de cuenta" in t:
        return "Comisiones Bancarias"
    if "nomina" in t or "salario" in t or "sueldo" in t or "personal" in t:
        return "Personal CREAR"
    if "inscripcion" in t or "venta" in t or "cobro" in t or "ingreso" in t:
        return "Inscripciones"
    return "Gastos Varios"

def extract_date_from_filename(filename):
    months_es = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
        "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
        "dicembre": 12
    }
    
    fn = filename.lower()
    # Pattern 1: DD - DD MONTH YYYY
    m1 = re.search(r"(\d{1,2})\s*-\s*(\d{1,2})\s+([a-zñáéíóú]+)\s+(\d{4})", fn)
    if m1:
        try:
            day = int(m1.group(2))
            month_name = m1.group(3)
            year = int(m1.group(4))
            month = months_es.get(month_name, 1)
            return f"{year:04d}-{month:02d}-{day:02d}"
        except:
            pass
            
    # Pattern 2: DD MONTH1 - DD MONTH2 YYYY
    m2 = re.search(r"(\d{1,2})\s+([a-zñáéíóú]+)\s*-\s*(\d{1,2})\s+([a-zñáéíóú]+)\s+(\d{4})", fn)
    if m2:
        try:
            day = int(m2.group(3))
            month_name = m2.group(4)
            year = int(m2.group(5))
            month = months_es.get(month_name, 1)
            return f"{year:04d}-{month:02d}-{day:02d}"
        except:
            pass
            
    # Pattern 3: YYYY
    m3 = re.search(r"\b(202\d)\b", fn)
    if m3:
        return f"{m3.group(1)}-01-01"
        
    return datetime.now().strftime("%Y-%m-%d")

def learn_file_structure(file_path):
    """Analyzes a file structure (Excel, CSV, PDF, TXT) and returns detected column/data mapping."""
    ext = os.path.splitext(file_path)[1].lower()
    mapping = {
        "file_name_pattern": os.path.basename(file_path),
        "extension": ext,
        "date_col": None,
        "concept_col": None,
        "amount_col": None,
        "type_col": None,
        "category_col": None,
        "is_tabular": False,
        "text_patterns": {}
    }

    date_terms = re.compile(r"fecha|date|dia|transac", re.IGNORECASE)
    concept_terms = re.compile(r"concept|descrip|detall|proveed|glosa|remit|proveedor", re.IGNORECASE)
    amount_terms = re.compile(r"monto|import|cantid|total|amount|valor|s/|\$|soles|dolares", re.IGNORECASE)
    type_terms = re.compile(r"tipo|operac|movim|type", re.IGNORECASE)
    category_terms = re.compile(r"categ|rubro|program", re.IGNORECASE)

    if ext in [".csv", ".xlsx", ".xls"]:
        mapping["is_tabular"] = True
        try:
            if ext == ".csv":
                df = pd.read_csv(file_path, nrows=5)
            else:
                df = pd.read_excel(file_path, nrows=5)
            
            headers = list(df.columns)
            for col in headers:
                col_str = str(col)
                if date_terms.search(col_str):
                    mapping["date_col"] = col_str
                elif concept_terms.search(col_str):
                    mapping["concept_col"] = col_str
                elif amount_terms.search(col_str):
                    mapping["amount_col"] = col_str
                elif type_terms.search(col_str):
                    mapping["type_col"] = col_str
                elif category_terms.search(col_str):
                    mapping["category_col"] = col_str
        except Exception as e:
            logging.error(f"Error reading tabular layout for {file_path}: {e}")
            
    elif ext in [".txt", ".pdf"]:
        content = ""
        try:
            if ext == ".txt":
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()
            else:
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        content += page.extract_text() or ""
            
            lines = content.splitlines()
            for line in lines:
                if ":" in line:
                    key, val = line.split(":", 1)
                    key_clean = key.strip().lower()
                    if "fecha" in key_clean or "date" in key_clean:
                        mapping["text_patterns"]["date"] = rf"{key.strip()}:\s*(.*)"
                    elif "concept" in key_clean or "detall" in key_clean:
                        mapping["text_patterns"]["concept"] = rf"{key.strip()}:\s*(.*)"
                    elif "monto" in key_clean or "total" in key_clean or "pagar" in key_clean:
                        mapping["text_patterns"]["amount"] = rf"{key.strip()}:\s*(.*)"
                    elif "tipo" in key_clean or "movim" in key_clean:
                        mapping["text_patterns"]["type"] = rf"{key.strip()}:\s*(.*)"
                    elif "categ" in key_clean or "rubro" in key_clean:
                        mapping["text_patterns"]["category"] = rf"{key.strip()}:\s*(.*)"
        except Exception as e:
            logging.error(f"Error reading text layout for {file_path}: {e}")

    return mapping

def parse_date_value(raw_val):
    if isinstance(raw_val, pd.Timestamp):
        return raw_val.strftime("%Y-%m-%d")
    if isinstance(raw_val, datetime):
        return raw_val.strftime("%Y-%m-%d")
    
    raw_str = str(raw_val).strip()
    # Check if it looks like an ISO date (e.g. YYYY-MM-DD or YYYY-MM-DD HH:MM:SS)
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", raw_str)
    if iso_match:
        return f"{iso_match.group(1)}-{iso_match.group(2)}-{iso_match.group(3)}"
        
    months_es = {
        "enero": "1", "febrero": "2", "marzo": "3", "abril": "4", "mayo": "5", "junio": "6",
        "julio": "7", "agosto": "8", "septiembre": "9", "octubre": "10", "noviembre": "11", "diciembre": "12"
    }
    date_clean = raw_str.lower()
    for mes_name, mes_num in months_es.items():
        if mes_name in date_clean:
            date_clean = date_clean.replace(f"de {mes_name} de", f"/{mes_num}/").replace(" de ", "")
            
    try:
        parsed_dt = date_parser.parse(date_clean, dayfirst=True)
        return parsed_dt.strftime("%Y-%m-%d")
    except Exception:
        match = re.search(r"\b\d{4}-\d{2}-\d{2}\b|\b\d{2}/\d{2}/\d{4}\b", raw_str)
        if match:
            try:
                parsed_dt = date_parser.parse(match.group(0), dayfirst=True)
                return parsed_dt.strftime("%Y-%m-%d")
            except Exception:
                pass
        return raw_str

def parse_numeric_value(raw_val, is_usd_hint=False):
    if isinstance(raw_val, (int, float)):
        return float(raw_val), is_usd_hint
        
    raw_str = str(raw_val).strip()
    is_usd = is_usd_hint or any(sym in raw_str.upper() for sym in ["USD", "US$", "$"])
    
    clean_str = re.sub(r"[^\d.,-]", "", raw_str)
    if "," in clean_str and "." in clean_str:
        clean_str = clean_str.replace(",", "")
    elif "," in clean_str and "." not in clean_str:
        parts = clean_str.split(",")
        if len(parts) == 2 and len(parts[1]) == 2:
            clean_str = clean_str.replace(",", ".")
        else:
            clean_str = clean_str.replace(",", "")
            
    try:
        return float(clean_str), is_usd
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", clean_str.replace(",", ""))
        if match:
            return float(match.group(0)), is_usd
        raise

def parse_text_patterns(file_path, text, patterns):
    """Extracts a transaction dict from raw text using regex patterns."""
    tx = {
        "id": os.path.basename(file_path),
        "fecha": None,
        "concepto": "Archivo " + os.path.basename(file_path),
        "monto_original": 0.0,
        "moneda": "S/",
        "monto_soles": 0.0,
        "tipo": "Egreso",
        "categoria": "Gastos Varios",
        "fuente": "Local: " + os.path.basename(file_path)
    }
    
    for key, pattern in patterns.items():
        if not pattern:
            continue
        match = re.search(pattern, text)
        if match:
            val = match.group(1).strip()
            if key == "date":
                tx["fecha"] = parse_date_value(val)
            elif key == "concept":
                tx["concepto"] = val
            elif key == "amount":
                try:
                    num_val, is_usd = parse_numeric_value(val)
                    tx["monto_original"] = num_val
                    if is_usd:
                        tx["moneda"] = "USD"
                        tx["monto_soles"] = round(num_val * 3.75, 2)
                    else:
                        tx["moneda"] = "S/"
                        tx["monto_soles"] = num_val
                except Exception:
                    pass
            elif key == "type":
                tx["tipo"] = "Ingreso" if any(kw in val.lower() for kw in ["ingreso", "abono", "deposito"]) else "Egreso"
            elif key == "category":
                tx["categoria"] = val
                
    if not tx.get("fecha"):
        tx["fecha"] = extract_date_from_filename(os.path.basename(file_path))
    return tx

def parse_tabular_file(file_path, mapping):
    """Parses rows of a tabular (CSV/Excel) file into transaction dicts based on the mappings."""
    ext = mapping["extension"]
    if ext == ".csv":
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)
        
    txs = []
    for idx, row in df.iterrows():
        # Skip empty rows or summary/totals rows
        row_str = str(row.to_dict()).lower()
        if "total" in row_str or "resumen" in row_str:
            continue
            
        row_id = hashlib.md5(f"{os.path.basename(file_path)}_{idx}".encode()).hexdigest()[:12]
        
        tx = {
            "id": row_id,
            "fecha": None,
            "concepto": f"Transaccion de {os.path.basename(file_path)}",
            "monto_original": 0.0,
            "moneda": "S/",
            "monto_soles": 0.0,
            "tipo": "Egreso",
            "categoria": "Gastos Varios",
            "fuente": "Local: " + os.path.basename(file_path)
        }
        
        # Concept
        if mapping.get("concept_col") and mapping["concept_col"] in row:
            val_concept = str(row[mapping["concept_col"]])
            if not val_concept or val_concept.lower().strip() in ["nan", "", "none", "totals"]:
                continue
            tx["concepto"] = val_concept
            
        # Amount
        if mapping.get("amount_col") and mapping["amount_col"] in row:
            raw_val = row[mapping["amount_col"]]
            if pd.isna(raw_val) or str(raw_val).strip() == "" or str(raw_val).strip().lower() == "nan":
                continue
            try:
                is_usd_col = "dolar" in str(mapping["amount_col"]).lower()
                num_val, is_usd = parse_numeric_value(raw_val, is_usd_col)
                if num_val <= 0:
                    continue
                tx["monto_original"] = num_val
                if is_usd:
                    tx["moneda"] = "USD"
                    tx["monto_soles"] = round(num_val * 3.75, 2)
                else:
                    tx["moneda"] = "S/"
                    tx["monto_soles"] = num_val
            except Exception:
                continue
        else:
            continue  # No amount column value, skip
            
        # Date
        if mapping.get("date_col") and mapping["date_col"] in row and not pd.isna(row[mapping["date_col"]]):
            raw_val = row[mapping["date_col"]]
            tx["fecha"] = parse_date_value(raw_val)
        else:
            tx["fecha"] = extract_date_from_filename(os.path.basename(file_path))
                
        # Type
        if mapping.get("type_col") and mapping["type_col"] in row:
            val = str(row[mapping["type_col"]]).lower()
            tx["tipo"] = "Ingreso" if any(kw in val for kw in ["ingreso", "abono", "deposito"]) else "Egreso"
            
        # Category
        if mapping.get("category_col") and mapping["category_col"] in row and not pd.isna(row[mapping["category_col"]]):
            tx["categoria"] = str(row[mapping["category_col"]])
        else:
            # Fallback to heuristics classification
            text_to_classify = (str(tx["concepto"]) + " " + str(row.get("Proveedor", "")) + " " + str(row.get("Detalle", ""))).lower()
            tx["categoria"] = classify_category(text_to_classify)
            
        txs.append(tx)
    return txs

def run_learning_engine(directory="./datos_financieros"):
    """Scans directory, saves mappings in maestro_estructura.json, and loads the data into budget master."""
    if not os.path.exists(directory):
        logging.warning(f"Directory '{directory}' does not exist. Creating it.")
        os.makedirs(directory, exist_ok=True)
        return

    schema_file = "maestro_estructura.json"
    if os.path.exists(schema_file):
        try:
            with open(schema_file, "r", encoding="utf-8") as f:
                mappings = json.load(f)
        except Exception:
            mappings = {}
    else:
        mappings = {}

    logging.info(f"Scanning '{directory}' for structural learning...")
    all_extracted_txs = []
    
    existing_ids = set()
    master_path = "presupuesto_maestro.xlsx"
    if os.path.exists(master_path):
        try:
            wb = openpyxl.load_workbook(master_path, read_only=True)
            if "Movimientos" in wb.sheetnames:
                ws_mov = wb["Movimientos"]
                for r in range(2, ws_mov.max_row + 1):
                    val = ws_mov.cell(row=r, column=1).value
                    if val:
                        existing_ids.add(str(val))
            wb.close()
        except Exception as e:
            logging.warning(f"Could not load existing IDs from budget master: {e}")
            
    for root, dirs, files in os.walk(directory):
        for f in files:
            file_path = os.path.join(root, f)
            if f.startswith("~$") or f.startswith("."):
                continue
            if f in existing_ids:
                if f not in mappings:
                    mappings[f] = learn_file_structure(file_path)
                continue
            logging.info(f"Learning structure of: {f}")
            structure = learn_file_structure(file_path)
            mappings[f] = structure
            
            # Extract data from the file based on the mappings
            try:
                if structure["is_tabular"]:
                    txs = parse_tabular_file(file_path, structure)
                    all_extracted_txs.extend(txs)
                    logging.info(f"Extracted {len(txs)} transactions from {f}")
                else:
                    # Read text content
                    ext = structure["extension"]
                    content = ""
                    if ext == ".txt":
                        with open(file_path, "r", encoding="utf-8", errors="ignore") as f_in:
                            content = f_in.read()
                    elif ext == ".pdf":
                        with pdfplumber.open(file_path) as pdf:
                            for page in pdf.pages:
                                content += page.extract_text() or ""
                    if content:
                        tx = parse_text_patterns(file_path, content, structure["text_patterns"])
                        all_extracted_txs.append(tx)
                        logging.info(f"Extracted transaction from {f}")
            except Exception as e:
                logging.error(f"Error extracting data from {f}: {e}")

    with open(schema_file, "w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=4, ensure_ascii=False)
    logging.info(f"Schema updated successfully in '{schema_file}'.")
    
    # Save these local transactions to the master budget sheet
    if all_extracted_txs:
        update_budget_master(all_extracted_txs)

# ============================================================
# 2. ACCESO A GMAIL (GMAIL OAUTH2 & PARSER)
# ============================================================

def get_gmail_service():
    """Initializes Google OAuth2 flow and returns Gmail API service client."""
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from google.auth.transport.requests import Request

    SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
    creds = None
    
    if os.path.exists('token.json'):
        try:
            creds = Credentials.from_authorized_user_file('token.json', SCOPES)
        except Exception as e:
            logging.warning(f"Error reading token.json: {e}")

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                creds = None
        
        if not creds:
            if not os.path.exists('credentials.json'):
                raise FileNotFoundError(
                    "Archivo 'credentials.json' no encontrado. Por favor descárgalo "
                    "desde Google Cloud Console e insértalo en este directorio."
                )
            
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            try:
                creds = flow.run_local_server(port=8080, open_browser=False)
            except Exception as e:
                logging.warning(f"Error running local server on port 8080: {e}. Trying random port.")
                creds = flow.run_local_server(open_browser=False)
                
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)

def parse_email_to_transaction(msg_id, subject, body, sender):
    """Heuristic parsing of BCP/provider email body into structured transaction data."""
    tx = {
        "id": msg_id,
        "fecha": None,
        "concepto": subject,
        "monto_original": 0.0,
        "moneda": "S/",
        "monto_soles": 0.0,
        "tipo": "Egreso",
        "categoria": "Gastos Varios",
        "fuente": f"Gmail ({sender})"
    }

    body_clean = body.replace("\r", "\n")
    
    # --- 1. Date Extraction ---
    date_patterns = [
        r"\b\d{4}-\d{2}-\d{2}\b",
        r"\b\d{2}/\d{2}/\d{4}\b",
        r"\b\d{1,2}\s+de\s+(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de\s+\d{4}\b"
    ]
    
    found_date = None
    for pattern in date_patterns:
        match = re.search(pattern, body_clean, re.IGNORECASE)
        if match:
            found_date = match.group(0)
            break
            
    if found_date:
        try:
            date_clean = found_date
            months_es = {
                "enero": "1", "febrero": "2", "marzo": "3", "abril": "4", "mayo": "5", "junio": "6",
                "julio": "7", "agosto": "8", "septiembre": "9", "octubre": "10", "noviembre": "11", "diciembre": "12"
            }
            for mes_name, mes_num in months_es.items():
                if mes_name in date_clean.lower():
                    date_clean = date_clean.lower().replace(f"de {mes_name} de", f"/{mes_num}/").replace(" de ", "")
            parsed_dt = date_parser.parse(date_clean, dayfirst=True)
            tx["fecha"] = parsed_dt.replace(tzinfo=LIMA_TZ).strftime("%Y-%m-%d")
        except Exception:
            tx["fecha"] = datetime.now(LIMA_TZ).strftime("%Y-%m-%d")
    else:
        tx["fecha"] = datetime.now(LIMA_TZ).strftime("%Y-%m-%d")

    # --- 2. Amount and Currency Extraction ---
    amount_pattern = re.compile(
        r"(?:(S/|S/\.|US\$|\$|USD)\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?))|"
        r"(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*(S/|S/\.|US\$|\$|USD)",
        re.IGNORECASE
    )
    
    matches = amount_pattern.findall(body_clean)
    if matches:
        first_match = matches[0]
        if first_match[1]:
            sym = first_match[0]
            val_str = first_match[1]
        else:
            sym = first_match[3]
            val_str = first_match[2]
            
        try:
            val = float(val_str.replace(",", ""))
            tx["monto_original"] = val
            if sym.upper() in ["USD", "US$", "$"]:
                tx["moneda"] = "USD"
                tx["monto_soles"] = round(val * 3.75, 2)
            else:
                tx["moneda"] = "S/"
                tx["monto_soles"] = val
        except ValueError:
            pass

    # --- 3. Transaction Type (Ingreso/Egreso) ---
    lower_body = body_clean.lower()
    ingreso_keywords = ["abono", "ingreso", "deposito", "transferencia recibida", "recibiste", "credito"]
    if any(kw in lower_body for kw in ingreso_keywords) or "ingreso" in subject.lower():
        tx["tipo"] = "Ingreso"
    else:
        tx["tipo"] = "Egreso"

    # --- 4. Category Classification (Heuristic Rules) ---
    tx["categoria"] = classify_category(body_clean + " " + subject)

    if tx["monto_soles"] <= 0:
        raise ValueError(f"Transaction amount is zero or negative: {tx['monto_soles']}")

    return tx

def fetch_gmail_transactions(days=7):
    """Queries Gmail for messages in the last X days and parses them."""
    try:
        service = get_gmail_service()
    except Exception as e:
        logging.error(f"Failed to authenticate with Gmail API: {e}")
        return []

    query = 'subject:("estado de cuenta" OR "movimiento" OR "factura" OR "pago" OR "boleta") OR from:bcp.com.pe'
    date_limit = (datetime.now(LIMA_TZ) - timedelta(days=days)).strftime("%Y/%m/%d")
    query += f" after:{date_limit}"

    logging.info(f"Searching Gmail with query: '{query}'")
    
    try:
        results = service.users().messages().list(userId='me', q=query).execute()
        messages = results.get('messages', [])
    except Exception as e:
        logging.error(f"Error listing messages: {e}")
        return []

    parsed_txs = []
    logging.info(f"Found {len(messages)} potential financial emails. Parsing details...")
    
    for msg in messages:
        msg_id = msg['id']
        try:
            msg_detail = service.users().messages().get(userId='me', id=msg_id, format='full').execute()
            
            headers = msg_detail.get('payload', {}).get('headers', [])
            subject = ""
            sender = ""
            for h in headers:
                if h['name'].lower() == 'subject':
                    subject = h['value']
                elif h['name'].lower() == 'from':
                    sender = h['value']
            
            body = ""
            payload = msg_detail.get('payload', {})
            
            def extract_body(part):
                if part.get('mimeType') == 'text/plain':
                    import base64
                    data = part.get('body', {}).get('data', '')
                    return base64.urlsafe_b64decode(data.encode('ASCII')).decode('utf-8', errors='ignore')
                elif part.get('mimeType') == 'text/html':
                    import base64
                    data = part.get('body', {}).get('data', '')
                    html = base64.urlsafe_b64decode(data.encode('ASCII')).decode('utf-8', errors='ignore')
                    soup = BeautifulSoup(html, "html.parser")
                    return soup.get_text(separator="\n")
                return ""
            
            parts = payload.get('parts', [])
            if parts:
                for part in parts:
                    body += extract_body(part)
            else:
                body = extract_body(payload)

            if not body:
                body = msg_detail.get('snippet', '')

            tx = parse_email_to_transaction(msg_id, subject, body, sender)
            parsed_txs.append(tx)
            
        except Exception as e:
            log_parsing_error(msg_id, subject, body, str(e))
            logging.warning(f"Message {msg_id} could not be parsed: {e}. Logged in {PARSING_ERRORS_LOG}.")

    return parsed_txs

# ============================================================
# 3. ACTUALIZACIÓN (BUDGET UPDATER & VALIDATION)
# ============================================================

def initialize_master_budget(path="presupuesto_maestro.xlsx"):
    """Initializes budget master workbook structure with default configuration."""
    wb = openpyxl.Workbook()
    
    # Tab 1: Movimientos
    ws_mov = wb.active
    ws_mov.title = "Movimientos"
    headers_mov = [
        "ID Movimiento", "Fecha", "Concepto", "Monto Original", 
        "Moneda", "Monto Soles", "Tipo", "Categoria", "Fuente"
    ]
    ws_mov.append(headers_mov)
    
    fill_header = PatternFill('solid', fgColor='FF1F3864')
    font_header = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
    for col_idx in range(1, len(headers_mov) + 1):
        cell = ws_mov.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_mov.row_dimensions[1].height = 24

    # Tab 2: Presupuestos (limits)
    ws_pres = wb.create_sheet("Presupuestos")
    headers_pres = ["Categoria", "Presupuesto Limite"]
    ws_pres.append(headers_pres)
    for col_idx in range(1, len(headers_pres) + 1):
        cell = ws_pres.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        ws_pres.column_dimensions[get_column_letter(col_idx)].width = 24
    
    default_limits = [
        ("Alquiler Oficinas", 5000.00),
        ("Aloj. Coordinadores", 6500.00),
        ("Alquiler Salones", 115000.00),
        ("Honorarios Entren.", 70000.00),
        ("Viático Entren.", 3000.00),
        ("Hospedaje Entren.", 9000.00),
        ("Pasajes Aéreos", 15000.00),
        ("Transporte Entren.", 1000.00),
        ("Servicios Básicos", 1500.00),
        ("Contables/Legal/Log.", 5000.00),
        ("Suministros Entrenamientos", 6000.00),
        ("Seguridad Social/IESS", 2500.00),
        ("Impuestos IVA", 35000.00),
        ("Comisiones Bancarias", 500.00),
        ("Personal CREAR", 30000.00),
        ("Gastos Varios", 2000.00)
    ]
    for cat, val in default_limits:
        ws_pres.append([cat, val])
        
    wb.save(path)
    logging.info(f"Created new budget master template at '{path}'.")

def update_budget_master(new_txs, path="presupuesto_maestro.xlsx"):
    """Appends new transactions to budget master workbook if not already present."""
    if not os.path.exists(path):
        initialize_master_budget(path)
        
    wb = openpyxl.load_workbook(path)
    ws_mov = wb["Movimientos"]
    ws_pres = wb["Presupuestos"]
    
    existing_ids = set()
    for row in range(2, ws_mov.max_row + 1):
        id_val = ws_mov.cell(row=row, column=1).value
        if id_val is not None:
            existing_ids.add(str(id_val))

    limits = {}
    for row in range(2, ws_pres.max_row + 1):
        cat = ws_pres.cell(row=row, column=1).value
        val = ws_pres.cell(row=row, column=2).value
        if cat is not None and val is not None:
            limits[cat] = float(val)

    rows_data = []
    for r in range(2, ws_mov.max_row + 1):
        rows_data.append({
            "fecha": ws_mov.cell(row=r, column=2).value,
            "monto_soles": ws_mov.cell(row=r, column=6).value,
            "tipo": ws_mov.cell(row=r, column=7).value,
            "categoria": ws_mov.cell(row=r, column=8).value,
        })
    df_existing = pd.DataFrame(rows_data)

    appended_count = 0
    
    for tx in new_txs:
        if str(tx["id"]) in existing_ids:
            continue
            
        if not tx.get("fecha"):
            logging.error(f"Validation failed: Missing date for transaction ID {tx['id']}. Skipping.")
            continue
            
        if tx["monto_soles"] < 0:
            logging.warning(f"Validation: Egress amount was negative ({tx['monto_soles']}). Converting to positive.")
            tx["monto_soles"] = abs(tx["monto_soles"])
            tx["monto_original"] = abs(tx["monto_original"])
            
        # Budget Alerts (gasto > 10% del presupuesto de la categoria)
        if tx["tipo"] == "Egreso" and tx["categoria"] in limits:
            limit = limits[tx["categoria"]]
            try:
                tx_date = datetime.strptime(tx["fecha"], "%Y-%m-%d")
                tx_month = tx_date.strftime("%Y-%m")
            except Exception:
                tx_month = datetime.now(LIMA_TZ).strftime("%Y-%m")
                
            cat_sum = tx["monto_soles"]
            if not df_existing.empty:
                df_existing["fecha_dt"] = pd.to_datetime(df_existing["fecha"], errors="coerce")
                df_existing["month"] = df_existing["fecha_dt"].dt.strftime("%Y-%m")
                mask = (df_existing["categoria"] == tx["categoria"]) & (df_existing["month"] == tx_month) & (df_existing["tipo"] == "Egreso")
                cat_sum += df_existing.loc[mask, "monto_soles"].sum()
                
            if cat_sum > limit * 1.10:
                diff_pct = ((cat_sum - limit) / limit) * 100
                alert_msg = (
                    f"CRITICAL ALERT: Category '{tx['categoria']}' in {tx_month} "
                    f"accumulates S/ {cat_sum:.2f}, exceeding budget limit S/ {limit:.2f} by +{diff_pct:.1f}%!"
                )
                print(alert_msg)
                logging.warning(alert_msg)

        row_values = [
            tx["id"], tx["fecha"], tx["concepto"], tx["monto_original"], 
            tx["moneda"], tx["monto_soles"], tx["tipo"], tx["categoria"], tx["fuente"]
        ]
        ws_mov.append(row_values)
        existing_ids.add(str(tx["id"]))
        
        df_new = pd.DataFrame([{
            "fecha": tx["fecha"],
            "monto_soles": tx["monto_soles"],
            "tipo": tx["tipo"],
            "categoria": tx["categoria"]
        }])
        df_existing = pd.concat([df_existing, df_new], ignore_index=True)
        appended_count += 1

    wb.save(path)
    logging.info(f"Budget master spreadsheet updated. Appended {appended_count} new entries.")

# ============================================================
# 4. GENERACIÓN DE BALANCE FORENSE (FINANCIERO)
# ============================================================

def generate_financial_balance(master_path="presupuesto_maestro.xlsx"):
    """Reads master spreadsheet data and builds a comprehensive visual balance sheet report."""
    if not os.path.exists(master_path):
        logging.error(f"Cannot generate balance report. Master budget spreadsheet '{master_path}' does not exist. Initializing empty.")
        initialize_master_budget(master_path)

    try:
        df_mov = pd.read_excel(master_path, sheet_name="Movimientos")
        df_pres = pd.read_excel(master_path, sheet_name="Presupuestos")
    except Exception as e:
        logging.error(f"Failed to read master datasets: {e}")
        return

    date_str = datetime.now(LIMA_TZ).strftime("%Y%m%d")
    report_path = f"balance_financiero_{date_str}.xlsx"
    
    wb = openpyxl.Workbook()
    
    FONT_TITLE = Font(name='Arial', bold=True, size=14, color='FFFFFFFF')
    FONT_SECT  = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
    FONT_TEXT = Font(name='Arial', size=10, color='FF000000')
    FONT_TEXT_B = Font(name='Arial', bold=True, size=10, color='FF1F3864')
    FONT_TOTAL = Font(name='Arial', bold=True, size=11, color='FF1F3864')
    FONT_NETO  = Font(name='Arial', bold=True, size=12, color='FFFFFFFF')
    
    FILL_TITLE = PatternFill('solid', fgColor='FF1F3864')
    FILL_SECT  = PatternFill('solid', fgColor='FF2E5B98')
    FILL_TOTAL = PatternFill('solid', fgColor='FFD9E1F2')
    FILL_BRUTO = PatternFill('solid', fgColor='FFFFE699')
    FILL_NETO  = PatternFill('solid', fgColor='FF1F3864')
    FILL_ALERT = PatternFill('solid', fgColor='FFFFC7CE')
    
    ALIGN_L = Alignment(horizontal='left', vertical='center', indent=1)
    ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_R = Alignment(horizontal='right', vertical='center')
    
    thin = Side(style='thin', color='FFB4B4B4')
    med  = Side(style='medium', color='FF1F3864')
    BTB  = Border(top=med, bottom=med)
    BALL = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    FMT_NUM = '#,##0.00;(#,##0.00);"—"'

    def set_cell(ws, r, c, val, font=None, fill=None, align=None, fmt=None, border=None):
        cell = ws.cell(row=r, column=c, value=val)
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        if fmt: cell.number_format = fmt
        if border: cell.border = border
        return cell

    # ------------------------------------------------------------
    # HOJA 1: RESUMEN GENERAL (INCOME VS EGRESS)
    # ------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Resumen General"
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 40
    
    ws1.merge_cells('A1:C1')
    set_cell(ws1, 1, 1, 'BALANCE FINANCIERO GENERAL', FONT_TITLE, FILL_TITLE, ALIGN_C)
    ws1.row_dimensions[1].height = 32
    
    ws1.merge_cells('A2:C2')
    subtitle_text = f"Generado el: {datetime.now(LIMA_TZ).strftime('%Y-%m-%d %H:%M:%S')} (Lima Time)"
    set_cell(ws1, 2, 1, subtitle_text, Font(name='Arial', italic=True, size=9, color='FF707070'), None, ALIGN_C)
    
    total_ingresos = 0.0
    total_egresos = 0.0
    if not df_mov.empty:
        total_ingresos = df_mov[df_mov["Tipo"] == "Ingreso"]["Monto Soles"].sum()
        total_egresos = df_mov[df_mov["Tipo"] == "Egreso"]["Monto Soles"].sum()
    balance_disponible = total_ingresos - total_egresos
    
    set_cell(ws1, 4, 1, "RESUMEN DE SALDOS", FONT_SECT, FILL_SECT, ALIGN_L)
    ws1.merge_cells('A4:C4')
    ws1.row_dimensions[4].height = 22
    
    set_cell(ws1, 5, 1, "Total Ingresos Acumulados", FONT_TEXT, None, ALIGN_L)
    set_cell(ws1, 5, 2, total_ingresos, FONT_TEXT, None, ALIGN_R, FMT_NUM)
    
    set_cell(ws1, 6, 1, "Total Egresos Acumulados", FONT_TEXT, None, ALIGN_L)
    set_cell(ws1, 6, 2, total_egresos, FONT_TEXT, None, ALIGN_R, FMT_NUM)
    
    set_cell(ws1, 7, 1, "SALDO DISPONIBLE NETO", FONT_NETO, FILL_NETO, ALIGN_L, border=BTB)
    set_cell(ws1, 7, 2, balance_disponible, FONT_NETO, FILL_NETO, ALIGN_R, FMT_NUM, BTB)
    set_cell(ws1, 7, 3, "", None, FILL_NETO, None, None, BTB)
    ws1.row_dimensions[7].height = 24

    # ------------------------------------------------------------
    # HOJA 2: DETALLE CATEGORIAS (BUDGET VS REAL)
    # ------------------------------------------------------------
    ws2 = wb.create_sheet("Detalle Categorias")
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18

    ws2.merge_cells('A1:E1')
    set_cell(ws2, 1, 1, 'COMPARATIVA DE GASTOS POR CATEGORÍA', FONT_TITLE, FILL_TITLE, ALIGN_C)
    ws2.row_dimensions[1].height = 32
    
    headers_ws2 = ["Categoría", "Presupuestado (S/)", "Ejecutado (S/)", "Diferencia (S/)", "Estado / Alerta"]
    for idx, col in enumerate(headers_ws2, 1):
        set_cell(ws2, 3, idx, col, Font(name='Arial', bold=True, size=10, color='FFFFFFFF'), FILL_SECT, ALIGN_C, border=BALL)
    ws2.row_dimensions[3].height = 22

    cat_expenses = {}
    if not df_mov.empty:
        cat_expenses = df_mov[df_mov["Tipo"] == "Egreso"].groupby("Categoria")["Monto Soles"].sum().to_dict()
    
    row_idx = 4
    for idx, row in df_pres.iterrows():
        cat = row["Categoria"]
        budget = float(row["Presupuesto Limite"])
        real = float(cat_expenses.get(cat, 0.0))
        diff = budget - real
        
        if real > budget * 1.10:
            status = "EXCEDIDO (+10%)"
            fill_st = FILL_ALERT
        elif real > budget:
            status = "EXCEDIDO"
            fill_st = PatternFill('solid', fgColor='FFFFE699')
        else:
            status = "DENTRO"
            fill_st = None
            
        set_cell(ws2, row_idx, 1, cat, FONT_TEXT, None, ALIGN_L, border=BALL)
        set_cell(ws2, row_idx, 2, budget, FONT_TEXT, None, ALIGN_R, FMT_NUM, BALL)
        set_cell(ws2, row_idx, 3, real, FONT_TEXT, None, ALIGN_R, FMT_NUM, BALL)
        set_cell(ws2, row_idx, 4, diff, FONT_TEXT, None, ALIGN_R, FMT_NUM, BALL)
        set_cell(ws2, row_idx, 5, status, FONT_TEXT_B, fill_st, ALIGN_C, border=BALL)
        ws2.row_dimensions[row_idx].height = 20
        row_idx += 1

    set_cell(ws2, row_idx, 1, "TOTALES EGRESOS", FONT_TOTAL, FILL_TOTAL, ALIGN_L, border=BTB)
    set_cell(ws2, row_idx, 2, f"=SUM(B4:B{row_idx-1})", FONT_TOTAL, FILL_TOTAL, ALIGN_R, FMT_NUM, BTB)
    set_cell(ws2, row_idx, 3, f"=SUM(C4:C{row_idx-1})", FONT_TOTAL, FILL_TOTAL, ALIGN_R, FMT_NUM, BTB)
    set_cell(ws2, row_idx, 4, f"=SUM(D4:D{row_idx-1})", FONT_TOTAL, FILL_TOTAL, ALIGN_R, FMT_NUM, BTB)
    set_cell(ws2, row_idx, 5, "", None, FILL_TOTAL, None, None, BTB)
    ws2.row_dimensions[row_idx].height = 22

    # ------------------------------------------------------------
    # HOJA 3: HISTORIAL (ÚLTIMOS 30 DÍAS)
    # ------------------------------------------------------------
    ws3 = wb.create_sheet("Historial Reciente")
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 24
    ws3.column_dimensions['F'].width = 24
    
    ws3.merge_cells('A1:F1')
    set_cell(ws3, 1, 1, 'MOVIMIENTOS DE LOS ÚLTIMOS 30 DÍAS', FONT_TITLE, FILL_TITLE, ALIGN_C)
    ws3.row_dimensions[1].height = 32
    
    headers_ws3 = ["Fecha", "Concepto", "Monto Original", "Moneda", "Monto Soles", "Categoría"]
    for idx, col in enumerate(headers_ws3, 1):
        set_cell(ws3, 3, idx, col, Font(name='Arial', bold=True, size=10, color='FFFFFFFF'), FILL_SECT, ALIGN_C, border=BALL)
    ws3.row_dimensions[3].height = 22

    if not df_mov.empty:
        limit_date = (datetime.now(LIMA_TZ) - timedelta(days=30)).strftime("%Y-%m-%d")
        df_recent = df_mov[df_mov["Fecha"] >= limit_date].sort_values(by="Fecha", ascending=False)
        
        row_idx = 4
        for idx, row in df_recent.iterrows():
            set_cell(ws3, row_idx, 1, str(row["Fecha"])[:10], FONT_TEXT, None, ALIGN_C, border=BALL)
            set_cell(ws3, row_idx, 2, row["Concepto"], FONT_TEXT, None, ALIGN_L, border=BALL)
            set_cell(ws3, row_idx, 3, row["Monto Original"], FONT_TEXT, None, ALIGN_R, FMT_NUM, BALL)
            set_cell(ws3, row_idx, 4, row["Moneda"], FONT_TEXT, None, ALIGN_C, border=BALL)
            set_cell(ws3, row_idx, 5, row["Monto Soles"], FONT_TEXT_B, None, ALIGN_R, FMT_NUM, BALL)
            set_cell(ws3, row_idx, 6, row["Categoria"], FONT_TEXT, None, ALIGN_L, border=BALL)
            ws3.row_dimensions[row_idx].height = 18
            row_idx += 1

    # ------------------------------------------------------------
    # GENERAR GRÁFICO MATPLOTLIB & EMBED
    # ------------------------------------------------------------
    chart_filename = f"balance_chart_{date_str}.png"
    try:
        categories_list = list(cat_expenses.keys())
        expenses_list = list(cat_expenses.values())
        
        if categories_list:
            plt.figure(figsize=(7, 4.5))
            plt.barh(categories_list, expenses_list, color='#2E5B98', edgecolor='#1F3864')
            plt.xlabel('Gastado (Soles S/)')
            plt.title('Ejecución de Gastos por Categoría', fontsize=12, fontweight='bold', color='#1F3864')
            plt.tight_layout()
            plt.savefig(chart_filename, dpi=120)
            plt.close()
            
            # Embed chart into Sheet 1
            img = OpenpyxlImage(chart_filename)
            ws1.add_image(img, 'A9')
            logging.info(f"Chart generated and embedded as {chart_filename}")
        else:
            logging.info("No categories to plot.")
    except Exception as e:
        logging.error(f"Failed to generate matplotlib balance chart: {e}")

    ws1.freeze_panes = 'A5'
    ws2.freeze_panes = 'A4'
    ws3.freeze_panes = 'A4'
    
    wb.save(report_path)
    logging.info(f"Financial Balance Report generated at '{report_path}'.")

# ============================================================
# 5. CLI INTERFACE
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Agente Financiero Autónomo - CREAR Poder Sin Límites"
    )
    parser.add_argument(
        "--learn",
        nargs="?",
        const="./datos_financieros",
        help="Aprende la estructura de archivos en la ruta indicada (por defecto ./datos_financieros)"
    )
    parser.add_argument(
        "--update",
        action="store_true",
        help="Conecta a Gmail para buscar transacciones nuevas y actualizar el presupuesto maestro"
    )
    parser.add_argument(
        "--balance",
        action="store_true",
        help="Genera balance financiero consolidado mensual con reporte gráfico y por categorías"
    )

    args = parser.parse_args()

    if not (args.learn or args.update or args.balance):
        parser.print_help()
        return

    if args.learn:
        run_learning_engine(args.learn)

    if args.update:
        logging.info("Starting Gmail statement fetch and update cycle...")
        new_txs = fetch_gmail_transactions(days=7)
        if new_txs:
            update_budget_master(new_txs)
        else:
            logging.info("No new transactions found in the specified range.")

    if args.balance:
        logging.info("Starting financial balance sheet generation...")
        generate_financial_balance()

if __name__ == "__main__":
    main()
