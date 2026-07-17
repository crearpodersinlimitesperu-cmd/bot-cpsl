import os, json, logging, re
import openpyxl
from datetime import datetime, timedelta

log = logging.getLogger("Calendario")
logging.basicConfig(level=logging.INFO)

EXCEL_PATH = r"C:\Users\josem\OneDrive - QUANTUM COACHING TECHNOLOGY BVS CIA. LTDA\CREAR LIMA\PROGRAMACION 2026 CREAR LIMA.xlsx"
DATA_DIR = "/data" if os.path.exists("/data") else os.path.dirname(os.path.abspath(__file__))
CALENDARIO_PATH = os.path.join(DATA_DIR, "calendario_entrenamientos.json")

# Regex to match Excel relative date formulas referencing column A or B
formula_regex = re.compile(r"^=\+?([A-B])(\d+)(?:([\+-])(\d+))?$")

def sincronizar_calendario():
    if not os.path.exists(EXCEL_PATH):
        log.error(f"Archivo Excel no encontrado: {EXCEL_PATH}")
        return False
        
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
        if "LIM" not in wb.sheetnames:
            log.error("Pestaña 'LIM' no encontrada en el Excel")
            wb.close()
            return False
            
        ws = wb["LIM"]
        
        # Cache to store evaluated dates for cells: (row, col) -> datetime or int (year)
        eval_cache = {}
        
        def get_evaluated_cell(row, col):
            key = (row, col)
            if key in eval_cache:
                return eval_cache[key]
                
            cell = ws.cell(row=row, column=col)
            val = cell.value
            
            if val is None:
                eval_cache[key] = None
                return None
                
            if isinstance(val, datetime):
                eval_cache[key] = val
                return val
                
            if isinstance(val, int) and col == 1:
                # Year separators like 2025, 2026, 2027
                eval_cache[key] = val
                return val
                
            if isinstance(val, str) and val.startswith("="):
                m = formula_regex.match(val.strip().replace(" ", ""))
                if m:
                    ref_col_letter = m.group(1)
                    ref_row = int(m.group(2))
                    sign = m.group(3)
                    offset_val = int(m.group(4)) if m.group(4) else 0
                    
                    ref_col = 1 if ref_col_letter == "A" else 2
                    
                    ref_date = get_evaluated_cell(ref_row, ref_col)
                    if isinstance(ref_date, datetime):
                        if sign == "-":
                            res = ref_date - timedelta(days=offset_val)
                        else:
                            res = ref_date + timedelta(days=offset_val)
                        eval_cache[key] = res
                        return res
                    else:
                        eval_cache[key] = None
                        return None
                else:
                    eval_cache[key] = None
                    return None
                    
            eval_cache[key] = val
            return val

        # Evaluate all rows in LIM worksheet
        log.info("Evaluando fórmulas de fecha en la pestaña LIM...")
        for r in range(2, ws.max_row + 1):
            get_evaluated_cell(r, 1) # INICIO
            get_evaluated_cell(r, 2) # FINAL
            
        eventos = []
        for r in range(2, ws.max_row + 1):
            inicio_dt = eval_cache.get((r, 1))
            final_dt = eval_cache.get((r, 2))
            
            # Skip year separators
            if isinstance(inicio_dt, int):
                continue
                
            if not isinstance(inicio_dt, datetime) or not isinstance(final_dt, datetime):
                continue
                
            entrenamiento = ws.cell(row=r, column=3).value
            if entrenamiento:
                entrenamiento = str(entrenamiento).strip()
                
            if not entrenamiento or entrenamiento.upper() in ('NAN', ''):
                continue
                
            eventos.append({
                "inicio": inicio_dt.strftime("%Y-%m-%d"),
                "fin": final_dt.strftime("%Y-%m-%d"),
                "nombre": entrenamiento
            })
            
        wb.close()
        
        # Ordenar eventos cronológicamente por fecha de inicio
        eventos.sort(key=lambda x: x["inicio"])
        
        # Guardar en JSON
        with open(CALENDARIO_PATH, 'w', encoding='utf-8') as f:
            json.dump(eventos, f, ensure_ascii=False, indent=2)
            
        log.info(f"Se sincronizaron con éxito {len(eventos)} eventos para LIMA.")
        return True
        
    except Exception as e:
        log.error(f"Error procesando el Excel: {e}")
        return False

if __name__ == "__main__":
    sincronizar_calendario()
